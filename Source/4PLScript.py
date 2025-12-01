"""
DORA-Result Analysis Script

Summary:
This script is designed to analyze data produced by the DORAScripts. The primary functionalities include
selecting and loading data files, selecting target groups and the 4PL-analysis type, extracting and validating group data,
performing actual fitting analysis and graph building, while storing he results on disk as .png files.
The script starts by ensuring the necessary libraries are installed and then initializes global attributes related to
the experiment setup. It includes user interaction functions for input validation and output formatting, ensuring
a smooth user experience.
Interruptions are handled gracefully, and the script provides clear progress and error messages throughout its execution.

Usage:
1. Ensure you have the required directories (`config_files`, `data_files and `result_files`) with appropriate files.
2. Run the script from the command line.
3. Follow the prompts to:
   - Select a data file.
   - Choose a data sheet from the data file.
   - Select the groups.
   - Select the analysis type.
4. The script will then:
   - Extract group data.
   - Perform the selected analysis.
   - Write the analysis results into the selected excel file.
   - Fit the draw the results into graphs.
   - Zip and store the graphs as .png files.
5. The updated data file will be saved with the new data.
"""




########## IMPORTS ##########


import os
import time
import textwrap
import warnings
import json
from json import JSONDecodeError
from sys import exit, argv as args, maxsize, float_info
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED
try:
    from xlrd import open_workbook                                  # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.\n')
try:
    from openpyxl import load_workbook as load                      # type: ignore
    from openpyxl.styles import Font, Alignment                     # type: ignore
    from openpyxl.workbook import Workbook                          # type: ignore
    from openpyxl.utils import get_column_letter                    # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xlsx files called "openpyxl". Run the installer to update your libraries.\n')
try:
    import numpy                                                    # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library for numerical operations called "numpy". Run the installer to update your libraries.\n')
try:
    import matplotlib.pyplot as pyplot                              # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library for plotting called "matplotlib". Run the installer to update your libraries.\n')
try:
    from scipy.optimize import OptimizeWarning, curve_fit as fit    # type: ignore
    from scipy.stats import t                                       # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library for curve fitting called "scipy". Run the installer to update your libraries.\n')




########## ATTRIBUTES ##########

SCRIPT_VERSION = "0.1.0"                    # The public version number of the script

action_iterator = 1                         # A global iterator used for tracking actions
DELAY = 0.05                                # The delay (in seconds) used for timing outputs and creating a user-friendly experience
ENVIRONMENT = "SCR"                         # The environment mode for the script, typically set to 'SCR' (short for scripted)
CUSTOM_TYPE = "Custom"                      # Custom analysis type key
ANALYSIS_TYPES = {}                         # Analysis types (keys) and their hill slope standards (values) that are read from the confg file
CONFIG_PARAMETER = "analysis_types"         # Parameter that is configured in the config file
CONFIG_FILE_NAME = "4PL_config"             # The name of the script's configure file
CONFIG_FILE_TYPE = ".json"                  # The type of the script's configure file
CONFIG_FILES_DIRECTORY = "./config_files/"  # The directory within which the script looks for its configuration (.json) files
DATA_FILES_DIRECTORY = "./data_files/"      # The directory within which the script looks for potential excel (.xls and.xlsx) data files
RESULT_FILES_DIRECTORY = "./result_files/"  # The directory within which the script places its result files
VALUES_IN_A_SAMPLE = 9                      # The number of values in a filled sample
ALPHA = 0.05                                # The maximum allowed probability of a false positive




########## CLASSES ##########


class Analysis:
    """Class to manage Analysis information."""

    def __init__(self, type, groups):
        self.type = type
        self.groups = groups

    def __str__(self):
        """String representation of the Analysis object"""
        return f"{self.type} analysis."
    
    def form_data(self):
        """Builds nested dict: group -> sample -> x,y arrays."""
        data = {}
        for group in self.groups:
            key_1 = group.title
            value_1 = {}
            data[key_1] = value_1
            for sample in group.samples:
                key_2 = sample.label
                value_2 = {}
                value_1[key_2] = value_2
                x = numpy.array(group.dils[:len(sample.values)])
                y = numpy.array(sample.values)
                value_2["x"] = x
                value_2["y"] = y
        return data
    
    def variable_slope_model(self, x, log_result, hill_slope):
        """4PL-like variable slope model."""
        return 100 / (1 + 10**((log_result - x) * hill_slope))
    
    def fit_group(self, group_data):
        """Fits each sample in group_data to the 4PL model."""
        if self.type not in ANALYSIS_TYPES:
            exit_by_error(f'Set analysis type "{self.type}" was not found in the configuration.')
        hill_slope_standard = ANALYSIS_TYPES[self.type]
        results = {}
        for sample_name, sample_data in group_data.items():
            x_data = sample_data['x']
            y_data = sample_data['y']
            # Initial guess for parameters
            initial_guess = [numpy.median(x_data), hill_slope_standard]
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", OptimizeWarning)
                params, cov = fit(self.variable_slope_model, x_data, y_data, p0=initial_guess, maxfev=10000) # In case errors about this are thrown, try adjusting the number by increasing it.
            # Compute confidence intervals
            perr = numpy.sqrt(numpy.diag(cov))
            dof = max(0, len(y_data) - len(params))
            tval = t.ppf(1 - ALPHA/2, dof)
            ci = [(param - perr[i]*tval, param + perr[i]*tval) for i, param in enumerate(params)]
            log_result = params[0]
            hill_slope = params[1]
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", RuntimeWarning)
                result = 10 ** log_result
            # Generate fit curve
            x_fit = numpy.linspace(min(x_data), max(x_data), 100)
            y_fit = self.variable_slope_model(x_fit, *params)
            # Gather and return the results
            results[sample_name] = {
                'params': params,
                'ci': ci,
                'result': result,
                'log_result': log_result,
                'hill_slope': hill_slope,
                'x_data': x_data,
                'y_data': y_data,
                'x_fit': x_fit,
                'y_fit': y_fit
            }
        return results

    def run(self):
        """Performs run the analysis by fitting its group data."""
        data = self.form_data()
        analysis_results = {}
        for group_title, group_data in data.items():
            analysis_results[group_title] = self.fit_group(group_data)
        return analysis_results


class Group:
    """Class to manage group information, including dilutions and samples."""

    def __init__(self, title, cell):
        """Initializes an Group object with a title."""
        self.title = title
        self.cell = cell
        self.dils = []
        self.samples = []

    def __str__(self):
        """String representation of the Group object."""
        return self.title

    def add_sample(self, sample):
        """Adds a sample to the group."""
        self.samples.append(sample)

    def add_dilution(self, dil):
        """Adds a dilution to the group."""
        self.dils.append(dil)


class Sample:
    """Class to store and manage its information."""

    def __init__(self, label):
        """Initializes a Sample object with a title and an empty list for values."""
        self.label = label
        self.values = []

    def __str__(self):
        """Returns a string representation of the Sample object."""
        return f'Sample {self.label}'
    
    def is_filled(self):
        """Returns True if all required values are set."""
        return len(self.values) == VALUES_IN_A_SAMPLE

    def add_value(self, value):
        """Adds an absorbance value."""
        self.values.append(value)
    
    def get_values(self):
        """Retrieves a list of absorbance values."""
        return self.values




########## ARGUMENT HANDLING ##########


def info():
    """Function to print out instructions on how to use the script for the user."""
    try:
        # Print tutorial started status
        print_status("Tutorial Started", tailing_line_break=True)
        # Print each section of the tutorial with detailed instructions
        print_info("1) Evaluating Environment:", [
            "The script begins by checking the structure of the required directories.",
            "It checks for the existence of 'config_files', 'data_files' and 'result_files' directories.",
            "It ensures that each directory contains valid files (like .json, .xls and .xlsx)."
        ])
        print_info("2) Configuring Script:", [
            "Next the script will look for its configuration file '4PL_config.json' from the 'config_files' directory.",
            "It will load and apply its parameters into the script.",
            "It ensures that the config file can be found and its parameters are valid."
        ])
        print_info("3) Selecting Data File:", [
            "The user will then be prompted to select a data file from the 'data_files' directory.",
            "The available excel files will be listed, and the user will again choose one by its corresponding number."
        ])
        print_info("4) Selecting Data Sheet:", [
            "The user will select a sheet from the chosen data file.",
            "The available sheets will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("5) Selecting Data Groups:", [
            "The user will select the to used groups.",
            'The user may select "SELECT ALL" to pick all available groups.',
        ])
        print_info("6) Selecting Analysis Type:", [
            "The user will select the 4PL analysis type.",
            f"Should the user select the {CUSTOM_TYPE} type, then they will also be asked to provide a custom name and hill slope standard for the analysis."
        ])
        print_info("7) Extracting Analysis Data:", [
            "The script will then read the selected data sheet and extract the data of the selected groups."
        ])
        print_info("8) Running Analysis:", [
            "The script will run the selected analysis and calculate the result values for each selected group using a fitting function.",
            "The script will then write the analysis results into the data file.",
        ])
        print_info("9) Plotting Graphs:", [
            "The script will plot a graph for group based on their analysis results.",
            "The graphs are stored as .png files. and named as \"Graph_[groupname]_[timestamp]_(original_file).png\" to make them distinct and recognizable.",
            "The files are zipped into a .zip container named \"Graphs_[analysistype]_[timestamp]_([origin_file]).zip\".",
        ])
        print_info("10) Saving Data File:", [
            "Finally the script will save the updated data file with the new results sheets to ensure that the intermediate results are preserved."
        ])
        
        # Print additional notes
        print_info("Notes:", [
            "The script may be interrupted at any point of its execution safely by hitting (CTRL+C).",
            "Should the config file be lost, the script will be able to restore its template as long as the config directory exists.",
            "Should the user pick a .xls file as data file, the script will automatically conver it into a .xlsx file and replace the old .xls file altogether.",
            "The script looks for config files within the './config_files/' directory.",
            "The script looks for data files within the './data_files/' directory.",
            "The script places generated files into the './result_files/' directory.",
            'For the script to recognize data groups, they should be labeled as "Group..." in the excel file.'
        ])
        # Print explanation for each type of user prompt that the script uses
        print_info("User Prompts:", [
            "Input Prompt: When the script requires user input, it will display a prompt with a '>' symbol (e.g., \"Please select the data file > \"). The user will enter their response after the '>' symbol.",
            "Choices: Whenever the script presents multiple options to choose from, they will be listed with a corresponding command in square brackets (e.g., \"[1] data_file_1.xlsx\"). The user will select an option by typing its command.",
            "Task: The script will print task status messages to indicate what it is currently doing. These messages will start with \"> Task: \" followed by the task description.",
            "Progress: During longer operations, the script will print progress messages to keep the user informed. These messages will start with \"> Progress: \" followed by the progress description.",
            "Success: Upon successful completion of an operation, the script will print success messages starting with \"> Success: \" followed by a brief description of the success.",
            "Error: If an error occurs during an operation, the script will print error messages starting with \"> Error: \" followed by the error description."
        ])
        print_status("Tutorial completed", leading_line_break=False)
    except KeyboardInterrupt:
        exit_by_interruption(True)


def handle_arguments(args):
    """Handles the command line arguments provided to the script."""
    # Check if any arguments were given at launch
    if 1 < len(args):
        param = args[1].lower()
        if param in ["i", "info", "h", "help"]:
            # If the first argument was either "info" or "help", start tutorial
            info()
            exit()




########## EXIT ROUTES ##########


def exit_by_error(message=None):
    """Exits the script with a given error message."""
    try:
        if message:
            time.sleep(DELAY)
            print_error(message)
        time.sleep(DELAY)
        print_status("Failed")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")


def exit_by_interruption(during_info=False):
    """Exits the script due to an interruption (e.g., KeyboardInterrupt, CTRL+C)."""
    try:
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        if during_info:
            print_status("Tutorial stopped")    
        else:
            print_status("Stopped")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")




########## USER INTERACTION ##########


def prompt_input(prompt_str):
    """Gets user input after displaying a prompt."""
    try:
        time.sleep(DELAY)
        user_input = input(f"{prompt_str} > ").strip()
    except KeyboardInterrupt:
        exit_by_interruption()
    return user_input


def get_user_input(prompt, validation_func=None, type_select=True):
    """Gets and validates user input.

    Parameters:
    prompt (str): The prompt message to display.
    validation_func (function): The function to validate user input. Default is None.
    type_select (bool): Determines if input is for selection (True) or entry (False). Default is True.

    Returns:
    str: The validated user input.
    """
    # Print a blank line if type_select is True
    if type_select:
        print()
    while True:
        # Get user input
        user_input = prompt_input(prompt)
        # If no validation function is provided, return the input
        if not validation_func:
            if type_select:
                print()
            return user_input
        # Validate the user input
        validation_result, validated_input = validation_func(user_input)
        # If validation is successful, return the validated input
        if validation_result is True:
            if type_select:
                print()
            return validated_input
        # Print error message if validation fails
        print_error(validated_input)




########## PRINT FUNCTIONS ##########


def print_status(status_string, leading_line_break=True, tailing_line_break=False):
    """Prints a status message with version in the top border."""
    try:
        time.sleep(DELAY)
        text = f"4PL SCRIPT {status_string.upper()}"
        # Middle line padding
        asterisks = "*" * 3
        spaces = " " * 5
        middle_line = f"{asterisks}{spaces}{text}{spaces}{asterisks}"
        # Top border with version
        top_border_text = f" v{SCRIPT_VERSION} "
        total_width = len(middle_line)
        top_line = "*" * ((total_width - len(top_border_text)) // 2) + top_border_text
        if len(top_line) < total_width:
            top_line += "*" * (total_width - len(top_line))
        bottom_line = "*" * len(middle_line)

        if leading_line_break:
            time.sleep(DELAY)
            print()
        time.sleep(DELAY)
        print(top_line)
        time.sleep(DELAY)
        print(middle_line)
        time.sleep(DELAY)
        print(bottom_line)
        if tailing_line_break or status_string.strip().lower() == "completed":
            time.sleep(DELAY)
            print()
    except KeyboardInterrupt:
        exit_by_interruption()


def print_action(action_str):
    """Prints a separator line for actions."""
    try:
        global action_iterator
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        print(f"========== Action {action_iterator}: {action_str.title()} ==========")
        action_iterator += 1
    except KeyboardInterrupt:
        exit_by_interruption()


def print_task(task_str):
    """Prints a task message."""
    try:
        time.sleep(DELAY)
        print(f"> Task: {task_str}..")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_choice(choice_cmd, choice_str):
    """Prints a choice option."""
    try:
        time.sleep(DELAY)
        print(f"[{choice_cmd}] {choice_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_progress(progress_str):
    """Prints a progress message."""
    try:
        time.sleep(DELAY)
        print(f"> Progress: {progress_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_success(success_str):
    """Prints a success message."""
    try:
        time.sleep(DELAY)
        print(f"> Success: {success_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_error(error_str):
    """Prints an error message."""
    try:
        time.sleep(DELAY)
        print(f"> Error: {error_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_info(header, messages, max_length=100):
    """Prints information with a delay and text wrapping."""
    try:
        INDENT = 3
        time.sleep(DELAY)
        print(header.title())
        for message in messages:
            wrapped_message = textwrap.fill(f" - {message}", max_length, subsequent_indent=' ' * INDENT)
            time.sleep(DELAY)
            print(wrapped_message)
        time.sleep(DELAY)
        print()
    except KeyboardInterrupt:
        exit_by_interruption(True)




########## VALIDATION FUNCTIONS ##########


def validate_digit(input_str):
    """Validate if the input string is a digit.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the digit value if valid, or an error message if not.
    """
    if input_str.isdigit():
        return True, int(input_str)
    return False, f'Your input of "{input_str}" was not a positive whole number.'


def validate_non_occupied_digit(input_str, occupied):
    """Validate if the input string is an digit that isn't already occupied.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if validation_result not in occupied:
        return True, validation_result
    return False, f'Your input "{input_str}" was already taken by another defined group.'


def validate_integer(input_str):
    """Validate if the input string is an integer.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    try:
        return True, int(input_str)
    except ValueError:
        return False, f'Your input of "{input_str}" was not a whole number.'


def validate_float(input_str):
    """Validate if the input string is a float, allowing for both Finnish (,) and English (.) decimal separators.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the float value if valid, or an error message if not.
    """
    try:
        input_str = str(input_str).replace(',', '.')
        float_value = float(input_str)
        return True, float_value
    except ValueError:
        return False, f'Your input "{input_str}" was not a number.'


def validate_non_empty(input_str):
    """Validate if the input string is not empty.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the input string if valid, or an error message if not.
    """
    if input_str.strip():
        return True, input_str
    return False, f"Your input was empty."


def validate_min_max(input_str, min_choice, max_choice, escape_str=None):
    """Validate if the input string is a digit within a specified range, with an optional escape string.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer or 0 if escaped, or an error message if not.
    """
    if escape_str is not None and input_str.strip() == escape_str:
        return True, 0
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if min_choice <= validation_result <= max_choice:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [{min_choice}, {max_choice}].'


def validate_yes_no(input_str):
    """Validate if the input string is a 'yes' or 'no' response.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated input string if valid, or an error message if not.
    """
    input_str = input_str.lower().strip()
    if input_str in ['y', 'yes', 'n', 'no']:
        return True, input_str
    return False, f'Your input "{input_str}" was invalid. Choose either yes or no.'


def validate_group_title(value):
    """Validate if the value contains an appropriate group title like 'Group 2'."""
    if isinstance(value, str) and value.casefold().startswith("group"):
        return True
    return False


def validate_log_string(value):
    """Validate if the input matches the string of 'log dil' case insensitively."""
    if isinstance(value, str) and "log dil" in value.casefold():
        return True
    return False


def validate_dilution(value):
    """Validate if the value is a dilution log value."""
    is_valid, validation_result = validate_float(value)
    if is_valid:
        return True
    return False


def validate_absorbance(value):
    """Validate if the value is a valid absorbance value."""
    is_valid, validation_result = validate_float(value)
    if is_valid:
        return True
    return False


def validate_sample(sheet, first_row, col):
    """Check if the col contains enough valid float values going down from the first value cell of the sample."""
    for row in range(first_row, first_row + VALUES_IN_A_SAMPLE):
        cell_value = sheet.cell(row=row, column=col).value
        if not validate_absorbance(cell_value):
            return False
    return True




########## SEARCH FUNCTIONS ##########


def find_group(sheet, start_row):
    """Finds the start, end, and next row of a group by scanning the entire sheet.
    
    Returns:
    dict: {'title', 'row', 'col', 'next_row'} or None if not found.
    """
    START_COL = 1
    END_COL = 5
    # Iterate through the entire sheet, checking each row's first 5 cols for a valid group start.
    for row in range(start_row, sheet.max_row + 1):
        for col in range(START_COL, END_COL + 1):
            group_cell_value = sheet.cell(row=row, column=col).value
            if not validate_group_title(group_cell_value):
                continue
            log_cell_value = sheet.cell(row=row+1, column=col).value
            if not validate_log_string(log_cell_value):
                continue
            log_val_cell_value = sheet.cell(row=row+2, column=col).value
            if not validate_dilution(log_val_cell_value):
                continue
            if not validate_sample(sheet, row+2, col+1):
                continue
            # Group was found, returning its details
            return {
                "title": group_cell_value,
                "cell": (row, col),
                "next_row": row + VALUES_IN_A_SAMPLE + 3
            }
    return None # No groups were found




########## FILE CONVERTER ##########


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Function to convert deprecated .xls files into .xlsx files."""
    # Open the .xls file using xlrd
    workbook_xls = open_workbook(xls_path)
    workbook_xlsx = Workbook()
    for sheet_index in range(workbook_xls.nsheets):
        # One sheet at a time
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        if sheet_index == 0:
            # Activate the workbook to create its first sheet
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
        else:
            # Add them otherwise
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
        # One row at a time
        for row_index in range(sheet_xls.nrows):
            for col_index in range(sheet_xls.ncols):
                # Write each cell
                sheet_xlsx.cell(row=row_index + 1, column=col_index + 1).value = sheet_xls.cell_value(row_index, col_index)
    try:
        # Save the new .xlsx file using openpyxl
        workbook_xlsx.save(xlsx_path)
        # Delete the old .xls file using xlrds
        os.remove(xls_path)
    except PermissionError:
        # If couldn't remove the original file, remove the new one to avoid duplicates
        os.remove(xlsx_path)
        exit_by_error(f'File "{xls_path}" was already open in another program. Please close it and try again.')




########## CORE FUNCTIONS ##########


def directory_contains_file(directory, filename):
    """Check if a specific file exists in the given directory."""
    return filename in os.listdir(directory)


def get_excel_files(directory):
    """Function to find all excel files (.xls and .xlsx) in the given directory that do not start with '~$'."""
    return [file for file in os.listdir(directory) if (file.endswith('.xlsx') or file.endswith('.xls')) and not file.startswith('~$')]


def generate_mock_template():
    """Function to create mock analysis types instance as a dictionary."""
    return {
        "Analysis type 1": -1,
        "Analysis type 2": 0,
        "Analysis type 3": 1
    }


def create_config_template():
    """Function to create a template configuration file."""
    # Ensuring the config directory still exists.
    os.makedirs(CONFIG_FILES_DIRECTORY, exist_ok=True)
    print_task(f'Creating config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" in config directory.')
    # Determine the configuration file path
    config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    # Create the configuration file
    with open(config_file_path, "w", encoding="utf-8") as file:
        # Add template content
        mock_analysis_types = generate_mock_template()
        json.dump({CONFIG_PARAMETER: mock_analysis_types}, file, indent=4)
    print_success("Config file created, terminating script.")


def evaluate_script_environment():
    """Check if the required directories with acceptable content exist in the current working directory."""
    # Check for the existence of config files directory
    if not os.path.isdir(CONFIG_FILES_DIRECTORY):
        exit_by_error(f'Config files directory "{CONFIG_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Config files directory "{CONFIG_FILES_DIRECTORY}" located.')
    # Check if config files directory contains the scirpt's config file
    config_file = directory_contains_file(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    if not config_file:
        # In case the config file cannot be located
        print_error(f'Config files directory didn\'t contain file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
        # Ask the user if the missing config file should be restored
        yes_no_value = get_user_input('The script will be terminated...\nRestore the missing config file template on exit? [y/n]', validate_yes_no)
        if yes_no_value in ['y', 'yes']:
            # Restore the config file template
            create_config_template()
            # Stop the script
            exit_by_error()
        else:
            # Terminate the script
            exit_by_error("Terminating script.")
    print_progress(f'Config files directory contained file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    # Check for the existence of data files directory
    if not os.path.isdir(DATA_FILES_DIRECTORY):
        exit_by_error(f'Data files directory "{DATA_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Data files directory "{DATA_FILES_DIRECTORY}" located.')
    # Check if data files directory contains at least one excel file that doesn't start with "~$"
    data_files = get_excel_files(DATA_FILES_DIRECTORY)
    if not data_files:
        exit_by_error(f'Data files directory did not contain any excel files.')
    print_progress(f'Data files directory contained excel files.')
    # Check for the existence of result files directory
    if not os.path.isdir(RESULT_FILES_DIRECTORY):
        exit_by_error(f'Result files directory "{RESULT_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Result files directory "{RESULT_FILES_DIRECTORY}" located.')


def load_configuration():
    """Attempt to Load configration from the configuration file."""
    try:
        # Determine the configuration file path
        config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
        with open(config_file_path, "r", encoding="utf-8") as file:
            # Load from the configuration file
            config = json.load(file)
            # Return the loaded configuration
            return config
    except FileNotFoundError:
        # Configuration file could not be found
        exit_by_error(f'Config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be found.')
    except JSONDecodeError:
        # Configuration could not be loaded
        exit_by_error(f'Configuration from file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be loaded.')
    

def apply_configuration(config):
    """Check that the configuration is valid and apply it."""
    # Fetch "analysis types" from configuration
    analysis_types = config.get(CONFIG_PARAMETER, None)
    # Check that configuration is valid
    if analysis_types is None:
        # Configuration didn't contain "analysis types"
        exit_by_error(f'Loaded configuration is missing the "{CONFIG_PARAMETER}" property. Please add it.')
    if not isinstance(analysis_types, dict):
        # Aanalysis types wasn't a dict
        exit_by_error(f'The "{CONFIG_PARAMETER}" property must be a dictionary, but it is currently of type {type(analysis_types).__name__}.')
    if len(analysis_types) < 1:
        # Analysis types was empty
        exit_by_error(f'The "{CONFIG_PARAMETER}" was empty. Please add at least one key-value pair.')
    # Validate each key-value pair in the dictionary
    for key, value in analysis_types.items():
        if not isinstance(key, str):
            exit_by_error(f'Each key in "{CONFIG_PARAMETER}" must be a string. Key "{key}" was of type {type(key).__name__}.')
        if not key.strip():
            exit_by_error(f'Each key in "{CONFIG_PARAMETER}" must be a non-empty string. Found a key that was empty or white space.')
        if not isinstance(value, int):
            exit_by_error(f'Each value in "{CONFIG_PARAMETER}" must be an integer. Value for "{key}" is of type {type(value).__name__}.')
    # Store validated config dictionary
    ANALYSIS_TYPES.update(analysis_types)
    ANALYSIS_TYPES.update({CUSTOM_TYPE: None})


def configure_script():
    """Initialize the script by loading and applying a configuration."""
    # Load the configuration
    config = load_configuration()
    print_progress("Configuration loaded.")
    # Apply the configuration
    apply_configuration(config)
    print_progress("Configuration applied.")


def select_excel_file(directory):
    """Selects an Excel file from the specified directory."""
    # List all Excel files in the directory, excluding temporary files
    try:
        excel_files = get_excel_files(directory)
    except FileNotFoundError:
        exit_by_error(f'Directory "{directory}" could not be located.')
    # If no Excel files are found, exit the script with an error message
    if not excel_files:
        exit_by_error(f'Directory "{directory}" did not contain any excel files.')
    # Print choices for the user to select from
    for i, file in enumerate(excel_files):
        print_choice(i+1, file)
    # User chooses a file
    int_value = get_user_input(f'Please select the file', lambda input_str: validate_min_max(input_str, 1, len(excel_files)))
    return excel_files[int_value - 1]


def load_excel_file(directory, file_name):
    """Load an Excel file from the given path and return the workbook object."""
    try:
        # Convert .xls to .xlsx if necessary
        if file_name.endswith('.xls'):
            xls_path = directory + file_name
            file_name += 'x'
            xlsx_path = directory + file_name
            convert_xls_to_xlsx(xls_path, xlsx_path)
            print_progress("Converted the old .xls file into a new .xlsx file.")
        else:
            xlsx_path = directory + file_name
        # Load the Excel file
        file = load(xlsx_path)
        # Save the unchanged file to check if it is accessible
        file.save(xlsx_path)
        return file, file_name
    except FileNotFoundError:
        # Handle file not found error
        exit_by_error(f'File "{file_name}" could not be located.')
    except PermissionError:
        # Handle permission error if the file is already open
        exit_by_error(f'File "{file_name}" was already open in another program. Please close it and try again.')
    except Exception as e:
        # Handle any other exceptions that occur during file loading
        exit_by_error(f'Failed loading file "{file_name}".  > REASON: {e}.')


def select_datasheet(data_file):
    """Prompt the user to select a data sheet from the given data file."""
    # List all sheet names in the data file
    datasheets = [sheet for sheet in data_file.sheetnames]
    for i, sheet in enumerate(datasheets):
        print_choice(i+1, sheet)
    # User chooses a sheet
    int_value = get_user_input('Please select the sheet', lambda input_str: validate_min_max(input_str, 1, len(data_file.sheetnames)))
    return datasheets[int_value - 1]


def load_datasheet(data_file, data_sheet_name):
    """Load the specified data sheet from the given data file."""
    return data_file[data_sheet_name]


def get_available_groups(data_sheet):
    """Extracts groups from the specified data sheet."""
    next_row = 1
    groups = []
    while True:
        # Find a group in the sheet
        group_info = find_group(data_sheet, next_row)
        # If no group was found then the sheet contains no more plates
        if not group_info:
            break
        # Create and add a new Group object with the gathered data
        group = Group(group_info["title"], group_info["cell"])
        groups.append(group)
        # Update the row to read onwards from
        next_row = group_info["next_row"]
    # Exit with an error in case no wellplates were found on the sheet
    if len(groups) < 1:
        exit_by_error(f"Could not locate any groups in data sheet '{data_sheet.title}'.")
    # Return the groups
    return groups


def select_groups(data_sheet):
    """Select the group by asking the user for a group number."""
    SELECT_ALL_STRING = "SELECT ALL"
    available_groups = get_available_groups(data_sheet)
    available_groups.append("SELECT ALL")
    selected_groups = []
    while True:
        for i, group in enumerate(available_groups):
            print_choice(i+1, group)
        # User chooses a group
        int_value = get_user_input('Please select a group', lambda input_str: validate_min_max(input_str, 1, len(available_groups)))
        if int_value == len(available_groups):
            # Remove select all option
            available_groups.pop(len(available_groups) - 1)
            # Move all groups from available to selected
            selected_groups.extend(available_groups)
            available_groups.clear()
            break
        else:
            # Move the selected group from available to selected
            group = available_groups.pop(int_value - 1)
            selected_groups.append(group)
        # If the available groups have all been spent
        if not available_groups or available_groups == [SELECT_ALL_STRING]:
            break
        # Ask if the user wants to select another group
        yes_no_value = get_user_input('Would you like to select another group? [y/n]', validate_yes_no, False)
        if yes_no_value in ['n', 'no']:
            break
    return selected_groups


def select_analysis_type():
    """Prompt user to select analysis type, and set value for 'Custom' if chosen."""
    type_names = list(ANALYSIS_TYPES.keys())
    # Display available analysis types for selection
    for i, type_name in enumerate(type_names):
        print_choice(i+1, type_name)
    # Get user input for the selected analysis type
    user_input = get_user_input("Please select the analysis type", lambda input_str: validate_min_max(input_str, 1, len(type_names)))
    selected_type = type_names[int(user_input) - 1]
    # If user selected "Custom", ask for the hill slope standard value
    if selected_type == CUSTOM_TYPE:
        selected_type = get_user_input("Please provide a custom analysis type", validate_non_empty, False)
        hill_slope_standard = get_user_input("Please provide a custom hill slope standard", validate_float, False)
        ANALYSIS_TYPES[selected_type] = hill_slope_standard
    return selected_type


def extract_segment(data_sheet, start_row, start_col, end_row, end_col):
    """Extracts a specified segment of data from a datasheet."""
    # Determine the coordinates of the first and last cell of the segment
    first_cell = data_sheet.cell(row=start_row, column=start_col).coordinate
    last_cell = data_sheet.cell(row=end_row, column=end_col).coordinate
    # Extract the segment
    segment = data_sheet[first_cell:last_cell]
    return segment


def add_data_to_group(group_data_segment, group):
    """Adds data to the existing group object."""
    # Extract the dilutions
    for row in group_data_segment[2:]:
        log_dil = row[0].value
        # Validate the log value
        if not validate_dilution(log_dil):
            exit_by_error(f'Dilution value {log_dil} was not valid.')    
        # Add the valid log value to the group
        group.add_dilution(log_dil)
    # Extract the samples
    for i, row in enumerate(group_data_segment[1:], 1):
        for j, cell in enumerate(row[1:], 1):
            if i == 1: # On label row
                # Create new sample
                title = cell.value
                sample = Sample(title)
                group.add_sample(sample)
            else: # On value rows
                # Validate and add the absorbance value
                absorbance = cell.value
                if not validate_absorbance(absorbance):
                    exit_by_error(f'Absorbance value {absorbance} was not valid.')
                group.samples[j - 1].add_value(absorbance)


def extract_analysis_data(data_sheet, groups, analysis_type):
    """Extracts wellplate data from the specified data sheet."""
    analysis = Analysis(analysis_type, groups)
    for group in analysis.groups:
        # Get the sheet matrix information for the group
        START_ROW, START_COL = group.cell
        END_ROW = START_ROW + VALUES_IN_A_SAMPLE + 1
        end_col = START_COL
        while True:
            if not validate_sample(data_sheet, START_ROW+2, end_col+1):
                break
            end_col += 1
        print_progress(f"{group.title} located in segment [{chr(64+START_COL)}{START_ROW}, {chr(64+end_col)}{END_ROW}].")
        # Extract the group segment based on the starting point
        group_data_segment = extract_segment(data_sheet, START_ROW, START_COL, END_ROW, end_col)
        # Add the group segment data to the group
        add_data_to_group(group_data_segment, group)
        samples = group.samples
        for sample in group.samples:
            if not sample.is_filled():
                exit_by_error(f"There weren't enough valid values to complement sample {sample.label} of group {group.title}.")
        print_progress(f"{group.title} extracted with {len(samples)} samples.")
    return analysis


def create_or_clear_sheet(xlsx_file, sheet_name, position=None):
    """Create a new sheet or clear it if it already exists."""
    # Check if the sheet already exists
    if sheet_name in xlsx_file.sheetnames:
        # Delete the existing sheet
        print_progress(f'Sheet "{sheet_name}" data wiped.')
        del xlsx_file[sheet_name]
    else:
        print_progress(f'New sheet "{sheet_name}" created.')
    # Create and return a new sheet
    sheet = xlsx_file.create_sheet(sheet_name)
    if position:
        sheets = xlsx_file._sheets
        sheets.insert(position, sheets.pop(sheets.index(sheet)))
    return sheet


def get_col_width(sheet, col_char, padding=2):
    """Calculate the width of a column based on the maximum length of the data in that column plus padding."""
    DEFAULT_WIDTH = 8.43
    lengths = [len(str(cell.value)) for cell in sheet[col_char] if cell.value is not None]
    if not lengths:
        return DEFAULT_WIDTH
    return max(lengths) + padding


def autosize_cols(sheet, padding=2):
    """Auto-size all columns in the given Excel worksheet using calculated content width plus padding."""
    for col_cells in sheet.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = get_col_width(sheet, col_letter, padding=padding)
        sheet.column_dimensions[col_letter].width = width


def write_result_data(results_sheet, analysis_type, group_title, group_results, write_row, write_col):
    """Function to write the group's result data to the results sheet."""
    # Write group label
    label_width = len(group_results)
    results_sheet.merge_cells(start_row=write_row, start_column=write_col, end_row=write_row, end_column=write_col+label_width)
    results_sheet.cell(row=write_row, column=write_col).value = group_title
    results_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write sample labels header
    results_sheet.cell(row=write_row + 1, column=write_col).value = "Sample"
    results_sheet.cell(row=write_row + 1, column=write_col).font = Font(bold=True)
    # Write sample labels
    for i, sample_label in enumerate(group_results.keys(), 1):
        results_sheet.cell(row=write_row + 1, column=write_col + i).value = sample_label
        results_sheet.cell(row=write_row + 1, column=write_col + i).alignment = Alignment(horizontal='right')
    # Write result values header
    results_sheet.cell(row=write_row + 2, column=write_col).value = analysis_type
    results_sheet.cell(row=write_row + 2, column=write_col).font = Font(bold=True)
    # Write log result values header
    results_sheet.cell(row=write_row + 3, column=write_col).value = f'Log{analysis_type}'
    results_sheet.cell(row=write_row + 3, column=write_col).font = Font(bold=True)
    # Write hill slope result values header
    results_sheet.cell(row=write_row + 4, column=write_col).value = "Hill slope"
    results_sheet.cell(row=write_row + 4, column=write_col).font = Font(bold=True)
    # Write result values
    for i, sample_label in enumerate(group_results.keys(), 1):
        results_sheet.cell(row=write_row + 2, column=write_col + i).value = group_results[sample_label]['result']
        results_sheet.cell(row=write_row + 2, column=write_col + i).alignment = Alignment(horizontal='right')
        results_sheet.cell(row=write_row + 2, column=write_col + i).number_format = "0.0##"
        results_sheet.cell(row=write_row + 3, column=write_col + i).value = group_results[sample_label]['log_result']
        results_sheet.cell(row=write_row + 3, column=write_col + i).alignment = Alignment(horizontal='right')
        results_sheet.cell(row=write_row + 3, column=write_col + i).number_format = "0.0##"
        results_sheet.cell(row=write_row + 4, column=write_col + i).value = group_results[sample_label]['hill_slope']
        results_sheet.cell(row=write_row + 4, column=write_col + i).alignment = Alignment(horizontal='right')
        results_sheet.cell(row=write_row + 4, column=write_col + i).number_format = "0.0##"
    # Return next available cell to write on
    return write_row + 6, write_col


def results_to_data_file(results_sheet, analysis_type, results):
    """Function to handle the writing into result sheet for each group."""
    write_row, write_col = 1, 1
    # Write the results of each group into the results sheet
    for group_title, group_results in results.items():
        write_row, write_col = write_result_data(results_sheet, analysis_type, group_title, group_results, write_row, write_col)
    autosize_cols(results_sheet)


def run_analysis(data_file, analysis):
    """Run the analysis while writing and plotting its results."""
    # Run the analysis
    results = analysis.run()
    print_progress(f'{analysis.type} data analysed.')
    # Create a new results sheet
    results_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_{analysis.type.lower()}_RESULTS")
    # Write the results into the results sheet
    results_to_data_file(results_sheet, analysis.type, results)
    print_progress(f'Analysis results written into "{results_sheet.title}".')
    return results


def plot_group_graph(data_file_name, timestamp, group_title, group_results):
    """Plots group results into visual graphs, saving the plots as .png files."""
    # Create the graph
    graph = pyplot.figure(figsize=(10, 6))
    ax = graph.add_subplot(1, 1, 1)
    # Add data to the graph
    for sample, res in group_results.items():
        scatter = pyplot.scatter(res['x_data'], res['y_data'], label=f'{sample}')
        color = scatter.get_facecolor()[0]  # Extract the RGBA color tuple
        pyplot.plot(res['x_fit'], res['y_fit'], '--', color=color)
    # Add information and formatting to the graph
    font_props = {'fontsize': 16, 'fontweight': 'bold'}
    pyplot.xlabel('Log dilution', **font_props)
    pyplot.ylabel('Inhibition (%)', **font_props)
    pyplot.title(f'Variable Slope Fit for {group_title}', **font_props)
    pyplot.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    ax.spines['top'].set_linewidth(3)
    ax.spines['right'].set_linewidth(3)
    ax.spines['left'].set_linewidth(3)
    ax.spines['bottom'].set_linewidth(3)
    ax.tick_params(axis='both', which='major', labelsize=16, width=3)
    for label in ax.get_xticklabels() + ax.get_yticklabels():
        label.set_fontweight('bold')
    ax.set_ylim(-20, 120)
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=14)
    # Graph .png file name e.g.: "Graph_group_3_20250101_235959_(data_file.xlsx).png"
    graph_file_name = f"Graph_{group_title.replace(' ','_').lower()}_{timestamp}_({data_file_name}).png"
    graph_file_path = os.path.join(RESULT_FILES_DIRECTORY, graph_file_name)
    # Save the graph
    graph.savefig(graph_file_path, dpi=300, bbox_inches='tight')
    pyplot.close(graph)
    return graph_file_path


def zip_graphs(data_file_name, analysis_type, timestamp, graph_file_paths):
    """Function to zip graph .png files."""
    # Zip file name e.g.: "Graphs_20250101_235959_(data_file.xlsx).zip"
    zip_file_name = f"Graphs_{analysis_type}_{timestamp}_({data_file_name}).zip"
    zip_file_path = os.path.join(RESULT_FILES_DIRECTORY, zip_file_name)
    # Create the zip file
    with ZipFile(zip_file_path, 'w', ZIP_DEFLATED) as zipf:
        # Loop through graph file paths
        for i, graph_file_path in enumerate(graph_file_paths):
            # Add the wellplate file to the zip
            zipf.write(graph_file_path, os.path.basename(graph_file_path))
            # Remove the individual wellplate file after adding to zip
            os.remove(graph_file_path)
    return zip_file_name


def results_to_graphs(data_file_name, analysis_type, results):
    """Draws the graph for each of the analysis results and saves them as result files."""
    # Get current DateTime in 'YYYYMMDD_HHmmSS' format
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    graph_file_paths = []
    for group_title, group_results in results.items():
        # Plot each group's result into a graph
        graph_file_path = plot_group_graph(data_file_name, timestamp, group_title, group_results)
        graph_file_paths.append(graph_file_path)
        print_progress(f'Graph for {group_title} done.')
    # Store the graphs into a centralized .zip
    zip_file_name = zip_graphs(data_file_name, analysis_type, timestamp, graph_file_paths)
    return zip_file_name


def save_file(file, file_name):
    """Save the Excel file to the specified path.

    Parameters:
    file (Workbook): The Excel workbook to be saved.
    file_name (str): The name of the file to save.
    """
    try:
        # Save data file to the 'data_files' directory
        file.save(f"./data_files/{file_name}")
    except PermissionError:
        # Handle permission error if the data file is open in another program
        exit_by_error(f'Unable to save data file "{file_name}". It might be opened in an Excel editor.')




########## ACTUAL SCRIPT ##########


def run_script():
    print_action("Evaluating environment")
    print_task(f'Checking script environment structure.')
    evaluate_script_environment()
    print_success("Script environment structure accepted.")
    # NECESSARY DIRECTORIES HAVE BEEN LOCATED #

    print_action("Configuring script")
    print_task(f'Fetching script configuration from "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    configure_script()
    print_success("Script configured.")
    # SCRIPT HAS BEEN CONFIGURED #

    print_action("Selecting data file")
    data_file_name = select_excel_file(DATA_FILES_DIRECTORY)
    print_task(f'Loading data file "{data_file_name}".')
    # DATA FILE HAS BEEN SELECTED #
    data_file, data_file_name = load_excel_file(DATA_FILES_DIRECTORY, data_file_name)
    print_success(f'Data file "{data_file_name}" loaded.')
    # DATA FILE HAS BEEN LOADED #

    print_action("Selecting data sheet")
    data_sheet_name = select_datasheet(data_file)
    print_task(f'Loading data sheet "{data_sheet_name}".')
    # DATA SHEET HAS BEEN SELECTED #
    data_sheet = load_datasheet(data_file, data_sheet_name)
    print_success(f'Data sheet "{data_sheet_name}" loaded.')
    # DATA SHEET HAS BEEN LOADED #

    print_action("Selecting data groups")
    groups = select_groups(data_sheet)
    print_success(f"Groups selected.")
    # GROUPS HAVE BEEN SELECTED #

    print_action("Select analysis type")
    analysis_type = select_analysis_type()
    print_success(f"Analysis type selected.")
    # ANALYSIS HAS BEEN SELECTED #

    print_action("Extracting analysis data")
    print_task(f'Extracting the data of the {analysis_type} analysis from data sheet "{data_sheet.title}".')
    analysis = extract_analysis_data(data_sheet, groups, analysis_type)
    print_success(f"Analysis data extracted.")
    # ANALYSIS DATA EXTRACTED #

    print_action("Running analysis")
    print_task(f'Running the {analysis_type} analysis and creating a results sheet in data file "{data_file_name}".')
    results = run_analysis(data_file, analysis)
    print_success(f"Analysis run completed.")
    # ANALYSIS RUN COMPLETED #

    print_action("Plotting graphs")
    print_task(f'Plotting graphs for analysis results and storing them in "{RESULT_FILES_DIRECTORY}".')
    zip_file_name = results_to_graphs(data_file_name, analysis_type, results)
    print_success(f'Graphs plotted and stored into "{zip_file_name}".')
    # PLOTTING GRAPHS COMPLETED #

    print_action("Saving data file")
    print_task(f'Saving data file "{data_file_name}".')
    save_file(data_file, data_file_name)
    print_success(f'Data file "{data_file_name}" saved.')
    # DATA FILE HAS BEEN SAVED #




########## MAIN ENTRY POINT ##########


if __name__ == "__main__":
    handle_arguments(args)
    try:
        print_status("Started")
        run_script()
        print_status("Completed")
    except KeyboardInterrupt:
        exit_by_interruption()
