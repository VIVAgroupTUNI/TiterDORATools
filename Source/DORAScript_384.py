"""
384-Wellplate Data Processing Script

Summary:
This script is designed to process and analyze data from 384-wellplate experiments. The primary functionalities include
selecting and loading data files, defining groups and their samples, extracting and validating wellplate data,
integrating this data into predefined structures, performing statistical analysis, and writing the results back to
Excel files.
The script starts by ensuring the necessary libraries are installed and then initializes global attributes related to
the experiment setup. It includes user interaction functions for input validation and output formatting, ensuring
a smooth user experience.
Interruptions are handled gracefully, and the script provides clear progress and error messages throughout its execution.

Usage:
1. Ensure you have the required directories (`config_files` and `data_files`) with appropriate files.
2. Run the script from the command line.
3. Follow the prompts to:
   - Select a data file.
   - Choose a data sheet from the data file.
   - Define the groups, including sample count, and parallels per sample.
4. The script will then:
   - Extract wellplate data.
   - Integrate the data into the defined groups.
   - Perform statistical analysis.
   - Write the results and draw graphs back into the selected file.
5. The updated data file will be saved with the new statistics and graphs.
"""




########## IMPORTS ##########


import os
import time
import math
import textwrap
import json
from json import JSONDecodeError
from sys import exit
from sys import argv as args
from statistics import stdev, mean
try:
    from xlrd import open_workbook                                  # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.\n')
try:
    from openpyxl import load_workbook as load                      # type: ignore
    from openpyxl.styles import Font, Alignment                     # type: ignore
    from openpyxl.utils import get_column_letter as get_col_char    # type: ignore
    from openpyxl.workbook import Workbook                          # type: ignore
    from openpyxl.chart import LineChart, Reference                 # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xlsx files called "openpyxl". Run the installer to update your libraries.\n')




########## ATTRIBUTES ##########


SCRIPT_VERSION = "0.1.0"                                                                    # The public version number of the script

action_iterator = 1                                                                         # A global iterator used for tracking actions
DELAY = 0.05                                                                                # The delay (in seconds) used for timing outputs and creating a user-friendly experience
ENVIRONMENT = "SCR"                                                                         # The environment mode for the script, typically set to 'SCR' (short for scripted)
MOCK_ANALYSIS_TYPE = "EXCLUDE_DATA"                                                         # Analysis type mock filler to fill in the non used values of the wellplate
ALL_ANALYSIS_TYPES = [MOCK_ANALYSIS_TYPE]                                                   # Analysis types that are read from the confg file
CONFIG_PARAMETER = "analysis_types"                                                         # Parameter that is configured in the config file
CONFIG_FILE_NAME = "DORA_config"                                                            # The name of the script's configure file
CONFIG_FILE_TYPE = ".json"                                                                  # The type of the script's configure file
CONFIG_FILES_DIRECTORY = "./config_files/"                                                  # The directory within which the script looks for its configuration (.json) files
DATA_FILES_DIRECTORY = "./data_files/"                                                      # The directory within which the script looks for potential excel (.xls and.xlsx) data files
WELLPLATE_ROWS = 16                                                                         # The number of rows in the wellplate
WELLPLATE_COLS = 24                                                                         # The number of columns in the wellplate
WELLS_IN_A_PARALLEL = 12                                                                    # The number of wells in a parallel sample
PARALLEL_FIRST_COLS = [i for i in range(1, WELLPLATE_COLS + 1, WELLS_IN_A_PARALLEL)]        # A list of the wellplate columns where new parallel values start from
PARALLEL_CHOICES = [1, 2, 4, 8]                                                             # Allowed numbers of parallels per sample
MAX_SAMPLES = int(round(len(PARALLEL_FIRST_COLS) * WELLPLATE_ROWS / PARALLEL_CHOICES[0]))   # Maximum amount of samples based on the size of the wellplate and allowed number of parallels
CHART_LINE_WIDTH = 6000                                                                     # Width of the lines drawn in the linecharts
MARKER_SIZE = 5                                                                             # Size of the markers drawn in the linecharts
MARKERS = ["circle",    # Circle                                                            # Markers drawn as the linechart datapoints
           "square",    # Square
           "diamond",   # Diamond
           "triangle"]  # Triangle
COLORS = ["B22222",  # Firebrick Red                                                        # Colors of datapoint markers in linecharts
          "228B22",  # Forest Green
          "1E90FF",  # Dodger Blue
          "D2691E",  # Chocolate
          "20B2AA",  # Light Sea Green
          "8B008B",  # Dark Magenta
          "FF4500",  # Orange Red
          "6A5ACD",  # Slate Blue
          "556B2F",  # Dark Olive Green
          "DAA520",  # Goldenrod
          "800000",  # Maroon
          "C71585",  # Medium Violet Red
          "CD5C5C",  # Indian Red
          "9932CC",  # Dark Orchid
          "FF8C00",  # Dark Orange
          "48D1CC"]  # Medium Turquoise




########## CLASSES ##########


class ResetGroupException(Exception):
    """Custom exception to trigger a reset of a current group creation."""
    pass


class Group:
    """Class to manage group data, including samples and their associated configurations."""

    def __init__(self, analysis_type, ordinal, is_filler=False):
        """Initializes an Group object with a title."""
        self.is_filler = is_filler
        self.analysis_type = analysis_type
        self.ordinal = ordinal
        self.dil_series = []
        self.samples = []

    def __str__(self):
        """String representation of the Group object."""
        return f"Group {self.ordinal}, Samples: {len(self.samples)}"

    def add_sample(self, label, parallels):
        """Adds a sample to the group."""
        new_sample = Sample(label, parallels)
        self.samples.append(new_sample)

    def set_dil_series(self, init_dil, dil_factor):
        """Sets the dilution series of the group."""
        # Empty the diluition series list first
        self.dil_series.clear()
        # Start with the initial dilution value
        self.dil_series = [init_dil]
        # Generate the remaining 8 well dilutions
        for _ in range(8):
            if dil_factor < 0:
                # In case the factor is negative, use division
                next_value = self.dil_series[-1] / abs(dil_factor)
            else:
                # Otherwise use multiplication
                next_value = self.dil_series[-1] * dil_factor
            self.dil_series.append(next_value)

    def get_all_neg_ctrls(self):
        """Gets all negative control values across all samples."""
        return [control for sample in self.samples for control in sample.get_all_neg_ctrls()]

    def get_all_vir_ctrls(self):
        """Gets all viral control values across all samples."""
        return [control for sample in self.samples for control in sample.get_all_vir_ctrls()]

    def calc_neg_ctrl_avg(self):
        """Calculates the average of all negative control values across all samples."""
        controls = self.get_all_neg_ctrls()
        return sum(control["value"] for control in controls) / len(controls) if controls else None

    def calc_vir_ctrl_avg(self):
        """Calculates the average of all viral control values across all samples."""
        controls = self.get_all_vir_ctrls()
        return sum(control["value"] for control in controls) / len(controls) if controls else None
    
    def calc_neg_ctrl_avg_half(self):
        """Returns 50% of the average of all negative control values across all samples."""
        avg = self.calc_neg_ctrl_avg()
        return avg * 0.5 if avg is not None else None

    def calc_neg_ctrl_stdev(self):
        """Calculates the standard deviation of all negative control values across all samples."""
        values = [control["value"] for control in self.get_all_neg_ctrls()]
        return stdev(values) if 1 < len(values) else 0

    def calc_vir_ctrl_stdev(self):
        """Calculates the standard deviation of all viral control values across all samples."""
        values = [control["value"] for control in self.get_all_vir_ctrls()]
        return stdev(values) if 1 < len(values) else 0
    
    def get_statistics(self):
        return {
            "neg_ctrl_avg": self.calc_neg_ctrl_avg(),
            "neg_ctrl_avg_half": self.calc_neg_ctrl_avg_half(),
            "vir_ctrl_avg": self.calc_vir_ctrl_avg(),
            "neg_ctrl_stdev": self.calc_neg_ctrl_stdev(),
            "vir_ctrl_stdev": self.calc_vir_ctrl_stdev()
        }

    def get_normalized(self):
        """Returns normalized absorbance data of the entire Group object with fitted values."""
        topline = self.calc_neg_ctrl_avg()
        baseline = self.calc_vir_ctrl_avg()
        if topline is None or baseline is None or topline == baseline:
            raise ValueError(f"Normalization error: Invalid topline ({str(topline)}) or baseline ({str(baseline)}) values.")
        # Get their logarithm values
        log_dils = [math.log10(dil) for dil in self.dil_series]
        # Handle the data values
        norm_samples = []
        for norm_sample in self.samples:
            avgs = norm_sample.get_col_avgs()
            # Get normalized values (almost in 0 to 1 range)
            norm_avgs = norm_sample.get_normalized_col_avgs(baseline, topline)
            norm_samples.append({
                "label": norm_sample.label,
                "avgs": avgs,
                "normalized_avgs": norm_avgs
            })
        # Multiply normalized averages by 100 for display purposes.
        for sample in norm_samples:
            sample["normalized_avgs"] = [val * 100 for val in sample["normalized_avgs"]]
        # Return the normalized data
        return {
            "topline": topline,
            "baseline": baseline,
            "ordinal": self.ordinal,
            "analysis_type": self.analysis_type,
            "dils": self.dil_series,
            "log_dils": log_dils,
            "samples": norm_samples
        }


class Sample:
    """Class to manage sample data, including parallels and methods to perform calculations on them."""

    def __init__(self, label, parallels):
        """Initializes a Sample object with a label and the specified number of parallels."""
        self.label = label
        self.parallels = [Parallel() for _ in range(parallels)]

    def __str__(self):
        """String representation of the Sample object."""
        return f"Label: {self.label}, Parallels: {len(self.parallels)}"
    
    def get_all_neg_ctrls(self):
        """Retrieves all negative control absorbance values from all parallels."""
        return [control for parallel in self.parallels for control in parallel.get_neg_ctrls()]

    def get_all_vir_ctrls(self):
        """Retrieves all viral control absorbance values from all parallels."""
        return [parallel.get_vir_ctrl() for parallel in self.parallels]
    
    def normalize_sample(self, topline, baseline):
        """Returns normalized absorbance data of the entire sample object."""
        return {
            "label": self.label,
            "parallels": [parallel.normalize_parallel(topline, baseline) for parallel in self.parallels]
        }
    
    def get_col_avgs(self):
        """Calculates and returns the averaged values for columns 3-11 across all parallels."""
        col_vals = [[] for _ in range(9)]
        for parallel in self.parallels:
            values = parallel.get_values()
            for i, value in enumerate(values):
                col_vals[i].append(value)
        col_avgs = [sum(col) / len(col) for col in col_vals]
        return col_avgs
    
    def get_normalized_col_avgs(self, baseline, topline):
        """Calculates and returns the averaged values for columns 3-11 across all parallels."""
        col_normalized_vals = [[] for _ in range(9)]
        for parallel in self.parallels:
            normalized_values = parallel.get_normalized_values(baseline, topline)
            for i, value in enumerate(normalized_values):
                col_normalized_vals[i].append(value)
        col_normalized_avgs = [sum(col) / len(col) for col in col_normalized_vals]
        return col_normalized_avgs


class Parallel:
    """Class to store and manage absorbance data for control and viral wells within a wellplate."""

    def __init__(self):
        """Initializes a Parallel object with empty lists for controls and virals."""
        self.negative_control_1 = None
        self.negative_control_2 = None
        self.virals = []
        self.viral_control = None

    def __str__(self):
        """Returns a string representation of the Parallel object."""
        return f"Controls: {self.negative_control_1}, Virals: {self.virals}"
    
    def is_filled(self):
        """Returns True if all required values (negctrl1, negctrl2, 9 virals, and virctrl) are set."""
        return (
            self.negative_control_1 is not None and
            self.negative_control_2 is not None and
            self.viral_control is not None and
            len(self.virals) == 9
        )

    def set_neg_ctrl_1(self, absorbance, row_char, col_num):
        """Sets the first negative control absorbance value along with its row and column information."""
        self.negative_control_1 = {
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        }

    def set_neg_ctrl_2(self, absorbance, row_char, col_num):
        """Sets the second negative control absorbance value along with its row and column information."""
        self.negative_control_2 = {
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        }

    def add_vir(self, absorbance, row_char, col_num):
        """Adds a viral absorbance value along with its row and column information."""
        self.virals.append({
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        })

    def set_vir_ctrl(self, absorbance, row_char, col_num):
        """Sets the viral control absorbance value along with its row and column information."""
        self.viral_control = {
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        }

    def get_neg_ctrl_1(self):
        """Retrieves the first negative control absorbance value."""
        return self.negative_control_1
    
    def get_neg_ctrl_2(self):
        """Retrieves the first negative control absorbance value."""
        return self.negative_control_2

    def get_neg_ctrls(self):
        """Retrieves a list of control absorbance values."""
        return [self.get_neg_ctrl_1(), self.get_neg_ctrl_2()]
    
    def get_virs(self):
        """Retrieves a list of viral absorbance values"""
        return self.virals
    
    def get_vir_ctrl(self):
        """Retrieves the viral control absorbance value."""
        return self.viral_control
    
    def get_wells(self):
        """Retrieves all well absorbance values in order."""
        return self.get_neg_ctrls() + self.get_virs() + [self.get_vir_ctrl()]
    
    def get_filled_well_count(self):
        """Returns the count of wells that contain valid numeric absorbance values."""
        wells = self.get_wells()
        return sum(1 for well in wells if well and isinstance(well.get("value"), (int, float)))
    
    def get_values(self):
        """Returns values for wells in columns 3-11."""
        values = []
        for well in self.get_virs():
            value = well["value"]
            values.append(value)
        return values
    
    def get_normalized_values(self, baseline, topline):
        """Returns normalized values for wells in columns 3-11."""
        normalized_values = []
        for well in self.get_virs():
            normalized_value = (well["value"] - baseline) / (topline - baseline)
            normalized_values.append(normalized_value)
        return normalized_values




########## ARGUMENT HANDLING ##########


def info():
    """Function to print out instructions on how to use the script for the user."""
    try:
        # Print tutorial started status
        print_status("Tutorial Started", tailing_line_break=True)
        # Print each section of the tutorial with detailed instructions
        print_info("1) Evaluating Environment:", [
            "The script begins by checking the structure of the required directories.",
            "It checks for the existence of 'config_files' and 'data_files' directories.",
            "It ensures that each directory contains valid files (like .json, .xls and .xlsx)."
        ])
        print_info("2) Configuring Script:", [
            "Next the script will look for its configuration file 'DORA_config.json' from the 'config_files' directory.",
            "It will load and apply its parameters into the script.",
            "It ensures that the config file can be found and its parameters are valid."
        ])
        print_info("3) Selecting Data File:", [
            "The user will then be prompted to select a data file from the 'data_files' directory.",
            "The available excel files will be listed, and the user will again choose one by its corresponding number."
        ])
        print_info("4) Selecting Data Sheet:", [
            "The user will select a sheet from the chosen data file.",
            "The available sheets will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("5) Defining Groups:", [
            "The user will define the to be runned groups.",
            "For each group, the user will input the initial dilution, dilution factor, total samples, and parallels per sample.",
            "The user will also provide a label for each sample.",
            "The user will be able to exclude parts of the wellplate data from being analyzed."
        ])
        print_info("6) Extracting Wellplate Data:", [
            "The script will then read the selected data sheet and extract the wellplate data.",
            f"It will look for the {WELLPLATE_ROWS}x{WELLPLATE_COLS} wellplate segment, starting from the appropriate row and column, and collect the absorbance values."
        ])
        print_info("7) Integrating Wellplate Data:", [
            "The script will integrate the extracted wellplate data into the configured groups.",
            "Each sample and parallel within the groups will be populated with the corresponding absorbance values from the wellplate data."
        ])
        print_info("8) Calculating Statistics:", [
            "The script will handle the integrated data by calculating statistics for each of the groups.",
            "The script will then write the statistical data and draw graphs into the data file.",
            "It will create new sheets in the data file as storage for the written data."
        ])
        print_info("9) Writing Exports:", [
            "The script will write the data required for a later stage statistical analysis into the data file.",
            "It will create a new sheet in the data file as storage for the exportable data."
        ])
        print_info("10) Saving Data File:", [
            "Finally the script will save the updated data file with the new statistics sheets to ensure that the intermediate results are preserved."
        ])
        
        # Print additional notes
        print_info("Notes:", [
            "The script may be interrupted at any point of its execution safely by hitting (CTRL+C).",
            "Whenever the script prompts the user to provide numeric data that involves decimals, it allows the user to use either Finnish commas (,) or English dots (.) as decimal separators.",
            "Should the user provide a negative dilution factor, the dilution series will be descending instead of ascending.",
            "Should the user provide an incorrect value for an group during its definition, they may reset defining the current group and start anew by giving a command of \"...\" instead of the actual asked value."
            "Should the config file be lost, the script will be able to restore its template as long as the config directory exists.",
            "Should the user pick a .xls file as data file, the script will automatically conver it into a .xlsx file and replace the old .xls file altogether.",
            "The script looks for config files within the './config_files/' directory.",
            "The script looks for data files within the './data_files/' directory.",
        ])
        # Print explanation for each type of user prompt that the script uses
        print_info("User Prompts:", [
            "Input Prompt: When the script requires user input, it will display a prompt with a '>' symbol (e.g., \"Please select the data file > \"). The user will enter their response after the '>' symbol.",
            "Choices: Whenever the script presents multiple options to choose from, they will be listed with a corresponding command in square brackets (e.g., \"[1] data_file_1.xlsx\"). The user will select an option by typing its command.",
            "Task: The script will print task status messages to indicate what it is currently doing. These messages will start with \"> Task: \" followed by the task description.",
            "Progress: During longer operations, the script will print progress messages to keep the user informed. These messages will start with \"> Progress: \" followed by the progress description.",
            "Success: Upon successful completion of an operation, the script will print success messages starting with \"> Success: \" followed by a brief description of the success.",
            "Error: If an error occurs during an operation, the script will print error messages starting with \"> Error: \" followed by the error description."
        ])
        print_status("Tutorial completed", leading_line_break=False)
    except KeyboardInterrupt:
        exit_by_interruption(True)


def handle_arguments(args):
    """Handles the command line arguments provided to the script."""
    # Check if any arguments were given at launch
    if 1 < len(args):
        param = args[1].lower()
        if param in ["i", "info", "h", "help"]:
            # If the first argument was either "info" or "help", start tutorial
            info()
            exit()




########## EXIT ROUTES ##########


def exit_by_error(message=None):
    """Exits the script with a given error message."""
    try:
        if message:
            time.sleep(DELAY)
            print_error(message)
        time.sleep(DELAY)
        print_status("Failed")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")


def exit_by_interruption(during_info=False):
    """Exits the script due to an interruption (e.g., KeyboardInterrupt, CTRL+C)."""
    try:
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        if during_info:
            print_status("Tutorial stopped")    
        else:
            print_status("Stopped")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")




########## USER INTERACTION ##########


def prompt_input(prompt_str):
    """Gets user input after displaying a prompt."""
    try:
        time.sleep(DELAY)
        user_input = input(f"{prompt_str} > ").strip()
    except KeyboardInterrupt:
        exit_by_interruption()
    return user_input


def get_user_input(prompt, validation_func=None, type_select=True, allow_interrupt=False):
    """Gets and validates user input.

    Parameters:
    prompt (str): The prompt message to display.
    validation_func (function): The function to validate user input. Default is None.
    type_select (bool): Determines if input is for selection (True) or entry (False). Default is True.

    Returns:
    str: The validated user input.
    """
    # Print a blank line if type_select is True
    if type_select:
        print()
    while True:
        # Get user input
        user_input = prompt_input(prompt)
        # Listen for input of "..." if interruptions are allowed
        if allow_interrupt and user_input.strip() == "...":
            raise ResetGroupException
        # If no validation function is provided, return the input
        if not validation_func:
            if type_select:
                print()
            return user_input
        # Validate the user input
        validation_result, validated_input = validation_func(user_input)
        # If validation is successful, return the validated input
        if validation_result is True:
            if type_select:
                print()
            return validated_input
        # Print error message if validation fails
        print_error(validated_input)




########## PRINT FUNCTIONS ##########


def print_status(status_string, leading_line_break=True, tailing_line_break=False):
    """Prints a status message with version in the top border."""
    try:
        time.sleep(DELAY)
        text = f"DORA-384 SCRIPT {status_string.upper()}"
        # Middle line padding
        asterisks = "*" * 3
        spaces = " " * 5
        middle_line = f"{asterisks}{spaces}{text}{spaces}{asterisks}"
        # Top border with version
        top_border_text = f" v{SCRIPT_VERSION} "
        total_width = len(middle_line)
        top_line = "*" * ((total_width - len(top_border_text)) // 2) + top_border_text
        if len(top_line) < total_width:
            top_line += "*" * (total_width - len(top_line))
        bottom_line = "*" * len(middle_line)

        if leading_line_break:
            time.sleep(DELAY)
            print()
        time.sleep(DELAY)
        print(top_line)
        time.sleep(DELAY)
        print(middle_line)
        time.sleep(DELAY)
        print(bottom_line)
        if tailing_line_break or status_string.strip().lower() == "completed":
            time.sleep(DELAY)
            print()
    except KeyboardInterrupt:
        exit_by_interruption()


def print_action(action_str):
    """Prints a separator line for actions."""
    try:
        global action_iterator
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        print(f"========== Action {action_iterator}: {action_str.title()} ==========")
        action_iterator += 1
    except KeyboardInterrupt:
        exit_by_interruption()


def print_task(task_str):
    """Prints a task message."""
    try:
        time.sleep(DELAY)
        print(f"> Task: {task_str}..")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_choice(choice_cmd, choice_str):
    """Prints a choice option."""
    try:
        time.sleep(DELAY)
        print(f"[{choice_cmd}] {choice_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_progress(progress_str):
    """Prints a progress message."""
    try:
        time.sleep(DELAY)
        print(f"> Progress: {progress_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_success(success_str):
    """Prints a success message."""
    try:
        time.sleep(DELAY)
        print(f"> Success: {success_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_error(error_str):
    """Prints an error message."""
    try:
        time.sleep(DELAY)
        print(f"> Error: {error_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_info(header, messages, max_length=100):
    """Prints information with a delay and text wrapping."""
    try:
        INDENT = 3
        time.sleep(DELAY)
        print(header.title())
        for message in messages:
            wrapped_message = textwrap.fill(f" - {message}", max_length, subsequent_indent=' ' * INDENT)
            time.sleep(DELAY)
            print(wrapped_message)
        time.sleep(DELAY)
        print()
    except KeyboardInterrupt:
        exit_by_interruption(True)




########## VALIDATION FUNCTIONS ##########


def validate_string(input_str):
    """Validates if the input is a valid non-empty string.
    
    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the string value if valid, or an error message if not.
    """
    if not isinstance(input_str, str):
        return False, "Input must be a string."
    cleaned = input_str.strip()
    if not cleaned:
        return False, "Input cannot be empty or whitespace."
    return True, cleaned


def validate_digit(input_str):
    """Validate if the input string is a digit.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the digit value if valid, or an error message if not.
    """
    if input_str.isdigit():
        return True, int(input_str)
    return False, f'Your input of "{input_str}" was not a positive whole number.'


def validate_non_occupied_string(input_str, occupied, analysis_type):
    """Validate if the input string is an digit that isn't already occupied.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    is_valid, validation_result = validate_string(input_str)
    if not is_valid:
        return False, validation_result
    if validation_result not in occupied:
        return True, validation_result
    return False, f'Your input "{input_str}" was already taken by another defined group of {analysis_type}.'


def validate_integer(input_str):
    """Validate if the input string is an integer.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    try:
        return True, int(input_str)
    except ValueError:
        return False, f'Your input of "{input_str}" was not a whole number.'


def validate_float(input_str):
    """Validate if the input string is a float, allowing for both Finnish (,) and English (.) decimal separators.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the float value if valid, or an error message if not.
    """
    try:
        input_str = input_str.replace(',', '.')
        float_value = float(input_str)
        return True, float_value
    except ValueError:
        return False, f'Your input "{input_str}" was not a number.'


def validate_non_empty(input_str):
    """Validate if the input string is not empty.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the input string if valid, or an error message if not.
    """
    if input_str.strip():
        return True, input_str
    return False, f"Your input was empty."


def validate_min_max(input_str, min_choice, max_choice, escape_str=None):
    """Validate if the input string is a digit within a specified range, with an optional escape string.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer or 0 if escaped, or an error message if not.
    """
    if escape_str is not None and input_str.strip() == escape_str:
        return True, 0
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if min_choice <= validation_result <= max_choice:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [{min_choice}, {max_choice}].'


def validate_yes_no(input_str):
    """Validate if the input string is a 'yes' or 'no' response.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated input string if valid, or an error message if not.
    """
    input_str = input_str.lower().strip()
    if input_str in ['y', 'yes', 'n', 'no']:
        return True, input_str
    return False, f'Your input "{input_str}" was invalid. Choose either yes or no.'


def validate_sample_count(input_str):
    """Validate if the input string is a digit within the range of [1, 32].

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer if valid, or an error message if not.
    """
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if 1 <= validation_result <= MAX_SAMPLES:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [1, {MAX_SAMPLES}].'


def validate_parallel_count(input_str):
    """Validate if the input string is a digit and is either 4 or 8.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer if valid, or an error message if not.
    """
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if validation_result in PARALLEL_CHOICES:
        return True, validation_result
    return False, f'Your input "{input_str}" was invalid. Choose either {" or ".join(map(str, PARALLEL_CHOICES))}.'


def validate_row_char(row_char):
    """Check if the row character is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated row character if valid, or an error message if not.
    """
    if row_char is None:
        return False, "Row character was missing."
    if isinstance(row_char, str):
        row_char_cap = row_char.upper()
        if len(row_char_cap) == 1 and row_char_cap.isalpha():
            return True, row_char_cap
    return False, f"Row character \"{row_char}\" was invalid. Expected a single character string like 'A' or 'a'."


def validate_start_row_char(row_char):
    """Validate the starting row character 'A'.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated row character if valid, or an error message if not.
    """
    is_valid, validation_result = validate_row_char(row_char)
    if not is_valid:
        return False, validation_result
    if validation_result == 'A':
        return True, validation_result
    return False, f"Starting row character \"{row_char}\" was invalid. Expected either 'A' or 'a'."


def validate_col_num(col_num):
    """Check if the column number value is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated column number as a zero-padded string if valid, or an error message if not.
    """
    if col_num is None:
        return False, "Column number was missing."
    col_num_str = str(col_num).zfill(2)
    if col_num_str.isdigit() and 1 <= int(col_num_str) <= 24:
        return True, col_num_str
    return False, f'Column number "{col_num}" was invalid. Expected a whole number in a format of "1" or "01".'


def validate_start_col_num(col_num):
    """Validate the starting column number '1'.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated column number as a zero-padded string if valid, or an error message if not.
    """
    is_valid, validation_result = validate_col_num(col_num)
    if not is_valid:
        return False, validation_result
    if validation_result == '01':
        return True, validation_result
    return False, f'Starting column number "{col_num}" was invalid. Expected either "1" or "01".'


def validate_absorbance(absorbance, row_char, col_num):
    """Check if the absorbance value is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the absorbance value if valid, or an error message if not.
    """
    if absorbance is None:
        return True, absorbance  # None is acceptable for absorbance
    if isinstance(absorbance, (int, float)):
        return True, absorbance
    return False, f'Absorbance value "{absorbance}" of well {row_char}{col_num} was invalid. Expected either a number or an empty value.'




########## SEARCH FUNCTIONS ##########


def find_wellplate_start(sheet):
    """Function to find the starting row and column for the wellplate data.
    
    Returns:
    tuple: A tuple containing the starting row and column.
    """
    START_COL = 1
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value == 'A':
            start_row = row - 1
            # Verify the row character 'A'
            row_char = sheet.cell(row=start_row + 1, column=START_COL).value
            is_valid, validation_result = validate_start_row_char(row_char)
            if not is_valid:
                continue
            # Verify the column number 1 starting from the row above the 'A' row and one column to the left
            col_num = sheet.cell(row=start_row, column=START_COL + 1).value
            is_valid, validation_result = validate_start_col_num(col_num)
            if not is_valid:
                continue
            # Verify the empty cell exists between the row character 'A' and column number 1
            if sheet.cell(row=start_row, column=START_COL).value:
                continue
            return start_row, START_COL
    exit_by_error(f'Could not locate wellplate in data sheet "{sheet.title}".')




########## FILE CONVERTER ##########


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Function to convert deprecated .xls files into .xlsx files."""
    # Open the .xls file using xlrd
    workbook_xls = open_workbook(xls_path)
    workbook_xlsx = Workbook()
    for sheet_index in range(workbook_xls.nsheets):
        # One sheet at a time
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        if sheet_index == 0:
            # Activate the workbook to create its first sheet
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
        else:
            # Add them otherwise
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
        # One row at a time
        for row_index in range(sheet_xls.nrows):
            for col_index in range(sheet_xls.ncols):
                # Write each cell
                sheet_xlsx.cell(row=row_index + 1, column=col_index + 1).value = sheet_xls.cell_value(row_index, col_index)
    try:
        # Save the new .xlsx file using openpyxl
        workbook_xlsx.save(xlsx_path)
        # Delete the old .xls file using xlrds
        os.remove(xls_path)
    except PermissionError:
        # If couldn't remove the original file, remove the new one to avoid duplicates
        os.remove(xlsx_path)
        exit_by_error(f'File "{xls_path}" was already open in another program. Please close it and try again.')




########## CORE FUNCTIONS ##########


def directory_contains_file(directory, filename):
    """Check if a specific file exists in the given directory."""
    return filename in os.listdir(directory)


def get_excel_files(directory):
    """Function to find all excel files (.xls and .xlsx) in the given directory that do not start with '~$'."""
    return [file for file in os.listdir(directory) if (file.endswith('.xlsx') or file.endswith('.xls')) and not file.startswith('~$')]


def generate_mock_template():
    """Function to create mock analysis types instance."""
    return ["Analysis type 1",
            "Analysis type 2",
            "Analysis type 3"]


def create_config_template():
    """Function to create a template configuration file."""
    # Ensuring the config directory still exists.
    os.makedirs(CONFIG_FILES_DIRECTORY, exist_ok=True)
    print_task(f'Creating config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" in config directory.')
    # Determine the configuration file path
    config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    # Create the configuration file
    with open(config_file_path, "w", encoding="utf-8") as file:
        # Add template content
        mock_analysis_types = generate_mock_template()
        json.dump({CONFIG_PARAMETER: mock_analysis_types}, file, indent=4)
    print_success("Config file created, terminating script.")


def evaluate_script_environment():
    """Check if the required directories with acceptable content exist in the current working directory."""
    # Check for the existence of config files directory
    if not os.path.isdir(CONFIG_FILES_DIRECTORY):
        exit_by_error(f'Config files directory "{CONFIG_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Config files directory "{CONFIG_FILES_DIRECTORY}" located.')
    # Check if config files directory contains the scirpt's config file
    config_file = directory_contains_file(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    if not config_file:
        # In case the config file cannot be located
        print_error(f'Config files directory didn\'t contain file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
        # Ask the user if the missing config file should be restored
        yes_no_value = get_user_input('The script will be terminated...\nRestore the missing config file template on exit? [y/n]', validate_yes_no)
        if yes_no_value in ['y', 'yes']:
            # Restore the config file template
            create_config_template()
            # Stop the script
            exit_by_error()
        else:
            # Terminate the script
            exit_by_error("Terminating script.")
    print_progress(f'Config files directory contained file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    # Check for the existence of data files directory
    if not os.path.isdir(DATA_FILES_DIRECTORY):
        exit_by_error(f'Data files directory "{DATA_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Data files directory "{DATA_FILES_DIRECTORY}" located.')
    # Check if data files directory contains at least one excel file that doesn't start with "~$"
    data_files = get_excel_files(DATA_FILES_DIRECTORY)
    if not data_files:
        exit_by_error(f'Data files directory did not contain any excel files.')
    print_progress(f'Data files directory contained excel files.')


def load_configuration():
    """Attempt to Load configration from the configuration file."""
    try:
        # Determine the configuration file path
        config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
        with open(config_file_path, "r", encoding="utf-8") as file:
            # Load from the configuration file
            config = json.load(file)
            # Return the loaded configuration
            return config
    except FileNotFoundError:
        # Configuration file could not be found
        exit_by_error(f'Config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be found.')
    except JSONDecodeError:
        # Configuration could not be loaded
        exit_by_error(f'Configuration from file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be loaded.')
    

def apply_configuration(config):
    """Check that the configuration is valid and apply it."""
    # Fetch "analysis types" from configuration
    analysis_types = config.get(CONFIG_PARAMETER, None)
    # Check that configuration is valid
    if analysis_types is None:
        # Configuration didn't contain "analysis types"
        exit_by_error(f'Loaded configuration is missing the "{CONFIG_PARAMETER}" property. Please add it.')
    if not isinstance(analysis_types, list):
        # Aanalysis types wasn't a list
        exit_by_error(f'The "{CONFIG_PARAMETER}" property must be a list, but it is currently of type {type(analysis_types).__name__}.')
    if len(analysis_types) < 1:
        # Analysis types was empty
        exit_by_error(f'The "{CONFIG_PARAMETER}" was empty. Please add at least one analysis type.')
    # Validate each analysis type
    for i, item in enumerate(analysis_types):
        a_num = i + 1
        if not isinstance(item, str):
            # Analysis type wasn't a string
            exit_by_error(f'Each analysis type must be a string, but analysis type {a_num} is currently of type {type(item).__name__}.')
        if not item.strip():
            # Name was empty or whitespace
            exit_by_error(f'Each analysis type must have an actual value, but analysis type {a_num} is currently empty or white space.')
        # Store validated analysis type in list
    ALL_ANALYSIS_TYPES[:0] = analysis_types


def configure_script():
    """Initialize the script by loading and applying a configuration."""
    # Load the configuration
    config = load_configuration()
    print_progress("Configuration loaded.")
    # Apply the configuration
    apply_configuration(config)
    print_progress("Configuration applied.")


def select_excel_file(directory):
    """Selects an Excel file from the specified directory."""
    # List all Excel files in the directory, excluding temporary files
    try:
        excel_files = get_excel_files(directory)
    except FileNotFoundError:
        exit_by_error(f'Directory "{directory}" could not be located.')
    # If no Excel files are found, exit the script with an error message
    if not excel_files:
        exit_by_error(f'Directory "{directory}" did not contain any excel files.')
    # Print choices for the user to select from
    for i, file in enumerate(excel_files):
        print_choice(i+1, file)
    # User chooses a file
    int_value = get_user_input(f'Please select the file', lambda input_str: validate_min_max(input_str, 1, len(excel_files)))
    return excel_files[int_value - 1]


def load_excel_file(directory, file_name):
    """Load an Excel file from the given path and return the workbook object."""
    try:
        # Convert .xls to .xlsx if necessary
        if file_name.endswith('.xls'):
            xls_path = directory + file_name
            file_name += 'x'
            xlsx_path = directory + file_name
            convert_xls_to_xlsx(xls_path, xlsx_path)
            print_progress("Converted the old .xls file into a new .xlsx file.")
        else:
            xlsx_path = directory + file_name
        # Load the Excel file
        file = load(xlsx_path)
        # Save the unchanged file to check if it is accessible
        file.save(xlsx_path)
        return file, file_name
    except FileNotFoundError:
        # Handle file not found error
        exit_by_error(f'File "{file_name}" could not be located.')
    except PermissionError:
        # Handle permission error if the file is already open
        exit_by_error(f'File "{file_name}" was already open in another program. Please close it and try again.')
    except Exception as e:
        # Handle any other exceptions that occur during file loading
        exit_by_error(f'Failed loading file "{file_name}".  > REASON: {e}.')


def select_datasheet(data_file):
    """Prompt the user to select a data sheet from the given data file."""
    # List all sheet names in the data file
    datasheets = [sheet for sheet in data_file.sheetnames]
    for i, sheet in enumerate(datasheets):
        print_choice(i+1, sheet)
    # User chooses a sheet
    int_value = get_user_input('Please select the sheet', lambda input_str: validate_min_max(input_str, 1, len(data_file.sheetnames)))
    return datasheets[int_value - 1]


def load_datasheet(data_file, data_sheet_name):
    """Load the specified data sheet from the given data file."""
    return data_file[data_sheet_name]


def define_groups():
    """Define the groups by collecting user inputs for each groups' details."""
    groups = []
    filler_ordinal = 0
    while True:
        try:
            # Display available analysis types for selection
            for i, analysis_type in enumerate(ALL_ANALYSIS_TYPES):
                print_choice(i+1, analysis_type)
            # Get user input for the selected analysis type
            user_input = get_user_input("Please select the analysis type", lambda input_str: validate_min_max(input_str, 1, len(ALL_ANALYSIS_TYPES)))
            analysis_type = ALL_ANALYSIS_TYPES[int(user_input) - 1]
            # Based on the type, determine wether an actual group or filler
            is_filler = analysis_type == MOCK_ANALYSIS_TYPE
            # Create a new Group object and collect its details
            if is_filler:
                filler_ordinal += 1
                group = Group(analysis_type, filler_ordinal, is_filler)
                sample_count = get_user_input(f"Please enter the sample count (Excluded {filler_ordinal})", validate_sample_count, False, True)
                parallel_count = get_user_input(f"Please enter the parallel count (Excluded {filler_ordinal})", validate_parallel_count, False, True)
                # Use mock naming for filler
                for i in range(sample_count):
                    label = f"excluded {i+1}"
                    group.add_sample(label, parallel_count)
                print_progress(f'Excluded dataset {filler_ordinal} added.')
            else:
                # Collect group details
                occupied_ordinals = [group.ordinal for group in groups if group.analysis_type == analysis_type]
                ordinal = get_user_input("Please provide a group label", lambda input_str: validate_non_occupied_string(input_str, occupied_ordinals, analysis_type), False, True)
                group = Group(analysis_type, ordinal, is_filler)
                init_dil = get_user_input(f'Please enter the initial dilution (Group {ordinal})', validate_float, False, True)
                dil_factor = get_user_input(f'Please enter the dilution factor (Group {ordinal})', validate_integer, False, True)
                group.set_dil_series(init_dil, dil_factor)
                sample_count = get_user_input(f"Please enter the sample count (Group {ordinal})", validate_sample_count, False, True)
                parallel_count = get_user_input(f"Please enter the parallel count (Group {ordinal})", validate_parallel_count, False, True)
                # Collect sample details
                for i in range(sample_count):
                    label = get_user_input(f"Please enter a label for sample {i+1} (Group {ordinal})", validate_non_empty, False, True)
                    group.add_sample(label, parallel_count)
                print_progress(f'Group {ordinal} defined.')
            # Add the defined group or filler to the list of groups
            groups.append(group)
            # Ask if the user wants to configure another group
            yes_no_value = get_user_input('Would you like to add another group? [y/n]', validate_yes_no)
            if yes_no_value in ['n', 'no']:
                break
        except ResetGroupException:
            print()
            time.sleep(DELAY)
            print(f"> Info: Group reset.")
            print()
    return groups


def extract_wellplate(data_sheet, start_row, start_col):
    """Get a segment of the worksheet based on the starting point, number of rows, and number of columns.

    Parameters:
    data_sheet (Worksheet): The sheet to get data from.
    start_row (int): The starting row of the segment.
    start_col (int): The starting column of the segment.

    Returns:
    list: A list of lists representing the segment.
    """
    # Determine the coordinate of the first cell in the segment
    first_cell = data_sheet.cell(row=start_row, column=start_col).coordinate
    # Determine the coordinate of the last cell in the segment
    last_cell = data_sheet.cell(row=start_row + WELLPLATE_ROWS, column=start_col + WELLPLATE_COLS).coordinate
    # Return the segment of the worksheet as a list of lists
    return data_sheet[first_cell:last_cell]


def get_wellplate_data(wellplate, groups):
    """Creates a dictionary of absorbance values from the wellplate segment.

    Returns:
    dict: A dictionary with (row_char, col_num) as keys and absorbance values as values.
    """
    # Determine required data counts
    parallel_count = sum(len(sample.parallels) for group in groups for sample in group.samples)
    required_total = parallel_count * WELLS_IN_A_PARALLEL
    realized_total = 0
    # Create the dictionary to hold the absorbance values
    wellplate_data = {}
    # Extract the data as new key (row character, column number) value (absorbance) pairs of the dictionary
    for i, row in enumerate(wellplate):
        if i == 0:
            continue  # Skip the first row which contains headers
        row_char = wellplate[i][0].value
        # Validate the row character
        is_valid, validation_result = validate_row_char(row_char)
        if not is_valid:
            exit_by_error(validation_result)
        row_char = validation_result
        for j, well in enumerate(row):
            if j == 0:
                continue  # Skip the first column which contains row headers
            col_num = wellplate[0][j].value
            # Validate the column number
            is_valid, validation_result = validate_col_num(col_num)
            if not is_valid:
                exit_by_error(validation_result)
            col_num = validation_result
            absorbance = well.value
            # Validate the absorbance value
            is_valid, validation_result = validate_absorbance(absorbance, row_char, int(col_num))
            if not is_valid:
                exit_by_error(validation_result)
            absorbance = validation_result
            # Add the valid absorbance value to the dictionary
            wellplate_data[(row_char, col_num)] = absorbance
            if isinstance(absorbance, (int, float)):
                # If the well contains a numerical value, increment the total found absorbance values counter
                realized_total += 1
    if realized_total < required_total:
        # If there weren't enough data in the wellplate to cover the requirements, terminate the script
        exit_by_error(f"Excess data provided: wellplate contains ({realized_total}), but you configured ({required_total}). Ensure that all data points are correctly configured.")
    elif required_total < realized_total:
        # If there was more data on the wellplate than required to cover the groups, terminate the script
        exit_by_error(f"Insufficient data provided: wellplate contains ({realized_total}), but you configured ({required_total}). Ensure that all data points are correctly configured.")
    return wellplate_data


def extract_wellplate_data(data_sheet, groups):
    """Extracts wellplate data from the specified data sheet."""
    # Find the starting row and column of the wellplate
    START_ROW, START_COL = find_wellplate_start(data_sheet)
    print_progress(f"Wellplate located in segment [{chr(64+START_COL)}{START_ROW}, {chr(64+START_COL+WELLPLATE_COLS)}{START_ROW+WELLPLATE_ROWS}].")
    # Extract the wellplate segment based on the starting point
    wellplate = extract_wellplate(data_sheet, START_ROW, START_COL)
    # Get the absorbance values from the wellplate segment
    wellplate_data = get_wellplate_data(wellplate, groups)
    print_progress("Wellplate data was valid.")
    return wellplate_data


def get_first_key(wellplate_data):
    """Find the first key in the wellplate data that contains a numeric value.

    Parameters:
    wellplate_data (dict): Dictionary with (row_char, col_num) as keys and absorbance values as values.

    Returns:
    tuple: The first (row_char, col_num_padded) key that contains a numeric value, or None if no valid key is found.
    """
    # Initialize the starting key
    current_key = ('A', '01')
    for _ in range(WELLPLATE_ROWS * WELLPLATE_COLS):
        value = wellplate_data.get(current_key)
        if value is not None:
            return current_key
        current_key = get_next_key(current_key)
    return None  # Return None if no valid key is found


def get_next_key(key):
    """Update the row character and column index as per the required logic and return the full key.

    Parameters:
    current_key (tuple): The current (row_char, col_num) key.

    Returns:
    tuple: Updated (row_char, col_num_padded) key.
    """
    ASCII_A = ord('A')
    ASCII_B = ord('B')
    ASCII_JUMP = ASCII_A + WELLPLATE_ROWS
    ASCII_RESET = ASCII_B + WELLPLATE_ROWS
    # Retrieve the current key values
    row_char, col_num = key
    row_char_ascii = ord(row_char)
    col_num_int = int(col_num)
    # Determine the parallel segment columns
    segment_start = max([start_col for start_col in PARALLEL_FIRST_COLS if start_col <= col_num_int])
    segment_end = segment_start + WELLS_IN_A_PARALLEL
    # Determine the next column number
    next_col_num_int = col_num_int + 1
    # By default the next row character will remain the same unless deemed otherwise
    next_row_char = row_char
    # Check if the column number has reached the end of the current segment
    if segment_end <= next_col_num_int:
        # Reset the col number
        next_col_num_int = segment_start
        # Determine the next row character
        next_row_char_ascii = row_char_ascii + 2
        if ASCII_RESET <= next_row_char_ascii:
            # Jump back to the start after exhausting both odd and even rows
            next_row_char_ascii = ASCII_A
            next_col_num_int = segment_end
            if segment_start == PARALLEL_FIRST_COLS[-1]:
                return None # If the next key is outside the bounds of the wellpexit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.')alte, return None
        elif ASCII_JUMP <= next_row_char_ascii:
            # Jump to the even row characters if the odd rows are exhausted
            next_row_char_ascii = ASCII_B
        next_row_char = chr(next_row_char_ascii)
    next_col_num = str(next_col_num_int).zfill(2)
    next_key = (next_row_char, next_col_num)
    return next_key


def integrate_wellplate_data(wellplate_data, groups):
    """Integrates the collected wellplate data into the appropriate groups."""
    current_key = get_first_key(wellplate_data)  # Find the first key with valid values
    if not current_key:
        # If the first key could not be determined, the wellplate didn't contain any valid absorbance data
        exit_by_error("Couldn't find any absorbance values in the wellplate.")
    for group in groups:
        group_title = "Excluded dataset" if group.is_filler else "Group"
        for sample in group.samples:
            for parallel in sample.parallels:
                while not parallel.is_filled(): # Ensure this parallel is fully filled before moving to the next one
                    if not current_key:
                        # If the current key is None, the wellplate ran out of data and the user input or data was incorrect
                        exit_by_error(f'The wellplate did not contain enough absorbance data to integrate group "Group {group.ordinal}", sample "{sample.label}".')
                    if current_key not in wellplate_data:
                        # If the wellplate didn't contain the key, the wellplate in the data file was faulty
                        exit_by_error(f"Couldn't find an absorbance value in well {current_key}.")
                    absorbance = wellplate_data[current_key]
                    # If absorbance value is not a number
                    if not isinstance(absorbance, (int, float)):
                        # Move to the next key and continue
                        current_key = get_next_key(current_key)
                        continue
                    # Add absorbance to controls or virals based on well number
                    next_well_nbr = parallel.get_filled_well_count() # Determine next available well number
                    if next_well_nbr == 0:
                        parallel.set_neg_ctrl_1(absorbance, current_key[0], current_key[1])
                    elif next_well_nbr == 1:
                        parallel.set_neg_ctrl_2(absorbance, current_key[0], current_key[1])
                    elif next_well_nbr == 11:
                        parallel.set_vir_ctrl(absorbance, current_key[0], current_key[1])
                    else:
                        parallel.add_vir(absorbance, current_key[0], current_key[1])
                    # Update to the next key
                    current_key = get_next_key(current_key)
            print_progress(f'Integrated data for {group_title} {group.ordinal}, sample "{sample.label}".')
        print_progress(f'{group_title} {group.ordinal} integrated.')
    return groups  # Return the modified groups list for further use


def create_or_clear_sheet(xlsx_file, sheet_name, position=None):
    """Create a new sheet or clear it if it already exists."""
    # Check if the sheet already exists
    if sheet_name in xlsx_file.sheetnames:
        # Delete the existing sheet
        print_progress(f'Sheet "{sheet_name}" data wiped.')
        del xlsx_file[sheet_name]
    else:
        print_progress(f'New sheet "{sheet_name}" created.')
    # Create and return a new sheet
    sheet = xlsx_file.create_sheet(sheet_name)
    if position:
        sheets = xlsx_file._sheets
        sheets.insert(position, sheets.pop(sheets.index(sheet)))
    return sheet


def calculate_wellplate_statistics(wellplate_data):
    """Function to calculate the statistics of the 384-wellplate and its 96-well subplates."""
    # Constants for rows and columns
    ODD_ROW_CHARS = ['A', 'C', 'E', 'G', 'I', 'K', 'M', 'O']
    EVEN_ROW_CHARS = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    LEFT_NEG_CTRL_COL_NUMS = [1, 2]
    RIGHT_NEG_CTRL_COL_NUMS = [13, 14]
    LEFT_VIR_CTRL_COL_NUM = 12
    RIGHT_VIR_CTRL_COL_NUM = 24
    # List to store subplate statistics
    sub_wellplates_statistics = []
    # Lists to store actual absorbance values for the full wellplate
    full_wellplate_neg_ctrl_values = []
    full_wellplate_vir_ctrl_values = []
    # Loop through the subplates, generating the lists for negative control and viral control values
    for i in range(4): # Loop for 4x 96-well subplates
        # Lists to store actual absorbance values for the subplate
        sub_wellplate_neg_ctrl_values = []
        sub_wellplate_vir_ctrl_values = []
        # Determine which row characters and control columns to use for each subplate
        if i == 0: # First 96-well subplate (Odd rows, Left cols)
            row_chars = ODD_ROW_CHARS
            neg_ctrl_cols = LEFT_NEG_CTRL_COL_NUMS
            vir_ctrl_col = LEFT_VIR_CTRL_COL_NUM
        elif i == 1: # Second 96-well subplate (Even rows, Left cols)
            row_chars = EVEN_ROW_CHARS
            neg_ctrl_cols = LEFT_NEG_CTRL_COL_NUMS
            vir_ctrl_col = LEFT_VIR_CTRL_COL_NUM
        elif i == 2: # Third 96-well subplate (Odd rows, Right cols)
            row_chars = ODD_ROW_CHARS
            neg_ctrl_cols = RIGHT_NEG_CTRL_COL_NUMS
            vir_ctrl_col = RIGHT_VIR_CTRL_COL_NUM
        else: # Fourth 96-well subplate (Even rows, Right cols)
            row_chars = EVEN_ROW_CHARS
            neg_ctrl_cols = RIGHT_NEG_CTRL_COL_NUMS
            vir_ctrl_col = RIGHT_VIR_CTRL_COL_NUM
        # Loop through the rows for the current subplate
        for row_char in row_chars:
            # Collect negative control values
            for neg_ctrl_col in neg_ctrl_cols:
                value = wellplate_data.get((row_char, str(neg_ctrl_col).zfill(2)), None)
                if value:
                    sub_wellplate_neg_ctrl_values.append(value)
            # Collect viral control values
            value = wellplate_data.get((row_char, str(vir_ctrl_col).zfill(2)), None)
            if value:
                sub_wellplate_vir_ctrl_values.append(value)
        # Calculate and store statistics for the subplate
        subplate_stats = {
            'neg_ctrl_avg': mean(sub_wellplate_neg_ctrl_values) if sub_wellplate_neg_ctrl_values else "None",
            'neg_ctrl_avg_half': mean(sub_wellplate_neg_ctrl_values) / 2 if sub_wellplate_neg_ctrl_values else "None",
            'neg_ctrl_stdev': stdev(sub_wellplate_neg_ctrl_values) if 1 < len(sub_wellplate_neg_ctrl_values) else "None",
            'vir_ctrl_avg': mean(sub_wellplate_vir_ctrl_values) if sub_wellplate_vir_ctrl_values else "None",
            'vir_ctrl_stdev': stdev(sub_wellplate_vir_ctrl_values) if 1 < len(sub_wellplate_vir_ctrl_values) else "None"
        }
        # Add the subplate statistics onto the list
        sub_wellplates_statistics.append(subplate_stats)
        # Add subplate's negative and viral control values to fullplate's lists of values as well
        full_wellplate_neg_ctrl_values.extend(sub_wellplate_neg_ctrl_values)
        full_wellplate_vir_ctrl_values.extend(sub_wellplate_vir_ctrl_values)
    # Calculate and store statistics for the full wellplate
    full_wellplate_statistics = {
        'neg_ctrl_avg': mean(full_wellplate_neg_ctrl_values) if full_wellplate_neg_ctrl_values else "None",
        'neg_ctrl_avg_half': mean(full_wellplate_neg_ctrl_values) / 2 if full_wellplate_neg_ctrl_values else "None",    
        'neg_ctrl_stdev': stdev(full_wellplate_neg_ctrl_values) if 1 < len(full_wellplate_neg_ctrl_values) else "None",
        'vir_ctrl_avg': mean(full_wellplate_vir_ctrl_values) if full_wellplate_vir_ctrl_values else "None",
        'vir_ctrl_stdev': stdev(full_wellplate_vir_ctrl_values) if 1 < len(full_wellplate_vir_ctrl_values) else "None"
    }
    # Return both the fullplate statistics as well as the list of its subplate statistics
    return full_wellplate_statistics, sub_wellplates_statistics


def get_col_width(sheet, col_char, padding=2):
    """
    Calculate the width of a column based on the maximum length of the data in that column plus padding.

    Parameters:
    sheet (Worksheet): The worksheet containing the column.
    col_letter (str): The letter of the column to calculate the width for.
    padding (int): The amount of excess space to add to the maximum length. Default is 2.

    Returns:
    int: The calculated width of the column.
    """
    DEFAULT_WIDTH = 8.43
    # Get the lengths of all non-None values in the column
    lengths = [len(str(cell.value)) for cell in sheet[col_char] if cell.value is not None]
    # If lengths is empty, return the padding only
    if not lengths:
        return DEFAULT_WIDTH
    # Return the maximum length plus padding
    return max(lengths) + padding


def write_wellplate_statistics(statistics_sheet, full_wellplate_statistics, sub_wellplates_statistics, write_row, write_col):
    """Function to write the statistics of the wellplates to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to write the wellplate statistics into.
    full_wellplate_statistics (dict): The dictionary containing 384-wellplate statistics of averages and standard deviations.
    sub_wellplates_statistics (list): A List containing dictionaries of 96-well subplate statistics of averages and standard deviations.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write the mainplate header
    statistics_sheet.cell(row=write_row, column=write_col).value = "384-wellplate"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write plate's control averages and standard deviations
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).value = "Avg"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).value = "Avg50"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).value = "StDev"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col).value = "100%"
    statistics_sheet.cell(row=write_row + 2, column=write_col).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 2, column=write_col).alignment = Alignment(horizontal='right')
    statistics_sheet.cell(row=write_row + 3, column=write_col).value = "0%"
    statistics_sheet.cell(row=write_row + 3, column=write_col).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 3, column=write_col).alignment = Alignment(horizontal='right')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).value = full_wellplate_statistics['neg_ctrl_avg']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).value = full_wellplate_statistics['neg_ctrl_avg_half']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).value = full_wellplate_statistics['neg_ctrl_stdev']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).value = full_wellplate_statistics['vir_ctrl_avg']
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).value = full_wellplate_statistics['vir_ctrl_stdev']
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).number_format = "0.0#"
    for i, sub_wellplate_statistics in enumerate(sub_wellplates_statistics):
        subplate_row = write_row + ((i+1) * 5)
        # Write the subplate header
        statistics_sheet.cell(row=subplate_row, column=write_col).value = f"96-subplate ({i+1})"
        statistics_sheet.cell(row=subplate_row, column=write_col).font = Font(bold=True)
        # Write subplate's control averages and standard deviations
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 1).value = "Avg"
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 1).font = Font(italic=True)
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 1).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 2).value = "Avg50"
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 2).font = Font(italic=True)
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 2).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 3).value = "StDev"
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 3).font = Font(italic=True)
        statistics_sheet.cell(row=subplate_row + 1, column=write_col + 3).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 2, column=write_col).value = "100%"
        statistics_sheet.cell(row=subplate_row + 2, column=write_col).font = Font(italic=True)
        statistics_sheet.cell(row=subplate_row + 2, column=write_col).alignment = Alignment(horizontal='right',)
        statistics_sheet.cell(row=subplate_row + 3, column=write_col).value = "0%"
        statistics_sheet.cell(row=subplate_row + 3, column=write_col).font = Font(italic=True)
        statistics_sheet.cell(row=subplate_row + 3, column=write_col).alignment = Alignment(horizontal='right')
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 1).value = sub_wellplate_statistics['neg_ctrl_avg']
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 1).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 1).number_format = "0.0#"
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 2).value = sub_wellplate_statistics['neg_ctrl_avg_half']
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 2).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 2).number_format = "0.0#"
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 3).value = sub_wellplate_statistics['neg_ctrl_stdev']
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 3).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 2, column=write_col + 3).number_format = "0.0#"
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 1).value = sub_wellplate_statistics['vir_ctrl_avg']
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 1).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 1).number_format = "0.0#"
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 3).value = sub_wellplate_statistics['vir_ctrl_stdev']
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 3).alignment = Alignment(horizontal='center')
        statistics_sheet.cell(row=subplate_row + 3, column=write_col + 3).number_format = "0.0#"
    # Set the width of the column to be proportional to its content
    subplate_col_char = get_col_char(write_col)
    col_width = get_col_width(statistics_sheet, subplate_col_char)
    statistics_sheet.column_dimensions[subplate_col_char].width = col_width
    # Return the next row and column values
    return write_row + 5 + len(sub_wellplates_statistics) * 4, write_col


def write_statistics_data(statistics_sheet, group, group_statistics, write_row, write_col):
    """Function to write the statistics of a group to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to write the group statistics into.
    group (Group): The Group object containing the samples and parallels.
    group_statistics (dict): The dictionary containing group statistics of averages and standard deviations.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write the Group header
    statistics_sheet.cell(row=write_row, column=write_col).value = f"Group {group.ordinal}"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write the Group title
    statistics_sheet.cell(row=write_row, column=write_col + 1).value = f"{group.analysis_type}"
    statistics_sheet.cell(row=write_row, column=write_col + 1).font = Font(bold=True)
    statistics_sheet.cell(row=write_row, column=write_col + 1).alignment = Alignment(horizontal='right')
    # Write control averages and standard deviations
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).value = "Avg"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).value = "Avg50"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 2).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).value = "StDev"
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 1, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col).value = "100%"
    statistics_sheet.cell(row=write_row + 2, column=write_col).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 2, column=write_col).alignment = Alignment(horizontal='right')
    statistics_sheet.cell(row=write_row + 3, column=write_col).value = "0%"
    statistics_sheet.cell(row=write_row + 3, column=write_col).font = Font(italic=True)
    statistics_sheet.cell(row=write_row + 3, column=write_col).alignment = Alignment(horizontal='right')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).value = group_statistics['neg_ctrl_avg']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 1).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).value = group_statistics['neg_ctrl_avg_half']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 2).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).value = group_statistics['neg_ctrl_stdev']
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 2, column=write_col + 3).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).value = group_statistics['vir_ctrl_avg']
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 3, column=write_col + 1).number_format = "0.0#"
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).value = group_statistics['vir_ctrl_stdev']
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).alignment = Alignment(horizontal='center')
    statistics_sheet.cell(row=write_row + 3, column=write_col + 3).number_format = "0.0#"
    # Set the width of the column to be proportional to its content
    subplate_col_char = get_col_char(write_col)
    col_width = get_col_width(statistics_sheet, subplate_col_char)
    statistics_sheet.column_dimensions[subplate_col_char].width = col_width
    return write_row + 5, write_col + 1


def write_sample_data(statistics_sheet, sample_num, sample, write_row, write_col):
    """Function to write sample data to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet where the grouped data will be written.
    sample (Sample): The sample object containing parallels an wells.
    write_row (int): The starting row for writing the grouped data.
    write_col (int): The starting column for writing the grouped data.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write the Sample header
    statistics_sheet.cell(row=write_row, column=write_col).value = f"Sample {sample_num}"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write the sample label
    statistics_sheet.cell(row=write_row, column=write_col + 1).value = sample.label
    statistics_sheet.cell(row=write_row, column=write_col + 1).font = Font(bold=True)
    statistics_sheet.cell(row=write_row, column=write_col + 1).alignment = Alignment(horizontal='right')
    # Write absorbance values for each parallel
    for j, parallel in enumerate(sample.parallels):
        absorbance_row = write_row + 2 + j
        statistics_sheet.cell(row=absorbance_row, column=write_col).value = parallel.viral_control['row_char']
        for k, well in enumerate(parallel.get_wells()):
            if j == 0:  # Only write column numbers during the first parallel iteration
                statistics_sheet.cell(row=write_row + 1, column=write_col + 1 + k).value = well['col_num']
            statistics_sheet.cell(row=absorbance_row, column=write_col + 1 + k).value = well['value']
    # Track and update the writing row for the next sample
    write_row = absorbance_row
    # Return the next row and column values
    return write_row + 2, write_col


def write_basic_data(statistics_sheet, group, write_row, write_col):
    """Function to write the basic data of an group to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to write the group statistics into.
    group (Group): The Group object containing the samples and parallels.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write dil label
    statistics_sheet.cell(row=write_row, column=write_col).value = "Dil"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write dil values
    for i, dil in enumerate(group.dil_series):
        dil_row = write_row + i + 1
        statistics_sheet.cell(row=dil_row, column=write_col).value = dil
        statistics_sheet.cell(row=dil_row, column=write_col).alignment = Alignment(horizontal='left')
    parallel_count = len(group.samples[0].parallels)
    for i, sample in enumerate(group.samples):
        sample_col = write_col + (i * parallel_count) + 1
        # Write the Sample label
        statistics_sheet.merge_cells(start_row=write_row, start_column=sample_col, end_row=write_row, end_column=sample_col + parallel_count - 1)
        statistics_sheet.cell(row=write_row, column=sample_col).value = sample.label
        statistics_sheet.cell(row=write_row, column=sample_col).font = Font(bold=True)
        statistics_sheet.cell(row=write_row, column=sample_col).alignment = Alignment(horizontal='center')
        # Write the parallel values
        for j, parallel in enumerate(sample.parallels):
            parallel_row = write_row + 1
            parallel_col = sample_col
            for k, well in enumerate(parallel.get_virs()):
                value_row = parallel_row + k
                value_col = parallel_col + j
                statistics_sheet.cell(row=value_row, column=value_col).value = well["value"]
                # Align left-side values to the right, right-side values to the left
                if j < parallel_count // 2:
                    statistics_sheet.cell(row=value_row, column=value_col).alignment = Alignment(horizontal='right')
                else:
                    statistics_sheet.cell(row=value_row, column=value_col).alignment = Alignment(horizontal='left')
                # statistics_sheet.cell(row=value_row, column=value_col).alignment = Alignment(horizontal='center')
    return write_row + 12, write_col


def write_averages_data(statistics_sheet, normalized_group, write_row, write_col):
    """Function to write the averages data to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to write the averages data into.
    normalized_group (dict): The dictionary containing normalized group data.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    label_width = len(normalized_group["samples"])
    # Write averages label
    statistics_sheet.merge_cells(start_row=write_row, start_column=write_col, end_row=write_row, end_column=write_col+label_width)
    statistics_sheet.cell(row=write_row, column=write_col).value = "Averages"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write dil values
    statistics_sheet.cell(row=write_row + 1, column=write_col).value = "Dil"
    statistics_sheet.cell(row=write_row + 1, column=write_col).font = Font(bold=True)
    for i, dil in enumerate(normalized_group["dils"]):
        dil_row = write_row + 2 + i
        statistics_sheet.cell(row=dil_row, column=write_col).value = dil
        statistics_sheet.cell(row=dil_row, column=write_col).alignment = Alignment(horizontal='left')
    # Write avg data
    for i, sample in enumerate(normalized_group["samples"]):
        sample_row = write_row + 1
        sample_col = write_col + i + 1
        statistics_sheet.cell(row=sample_row, column=sample_col).value = sample["label"]
        statistics_sheet.cell(row=sample_row, column=sample_col).font = Font(bold=True)
        statistics_sheet.cell(row=sample_row, column=sample_col).alignment = Alignment(horizontal='right')
        for j, value in enumerate(sample["avgs"]):
            value_row = sample_row + j + 1
            statistics_sheet.cell(row=value_row, column=sample_col).value = value
            statistics_sheet.cell(row=value_row, column=sample_col).alignment = Alignment(horizontal='right')
            statistics_sheet.cell(row=value_row, column=sample_col).number_format = "0.0"
    # Return the next row and column to write into
    return write_row + WELLS_IN_A_PARALLEL + 1, write_col


def draw_graph(sheet, title, row_min, row_max, col_min, col_max, bump_up=False):
    """Function to draw a line graph using the first column as the X-axis and subsequent columns as Y-series.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to draw the graph into.
    write_row (int): The starting row to write the statistics.
    title (string): The title of the graph.
    row_min (int): The first row containing graph data.
    row_max (int): The last row containing graph data.
    col_min (int): The first col containing graph data.
    col_max (int): The last col containing graph data.
    bump_up (boolean): Decides if the graph is to be moved up by so much.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    graph = LineChart()
    # Define data range: all columns except the first one
    graphData = Reference(sheet, min_col=col_min+1, min_row=row_min, max_col=col_max, max_row=row_max)
    graph.add_data(graphData, titles_from_data=True)
    # Define X-axis categories (first column)
    categories = Reference(sheet, min_col=col_min, min_row=row_min+1, max_row=row_max)
    graph.set_categories(categories)
    # Calculate min and max values for the Y-axis from the data
    y_values = []
    for col in range(col_min+1, col_max+1):
        for row in range(row_min+1, row_max+1):  # Starting from row_min+1 (second row)
            cell_value = sheet.cell(row=row, column=col).value
            y_values.append(cell_value)
    y_min = min(y_values)
    y_max = max(y_values)
    y_range = y_max - y_min
    # Calculate min and max values for the X-axis (based on the first column data)
    x_values = [sheet.cell(row=row, column=col_min).value for row in range(row_min+1, row_max+1)]
    x_min = min(x_values)
    x_max = max(x_values)
    x_range = x_max - x_min
    # Set the Y-axis min, max, and step
    graph.y_axis.min = y_min - (0.01 * y_range)  # Add slight buffer to the min
    graph.y_axis.max = y_max + (0.01 * y_range)  # Add slight buffer to the max
    graph.y_axis.majorUnit = (y_range / 10)  # Set the step for the Y-axis
    # Set the X-axis min and max (no excess space around the data)
    graph.x_axis.min = x_min
    graph.x_axis.max = x_max
    graph.x_axis.majorUnit = (x_range / 10)  # Adjust the step for the X-axis
    # Style the datapoint markers
    for i, series in enumerate(graph.series):
        series.smooth = False
        series.graphicalProperties.line.solidFill = COLORS[i % len(COLORS)]
        series.graphicalProperties.line.width = CHART_LINE_WIDTH
        series.marker.symbol = MARKERS[i % len(MARKERS)]
        series.marker.size = MARKER_SIZE
        series.marker.graphicalProperties.line.solidFill = COLORS[i % len(COLORS)]
        series.marker.graphicalProperties.line.width = 0
        series.marker.graphicalProperties.solidFill = COLORS[i % len(COLORS)]
    # Graph titles and labels
    graph.title = title
    graph.y_axis.title = 'Inhibition (%)'
    graph.y_axis.number_format = '0.0#'
    graph.x_axis.title = 'Log dilution'
    graph.x_axis.number_format = '0.0#'
    # Chart styling
    graph.style = 13
    graph.width = 20
    graph.height = 13
    # Position the chart
    if bump_up:
        sheet.add_chart(graph, f"{sheet.cell(row=row_min-14, column=col_max+2).coordinate}")
    else:
        sheet.add_chart(graph, f"{sheet.cell(row=row_min, column=col_max+2).coordinate}")


def write_normalized_data(statistics_sheet, normalized_group, write_row, write_col):
    """Function to write the normalized data and draw graph to the statistics sheet.

    Parameters:
    statistics_sheet (Worksheet): The worksheet to write the normalized data and draw graph into.
    normalized_group (dict): The dictionary containing normalized group data.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    label_width = len(normalized_group["samples"])
    # Write normalized label
    statistics_sheet.merge_cells(start_row=write_row, start_column=write_col, end_row=write_row, end_column=write_col+label_width)
    statistics_sheet.cell(row=write_row, column=write_col).value = "Normalized values"
    statistics_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
   # Write log dil values
    statistics_sheet.cell(row=write_row + 1, column=write_col).value = "Log dil"
    statistics_sheet.cell(row=write_row + 1, column=write_col).font = Font(bold=True)
    for i, dil in enumerate(normalized_group["log_dils"]):
        dil_row = write_row + 2 + i
        statistics_sheet.cell(row=dil_row, column=write_col).value = dil
        statistics_sheet.cell(row=dil_row, column=write_col).alignment = Alignment(horizontal='left')
        statistics_sheet.cell(row=dil_row, column=write_col).number_format = "0.0##"
    # Write normalized data
    for i, sample in enumerate(normalized_group["samples"]):
        sample_row = write_row + 1
        sample_col = write_col + i + 1
        statistics_sheet.cell(row=sample_row, column=sample_col).value = sample["label"]
        statistics_sheet.cell(row=sample_row, column=sample_col).font = Font(bold=True)
        statistics_sheet.cell(row=sample_row, column=sample_col).alignment = Alignment(horizontal='right')
        for j, value in enumerate(sample["normalized_avgs"]):
            value_row = sample_row + j + 1
            statistics_sheet.cell(row=value_row, column=sample_col).value = value
            statistics_sheet.cell(row=value_row, column=sample_col).alignment = Alignment(horizontal='right')
            statistics_sheet.cell(row=value_row, column=sample_col).number_format = "0.0##"
    # Determine row/column range for graph
    row_min = write_row + 1
    row_max = write_row + len(normalized_group["log_dils"]) + 1  # Includes "Dil" row
    col_min = write_col
    col_max = write_col + label_width
    # Draw the graph on the right
    draw_graph(statistics_sheet, "Normalized values -chart", row_min, row_max, col_min, col_max, True)
    # Return the next row and column to write into
    return write_row + WELLS_IN_A_PARALLEL + 1, write_col


def run_statistics(statistics_sheet, group):
    """Function to calculate the statistics of an group and write them to the statistics sheet."""
    next_row, next_col = 1, 1
    # Get statistics for the group
    group_statistics = group.get_statistics()
    # Write group statistics to the statistics sheet
    next_row, next_col = write_statistics_data(statistics_sheet, group, group_statistics, next_row, next_col)
    # Write the data of each sample to the statistics sheet
    for i, sample in enumerate(group.samples):
        sample_num = i + 1
        next_row, next_col = write_sample_data(statistics_sheet, sample_num, sample, next_row, next_col)
    next_row += 1
    # Write basic data to the statistics sheet
    next_row, next_col = write_basic_data(statistics_sheet, group, next_row, next_col)
    # Get normalized data for the graphs
    normalized_group = group.get_normalized()
    # Write averages data to statistics sheet
    next_row, next_col = write_averages_data(statistics_sheet, normalized_group, next_row, next_col)
    # Write normalized data and draw graph to statistics sheet
    next_row, next_col = write_normalized_data(statistics_sheet, normalized_group, next_row, next_col)


def statistics_to_data_file(data_file, wellplate_data, groups):
    """Function to log the statistics of the wellplate as well as each of the groups to the data file."""
    # Get wellplate statistics
    full_wellplate_statistics, sub_wellplates_statistics = calculate_wellplate_statistics(wellplate_data)
    # Write wellplate statistics to the statistics sheet
    statistics_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_plate_STATISTICS")
    write_wellplate_statistics(statistics_sheet, full_wellplate_statistics, sub_wellplates_statistics, 1, 1)
    print_progress(f'Statistics of full 384-wellplate and its 96-subplates written to sheet "{statistics_sheet.title}".')
    for group in groups:
        # Skip filler datasets
        if group.is_filler:
            continue
        # Prepare new sheet for statistics
        statistics_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_group-{group.ordinal.replace(' ', '-')}_STATISTICS")
        # Run plate and group statistics
        run_statistics(statistics_sheet, group)
        print_progress(f'Statistics of Group {group.ordinal} written to sheet "{statistics_sheet.title}".')


def write_export_data(logs_sheet, normalized_group, write_row, write_col):
    """Function to write the normalized data and draw graph to the logs sheet.

    Parameters:
    logs_sheet (Worksheet): The worksheet to write the normalized data into.
    normalized_group (dict): The dictionary containing normalized group data.
    write_row (int): The starting row to write the statistics.
    write_col (int): The starting column to write the statistics.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    label_width = len(normalized_group["samples"])
    # Write normalized label
    logs_sheet.merge_cells(start_row=write_row, start_column=write_col, end_row=write_row, end_column=write_col+label_width)
    logs_sheet.cell(row=write_row, column=write_col).value = f'Group {normalized_group["ordinal"]} ({normalized_group["analysis_type"]})'
    logs_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
   # Write log dil values
    logs_sheet.cell(row=write_row + 1, column=write_col).value = "Log dil"
    logs_sheet.cell(row=write_row + 1, column=write_col).font = Font(bold=True)
    for i, dil in enumerate(normalized_group["log_dils"]):
        dil_row = write_row + 2 + i
        logs_sheet.cell(row=dil_row, column=write_col).value = dil
        logs_sheet.cell(row=dil_row, column=write_col).alignment = Alignment(horizontal='left')
        logs_sheet.cell(row=dil_row, column=write_col).number_format = "0.0##"
    # Write normalized data
    for i, sample in enumerate(normalized_group["samples"]):
        sample_row = write_row + 1
        sample_col = write_col + i + 1
        logs_sheet.cell(row=sample_row, column=sample_col).value = sample["label"]
        logs_sheet.cell(row=sample_row, column=sample_col).font = Font(bold=True)
        logs_sheet.cell(row=sample_row, column=sample_col).alignment = Alignment(horizontal='right')
        for j, value in enumerate(sample["normalized_avgs"]):
            value_row = sample_row + j + 1
            logs_sheet.cell(row=value_row, column=sample_col).value = value
            logs_sheet.cell(row=value_row, column=sample_col).alignment = Alignment(horizontal='right')
            logs_sheet.cell(row=value_row, column=sample_col).number_format = "0.0##"
    # Return the next row and column to write into
    return write_row + WELLS_IN_A_PARALLEL, write_col


def exports_to_data_file(data_file, groups):
    """Function to write the exportable data of each group to the export sheet of the data file."""
    # Prepare new sheet for exports
    export_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_group_EXPORT", 1)
    next_row, next_col = 1, 1
    for group in groups:
        # Skip filler datasets
        if group.is_filler:
            continue
        # Get normalized data
        normalized_group = group.get_normalized()
        # Write exportable data to export sheet
        next_row, next_col = write_export_data(export_sheet, normalized_group, next_row, next_col)


def save_file(file, file_name):
    """Save the Excel file to the specified path.

    Parameters:
    file (Workbook): The Excel workbook to be saved.
    file_name (str): The name of the file to save.
    """
    try:
        # Save data file to the 'data_files' directory
        file.save(f"./data_files/{file_name}")
    except PermissionError:
        # Handle permission error if the data file is open in another program
        exit_by_error(f'Unable to save data file "{file_name}". It might be opened in an Excel editor.')




########## ACTUAL SCRIPT ##########


def run_script():
    print_action("Evaluating environment")
    print_task(f'Checking script environment structure.')
    evaluate_script_environment()
    print_success("Script environment structure accepted.")
    # NECESSARY DIRECTORIES HAVE BEEN LOCATED #

    print_action("Configuring script")
    print_task(f'Fetching script configuration from "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    configure_script()
    print_success("Script configured.")
    # SCRIPT HAS BEEN CONFIGURED #

    print_action("Selecting data file")
    data_file_name = select_excel_file(DATA_FILES_DIRECTORY)
    print_task(f'Loading data file "{data_file_name}".')
    # DATA FILE HAS BEEN SELECTED #
    data_file, data_file_name = load_excel_file(DATA_FILES_DIRECTORY, data_file_name)
    print_success(f'Data file "{data_file_name}" loaded.')
    # DATA FILE HAS BEEN LOADED #

    print_action("Selecting data sheet")
    data_sheet_name = select_datasheet(data_file)
    print_task(f'Loading data sheet "{data_sheet_name}".')
    # DATA SHEET HAS BEEN SELECTED #
    data_sheet = load_datasheet(data_file, data_sheet_name)
    print_success(f'Data sheet "{data_sheet_name}" loaded.')
    # DATA SHEET HAS BEEN LOADED #

    print_action("Defining groups")
    groups = define_groups()
    print_success(f"Groups ({len(groups)}) have been defined.")
    # GROUPS HAVE BEEN CONFIGURED #

    print_action("Extracting wellplate data")
    print_task(f'Extracting wellplate data from data sheet "{data_sheet_name}".')
    wellplate_data = extract_wellplate_data(data_sheet, groups)
    print_success("Wellplate data extracted.")
    # DATA HAS BEEN EXTRACTED #

    print_action("Integrating wellplate data")
    print_task("Integrating wellplate data into groups.")
    groups = integrate_wellplate_data(wellplate_data, groups)
    print_success("Wellplate data integrated.")
    # DATA HAS BEEN INTEGRATED #    

    print_action("Calculating statistics")
    print_task(f'Calculating statistics and creating statistics sheets in data file "{data_file_name}".')
    statistics_to_data_file(data_file, wellplate_data, groups)
    print_success("Group statistics completed.")
    # STATISTICS CALCULATIONS COMPLETED #

    print_action("Writing exports")
    print_task(f'Writing exportable data into an export sheet in data file "{data_file_name}".')
    exports_to_data_file(data_file, groups)
    print_success("Exportable data written.")
    # EXPORTABLE DATA WRITTEN #

    print_action("Saving data file")
    print_task(f'Saving data file "{data_file_name}".')
    save_file(data_file, data_file_name)
    print_success(f'Data file "{data_file_name}" saved.')
    # DATA FILE HAS BEEN SAVED #




########## MAIN ENTRY POINT ##########


if __name__ == "__main__":
    handle_arguments(args)
    try:
        print_status("Started")
        run_script()
        print_status("Completed")
    except KeyboardInterrupt:
        exit_by_interruption()
