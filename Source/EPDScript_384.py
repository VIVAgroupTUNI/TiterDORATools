"""
384-Wellplate Data Processing Script

Summary:
This script is designed to process and analyze data from 384-wellplate experiments. The primary functionalities include
selecting and loading master and data files, defining tissues and their samples, extracting and validating wellplate data,
integrating this data into predefined structures, performing statistical analyses, and writing the results back to
Excel files.
The script starts by ensuring the necessary libraries are installed and then initializes global attributes related to
the experiment setup. It includes user interaction functions for input validation and output formatting, ensuring
a smooth user experience.
Interruptions are handled gracefully, and the script provides clear progress and error messages throughout its execution.

Usage:
1. Ensure you have the required directories (`config_files`, `master_files` and `data_files`) with appropriate files.
2. Run the script from the command line.
3. Follow the prompts to:
   - Select a master file.
   - Select a data file.
   - Choose a data sheet from the data file.
   - Define the tissues, including dilution rate, first dilution, working volume, sample count, and parallels per sample.
4. The script will then:
   - Extract wellplate data.
   - Integrate the data into the defined tissues.
   - Perform data analysis.
   - Run necessary calculations.
   - Write the results back into the selected files.
5. The updated data and master files will be saved with the new analysis and calculation results.
"""




########## IMPORTS ##########


import os
import time
import textwrap
import json
from json import JSONDecodeError
from sys import exit
from sys import argv as args
from statistics import stdev, mean
try:
    from xlrd import open_workbook                                  # type: ignore
except ImportError as e:
    exit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.\n')
try:
    from openpyxl import load_workbook as load                      # type: ignore
    from openpyxl.styles import Font, Alignment                     # type: ignore
    from openpyxl.utils import get_column_letter as get_col_char    # type: ignore
    from openpyxl.workbook import Workbook                          # type: ignore
except ImportError as e:
    exit('\n> Error: You are missing the library to read .xlsx files called "openpyxl". Run the installer to update your libraries.\n')




########## ATTRIBUTES ##########

SCRIPT_VERSION = "0.1.0"                                                                    # The public version number of the script

action_iterator = 1                                                                         # A global iterator used for tracking actions
DELAY = 0.05                                                                                # The delay (in seconds) used for timing outputs and creating a user-friendly experience
ENVIRONMENT = "SCR"                                                                         # The environment mode for the script, typically set to 'SCR' (short for scripted)
ALL_TISSUES = []                                                                            # Usable tissues that are read from the confg file
CONFIG_PARAMETER = "tissues"                                                                # Parameter that is configured in the config file
CONFIG_FILE_NAME = "EPD_config"                                                             # The name of the script's configure file
CONFIG_FILE_TYPE = ".json"                                                                  # The type of the script's configure file
CONFIG_FILES_DIRECTORY = "./config_files/"                                                  # The directory within which the script looks for its configuration (.json) files
MASTER_FILES_DIRECTORY = "./master_files/"                                                  # The directory within which the script looks for potential excel (.xls and.xlsx) master files.
DATA_FILES_DIRECTORY = "./data_files/"                                                      # The directory within which the script looks for potential excel (.xls and.xlsx) data files.
WELLPLATE_ROWS = 16                                                                         # The number of rows in the wellplate
WELLPLATE_COLS = 24                                                                         # The number of columns in the wellplate
WELLS_IN_A_PARALLEL = 12                                                                    # The number of wells in a parallel sample
PARALLEL_FIRST_COLS = [i for i in range(1, WELLPLATE_COLS + 1, WELLS_IN_A_PARALLEL)]        # A list of the wellplate columns where new parallel values start from
PARALLEL_CHOICES = [4, 8]                                                                   # Allowed numbers of parallels per sample
MAX_SAMPLES = int(round(len(PARALLEL_FIRST_COLS) * WELLPLATE_ROWS / PARALLEL_CHOICES[0]))   # Maximum amount of samples based on the size of the wellplate and allowed number of parallels




########## CLASSES ##########


class Tissue:
    """Class to manage tissue data, including samples and their associated configurations."""

    def __init__(self, tissue_title, ordinal):
        """Initializes a Tissue object with a title."""
        self.title = tissue_title
        self.ordinal = ordinal
        self.dil_rate = 0
        self.first_dil = 0
        self.working_vol = 0
        self.samples = []

    def __str__(self):
        """String representation of the Tissue object."""
        return f"Tissue Label: {self.title}{self.ordinal}, Samples: {len(self.samples)}"

    def add_sample(self, tube_label, parallels):
        """Adds a sample to the tissue."""
        new_sample = Sample(tube_label, parallels)
        self.samples.append(new_sample)


class Sample:
    """Class to manage sample data, including parallels and methods to perform calculations on them."""

    def __init__(self, tube_label, parallels):
        """Initializes a Sample object with a tube label and the specified number of parallels."""
        self.tube_label = tube_label
        self.parallels = [Parallel() for _ in range(parallels)]

    def __str__(self):
        """String representation of the Sample object."""
        return f"Tube Label: {self.tube_label}, Parallels: {len(self.parallels)}"
    
    def get_control_stdev(self):
        """Calculates the standard deviation of the control absorbance values across all parallels."""
        control_values = [control for parallel in self.parallels for control in parallel.get_controls()]
        return stdev(control_values) if 1 < len(control_values) else 0

    def get_control_avg(self):
        """Calculates the average of the control absorbance values across all parallels."""
        control_values = [control for parallel in self.parallels for control in parallel.get_controls()]
        return mean(control_values) if control_values else 0
    
    def get_analysis_results(self):
        """Performs analysis on the sample data and returns a summary of the results."""
        avg = self.get_control_avg()
        avg_50 = avg / 2
        dev = self.get_control_stdev()
        well_info = [parallel.get_well_info(avg_50) for parallel in self.parallels]
        infection_avgs = self.get_infection_avgs()
        infection_avgs_sum = self.get_infection_avgs_sum()
        return {
            "avg": avg,
            "avg_50": avg_50,
            "dev": dev,
            "well_info": well_info,
            "infection_avgs": infection_avgs,
            "infection_avgs_sum": infection_avgs_sum
        }
    
    def get_infection_avgs(self):
        """Calculates the infection averages for the sample."""
        avg_50 = self.get_control_avg() / 2
        well_infos = [parallel.get_well_info(avg_50) for parallel in self.parallels]
        # Initialize a list to store the sums of each well across all parallels
        sum_wells = [0] * WELLS_IN_A_PARALLEL
        count_wells = [0] * WELLS_IN_A_PARALLEL
        # Iterate through all parallels to sum the values of the same index wells
        for well_info in well_infos:
            for i, info in enumerate(well_info):
                sum_wells[i] += 0 if info["is_infected"] else 1
                count_wells[i] += 1
        # Calculate the average for each well
        infection_avgs = [sum_wells[i] / count_wells[i] for i in range(WELLS_IN_A_PARALLEL)]
        return infection_avgs

    def get_infection_avgs_sum(self):
        """Calculates the sum of infection averages, excluding the first 2 control values."""
        infection_avgs = self.get_infection_avgs()
        return sum(infection_avgs[2:])  # Exclude the first 2 control values

    def get_calc_results(self, dil_rate, first_dil, working_vol):
        """Performs calculations based on the infection averages and provided parameters."""
        sum = self.get_infection_avgs_sum()
        sum_minus_half = sum - 0.5 if 0 < sum - 0.5 else 0
        dil_rate_to_the_power_of = pow(dil_rate, sum_minus_half) if sum_minus_half != 0 else 0
        first_dil_multiplied = first_dil * dil_rate
        dil_rate_x_first_dil_multiplied = dil_rate_to_the_power_of * first_dil_multiplied if sum_minus_half != 0 else 0
        tcid = dil_rate_x_first_dil_multiplied * (1 / working_vol) if sum_minus_half != 0 else 0
        pfu = tcid * 0.69 if sum_minus_half != 0 else 0
        return {
            "sum": sum,
            "sum_minus_half": sum_minus_half,
            "dil_rate_to_the_power_of": dil_rate_to_the_power_of,
            "first_dil_multiplied": first_dil_multiplied,
            "dil_rate_x_first_dil_multiplied": dil_rate_x_first_dil_multiplied,
            "tcid": tcid,
            "pfu": pfu
        }
    
    def get_epd(self, dil_rate, first_dil, working_vol):
        """Calculates the EPD (Effective Plating Dose) for the sample."""
        calc_results = self.get_calc_results(dil_rate, first_dil, working_vol)
        return round(calc_results["tcid"], 2)


class Parallel:
    """Class to store and manage absorbance data for control and viral wells within a wellplate."""

    def __init__(self):
        """Initializes a Parallel object with empty lists for controls and virals."""
        self.controls = []
        self.virals = []

    def __str__(self):
        """Returns a string representation of the Parallel object."""
        return f"Controls: {self.controls}, Virals: {self.virals}"

    def add_control(self, absorbance, row_char, col_num):
        """Adds a control absorbance value along with its row and column information."""
        self.controls.append({
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        })

    def add_viral(self, absorbance, row_char, col_num):
        """Adds a viral absorbance value along with its row and column information."""
        self.virals.append({
            "value": absorbance,
            "row_char": row_char,
            "col_num": col_num
        })

    def get_controls(self):
        """Retrieves a list of control absorbance values."""
        return [control['value'] for control in self.controls]

    def get_well_info(self, avg50):
        """Retrieves detailed information for all wells, including infection status."""
        # Collect well info for controls
        well_info = [{
            "value": control['value'],
            "row_char": control['row_char'],
            "col_num": control['col_num'],
            "is_infected": avg50 < control['value'],
        } for control in self.controls]
        # Collect well info for virals
        well_info.extend([{
            "value": viral['value'],
            "row_char": viral['row_char'],
            "col_num": viral['col_num'],
            "is_infected": avg50 < viral['value']
        } for viral in self.virals])
        return well_info
    



########## ARGUMENT HANDLING ##########


def info():
    """Function to print out instructions on how to use the script for the user."""
    try:
        # Print tutorial started status
        print_status("Tutorial Started", tailing_line_break=True)
        # Print each section of the tutorial with detailed instructions
        print_info("1) Evaluating Environment:", [
            "The script begins by checking the structure of the required directories.",
            "It checks for the existence of 'config_files', 'master_files' and 'data_files' directories.",
            "It ensures that each directory contains valid files (like .json, .xls and .xlsx)."
        ])
        print_info("2) Configuring Script:", [
            "Next the script will look for its configuration file 'EPD_config.json' from the 'config_files' directory.",
            "It will load and apply its parameters into the script.",
            "It ensures that the config file can be found and its parameters are valid."
        ])
        print_info("3) Selecting Master File:", [
            "After that, the user will be prompted to select a master file from the 'master_files' directory.",
            "The available excel files will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("4) Selecting Data File:", [
            "The user will then be prompted to select a data file from the 'data_files' directory.",
            "The available excel files will be listed, and the user will again choose one by its corresponding number."
        ])
        print_info("5) Selecting Data Sheet:", [
            "The user will select a sheet from the chosen data file.",
            "The available sheets will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("6) Defining Tissues:", [
            "The user will define the tissues to be analysed.",
            "For each tissue, the user will input the dilution rate, first dilution, working volume, total samples, and parallels per sample.",
            "The user will also provide the tube labels for each sample."
        ])
        print_info("7) Extracting Wellplate Data:", [
            "The script will then read the selected data sheet and extract the wellplate data.",
            f"It will look for the {WELLPLATE_ROWS}x{WELLPLATE_COLS} wellplate segment, starting from the appropriate row and column, and collect the absorbance values."
        ])
        print_info("8) Integrating Wellplate Data:", [
            "The script will integrate the extracted wellplate data into the defined tissues.",
            "Each sample and parallel within the tissues will be populated with the corresponding absorbance values from the wellplate data."
        ])
        print_info("9) Analyzing Data:", [
            "The script will analyze the integrated data by calculating averages, standard deviations, and infection statuses for each sample.",
            "It will create new sheets in the data file to store the analysis results for each tissue."
        ])
        print_info("10) Running Calculations:", [
            "The script will perform further calculations using the analyzed data.",
            "It will calculate EPD results for each sample and create additional sheets in the data file to store the calculation results for each tissue."
        ])
        print_info("11) Saving Data File:", [
            "The script will save the updated data file with the new analysis and calculation sheets to ensure that the intermediate results are preserved."
        ])
        print_info("12) Writing Results:", [
            "The script will then proceed to write the calculated EPD results into the selected master file.",
            "It will locate the appropriate rows and columns for each tissue and sample, and write the EPD results accordingly."
        ])
        print_info("13) Updating Measurement Averages:", [
            "If there are already three measurements recorded for the same sample in the master file, the script will calculate the average of these measurements and update the corresponding cell in the master file."
        ])
        print_info("14) Saving Master File:", [
            "Finally, the script will save the updated master file, ensuring that all changes are preserved."
        ])
        # Print additional notes
        print_info("Notes:", [
            "The script may be interrupted at any point of its execution safely by hitting (CTRL+C).",
            "The user must provide the working volume in milliliters (mL).",
            "Whenever the script prompts the user to provide numeric data that involves decimals, it allows the user to use either Finnish commas (,) or English dots (.) as decimal separators.",
            "Should the config file be lost, the script will be able to restore its template as long as the config directory exists.",
            "Should the user pick a .xls file as either master or data file, the script will automatically conver it into a .xlsx file and replace the old .xls file altogether.",
            "Should the user provide incorrect values for a script run, they must manually delete the incorrectly written EPD result values from the master file before running the script again.",
            "The script looks for config files within the './config_files/' directory.",
            "The script looks for master files within the './master_files/' directory.",
            "The script looks for data files within the './data_files/' directory.",
            "The master file must have its results' main label row start from column A and there must not be any cells with content in column A before the label row.",
            "The label for tissues column must contains the word 'tissue' in any form.",
            "The label for tube label column has to contain the words 'tube' and 'label in any form'.",
            "The label for TCID average column must contain the words 'tcid' and 'average' in any form.",
            "The label for each EPD result column must contain the words 'tcid' and 'result' in any form.",
        ])
        # Print explanation for each type of user prompt that the script uses
        print_info("User Prompts:", [
            "Input Prompt: When the script requires user input, it will display a prompt with a '>' symbol (e.g., \"Please select the master file > \"). The user will enter their response after the '>' symbol.",
            "Choices: Whenever the script presents multiple options to choose from, they will be listed with a corresponding command in square brackets (e.g., \"[1] master_file_1.xlsx\"). The user will select an option by typing its command.",
            "Task: The script will print task status messages to indicate what it is currently doing. These messages will start with \"> Task: \" followed by the task description.",
            "Progress: During longer operations, the script will print progress messages to keep the user informed. These messages will start with \"> Progress: \" followed by the progress description.",
            "Success: Upon successful completion of an operation, the script will print success messages starting with \"> Success: \" followed by a brief description of the success.",
            "Error: If an error occurs during an operation, the script will print error messages starting with \"> Error: \" followed by the error description."
        ])
        print_status("Tutorial completed", leading_line_break=False)
    except KeyboardInterrupt:
        exit_by_interruption(True)


def handle_arguments(args):
    """Handles the command line arguments provided to the script."""
    # Check if any arguments were given at launch
    if 1 < len(args):
        param = args[1].lower()
        if param in ["i", "info", "h", "help"]:
            # If the first argument was either "info" or "help", start tutorial
            info()
            exit()




########## EXIT ROUTES ##########


def exit_by_error(message):
    """Exits the script with a given error message."""
    try:
        time.sleep(DELAY)
        print_error(message)
        time.sleep(DELAY)
        print_status("Failed")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")


def exit_by_interruption(during_info=False):
    """Exits the script due to an interruption (e.g., KeyboardInterrupt, CTRL+C)."""
    try:
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        if during_info:
            print_status("Tutorial stopped")    
        else:
            print_status("Stopped")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")




########## USER INTERACTION ##########


def prompt_input(prompt_str):
    """Gets user input after displaying a prompt."""
    try:
        time.sleep(DELAY)
        user_input = input(f"{prompt_str} > ").strip()
    except KeyboardInterrupt:
        exit_by_interruption()
    return user_input


def get_user_input(prompt, validation_func=None, type_select=True):
    """Gets and validates user input.

    Parameters:
    prompt (str): The prompt message to display.
    validation_func (function): The function to validate user input. Default is None.
    type_select (bool): Determines if input is for selection (True) or entry (False). Default is True.

    Returns:
    str: The validated user input.
    """
    # Print a blank line if type_select is True
    if type_select:
        print()
    while True:
        # Get user input
        user_input = prompt_input(prompt)
        # If no validation function is provided, return the input
        if not validation_func:
            if type_select:
                print()
            return user_input
        # Validate the user input
        validation_result, validated_input = validation_func(user_input)
        # If validation is successful, return the validated input
        if validation_result is True:
            if type_select:
                print()
            return validated_input
        # Print error message if validation fails
        print_error(validated_input)




########## PRINT FUNCTIONS ##########


def print_status(status_string, leading_line_break=True, tailing_line_break=False):
    """Prints a status message with version in the top border."""
    try:
        time.sleep(DELAY)
        text = f"EPD-384 SCRIPT {status_string.upper()}"
        # Middle line padding
        asterisks = "*" * 3
        spaces = " " * 5
        middle_line = f"{asterisks}{spaces}{text}{spaces}{asterisks}"
        # Top border with version
        top_border_text = f" v{SCRIPT_VERSION} "
        total_width = len(middle_line)
        top_line = "*" * ((total_width - len(top_border_text)) // 2) + top_border_text
        if len(top_line) < total_width:
            top_line += "*" * (total_width - len(top_line))
        bottom_line = "*" * len(middle_line)

        if leading_line_break:
            time.sleep(DELAY)
            print()
        time.sleep(DELAY)
        print(top_line)
        time.sleep(DELAY)
        print(middle_line)
        time.sleep(DELAY)
        print(bottom_line)
        if tailing_line_break or status_string.strip().lower() == "completed":
            time.sleep(DELAY)
            print()
    except KeyboardInterrupt:
        exit_by_interruption()


def print_action(action_str):
    """Prints a separator line for actions."""
    try:
        global action_iterator
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        print(f"========== Action {action_iterator}: {action_str.title()} ==========")
        action_iterator += 1
    except KeyboardInterrupt:
        exit_by_interruption()


def print_task(task_str):
    """Prints a task message."""
    try:
        time.sleep(DELAY)
        print(f"> Task: {task_str}..")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_choice(choice_cmd, choice_str):
    """Prints a choice option."""
    try:
        time.sleep(DELAY)
        print(f"[{choice_cmd}] {choice_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_progress(progress_str):
    """Prints a progress message."""
    try:
        time.sleep(DELAY)
        print(f"> Progress: {progress_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_success(success_str):
    """Prints a success message."""
    try:
        time.sleep(DELAY)
        print(f"> Success: {success_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_error(error_str):
    """Prints an error message."""
    try:
        time.sleep(DELAY)
        print(f"> Error: {error_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_info(header, messages, max_length=100):
    """Prints information with a delay and text wrapping."""
    try:
        INDENT = 3
        time.sleep(DELAY)
        print(header.title())
        for message in messages:
            wrapped_message = textwrap.fill(f" - {message}", max_length, subsequent_indent=' ' * INDENT)
            time.sleep(DELAY)
            print(wrapped_message)
        time.sleep(DELAY)
        print()
    except KeyboardInterrupt:
        exit_by_interruption(True)




########## VALIDATION FUNCTIONS ##########


def validate_digit(input_str):
    """Validate if the input string is a digit.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    if input_str.isdigit():
        return True, int(input_str)
    return False, f'Your input of "{input_str}" was not a whole number.'


def validate_float(input_str):
    """Validate if the input string is a float, allowing for both Finnish (,) and English (.) decimal separators.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the float value if valid, or an error message if not.
    """
    try:
        input_str = input_str.replace(',', '.')
        float_value = float(input_str)
        return True, float_value
    except ValueError:
        return False, f'Your input "{input_str}" was not a number.'


def validate_non_empty(input_str):
    """Validate if the input string is not empty.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the input string if valid, or an error message if not.
    """
    if input_str.strip():
        return True, input_str
    return False, f"Your input was empty."


def validate_min_max(input_str, min_choice, max_choice, escape_str=None):
    """Validate if the input string is a digit within a specified range, with an optional escape string.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer or 0 if escaped, or an error message if not.
    """
    if escape_str is not None and input_str.strip() == escape_str:
        return True, 0
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if min_choice <= validation_result <= max_choice:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [{min_choice}, {max_choice}].'


def validate_yes_no(input_str):
    """Validate if the input string is a 'yes' or 'no' response.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated input string if valid, or an error message if not.
    """
    input_str = input_str.lower().strip()
    if input_str in ['y', 'yes', 'n', 'no']:
        return True, input_str
    return False, f'Your input "{input_str}" was invalid. Choose either yes or no.'


def validate_sample_count(input_str):
    """Validate if the input string is a digit within the range of [1, 32].

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer if valid, or an error message if not.
    """
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if 1 <= validation_result <= MAX_SAMPLES:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [1, {MAX_SAMPLES}].'


def validate_parallel_count(input_str):
    """Validate if the input string is a digit and is either 4 or 8.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer if valid, or an error message if not.
    """
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if validation_result in PARALLEL_CHOICES:
        return True, validation_result
    return False, f'Your input "{input_str}" was invalid. Choose either {" or ".join(map(str, PARALLEL_CHOICES))}.'


def validate_row_char(row_char):
    """Check if the row character is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated row character if valid, or an error message if not.
    """
    if row_char is None:
        return False, "Row character was missing."
    if isinstance(row_char, str):
        row_char_cap = row_char.upper()
        if len(row_char_cap) == 1 and row_char_cap.isalpha():
            return True, row_char_cap
    return False, f"Row character \"{row_char}\" was invalid. Expected a single character string like 'A' or 'a'."


def validate_start_row_char(row_char):
    """Validate the starting row character 'A'.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated row character if valid, or an error message if not.
    """
    is_valid, validation_result = validate_row_char(row_char)
    if not is_valid:
        return False, validation_result
    if validation_result == 'A':
        return True, validation_result
    return False, f"Starting row character \"{row_char}\" was invalid. Expected either 'A' or 'a'."


def validate_col_num(col_num):
    """Check if the column number value is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated column number as a zero-padded string if valid, or an error message if not.
    """
    if col_num is None:
        return False, "Column number was missing."
    col_num_str = str(col_num).zfill(2)
    if col_num_str.isdigit() and 1 <= int(col_num_str) <= 24:
        return True, col_num_str
    return False, f'Column number "{col_num}" was invalid. Expected a whole number in a format of "1" or "01".'


def validate_start_col_num(col_num):
    """Validate the starting column number '1'.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated column number as a zero-padded string if valid, or an error message if not.
    """
    is_valid, validation_result = validate_col_num(col_num)
    if not is_valid:
        return False, validation_result
    if validation_result == '01':
        return True, validation_result
    return False, f'Starting column number "{col_num}" was invalid. Expected either "1" or "01".'


def validate_absorbance(absorbance, row_char, col_num):
    """Check if the absorbance value is valid.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the absorbance value if valid, or an error message if not.
    """
    if absorbance is None:
        return True, absorbance  # None is acceptable for absorbance
    if isinstance(absorbance, (int, float)):
        return True, absorbance
    return False, f'Absorbance value "{absorbance}" of well {row_char}{col_num} was invalid. Expected either a number or an empty value.'




########## SEARCH FUNCTIONS ##########


def find_wellplate_start(sheet):
    """Function to find the starting row and column for the wellplate data.
    
    Returns:
    tuple: A tuple containing the starting row and column.
    """
    START_COL = 1
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value == 'A':
            start_row = row - 1
            # Verify the row character 'A'
            row_char = sheet.cell(row=start_row + 1, column=START_COL).value
            is_valid, validation_result = validate_start_row_char(row_char)
            if not is_valid:
                continue
            # Verify the column number 1 starting from the row above the 'A' row and one column to the left
            col_num = sheet.cell(row=start_row, column=START_COL + 1).value
            is_valid, validation_result = validate_start_col_num(col_num)
            if not is_valid:
                continue
            # Verify the empty cell exists between the row character 'A' and column number 1
            if sheet.cell(row=start_row, column=START_COL).value:
                continue
            return start_row, START_COL
    exit_by_error(f'Could not locate wellplate in data sheet "{sheet.title}".')


def find_main_labels_row(master_sheet):
    """Function to find the first row in column 1 of the sheet that contains content."""
    # Iterate over each row in the sheet
    for row in range(1, master_sheet.max_row + 1):
        cell = master_sheet.cell(row=row, column=1)
        # If the cell has content, return success and the row index
        if cell.value:
            return row
    return None


def find_label_col(master_sheet, label_row, col_identifiers):
    """Function to find a desired label column in the main label row based on identifiers.

    Parameters:
    label_row (int): The row containing the column labels.
    col_identifiers (list): List of strings that should be present in the column header.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the column index if found, or an error message if not.
    """
    # Iterate through all of the cells in main label row
    for col in range(1, master_sheet.max_column + 1):
        cell = master_sheet.cell(row=label_row, column=col)
        cell_value = str(cell.value).strip().lower() if cell.value else ""
        # Should the cell's content satisfy all the column identifiers, return the column
        if all(identifier in cell_value for identifier in col_identifiers):
            return col
    return None


def find_tube_label_col(master_sheet, label_row):
    """Function to find the tube label column in the main label row based on identifiers.

    Parameters:
    label_row (int): The row containing the column labels.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the column index if found, or an error message if not.
    """
    TUBE_LABEL_COL_IDENTIFIERS = ["tube", "label"]
    # Find the tube label column
    tube_label_col = find_label_col(master_sheet, label_row, TUBE_LABEL_COL_IDENTIFIERS)
    if not tube_label_col:
        return False, "Could not locate the tube label column."
    return True, tube_label_col


def find_tissue_col(master_sheet, label_row):
    """Function to find the tissue column in the main label row based on identifiers.

    Parameters:
    label_row (int): The row containing the column labels.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the column index if found, or an error message if not.
    """
    TISSUE_COL_IDENTIFIERS = ["tissue"]
    # Find the tissue column
    tissue_col = find_label_col(master_sheet, label_row, TISSUE_COL_IDENTIFIERS)
    if not tissue_col:
        return False, "Could not locate the tissue column."
    return True, tissue_col


def find_tcid_avg_col(master_sheet, label_row):
    """Function to find the tcid average column in the master sheet based on identifiers.

    Parameters:
    label_row (int): The row containing the column labels.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the column index if found, or an error message if not.
    """
    COL_IDENTIFIERS = ["tcid", "average"]
    # Find the TCID average column
    tcid_avg_col = find_label_col(master_sheet, label_row, COL_IDENTIFIERS)
    if not tcid_avg_col:
        return False, "Could not locate the TCID average column."
    return True, tcid_avg_col


def find_tube_match_row(master_sheet, tube_label, tissue_title):
    """Function to find the row for a given tube label and tissue title in the master sheet.
    
    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the row index if found, or an error message if not.
    """
    # Find the main labels row in the master sheet
    label_row = find_main_labels_row(master_sheet)
    if not label_row:
        exit_by_error("No label row found.")
    first_data_row = label_row + 1
    # Find the tube label column in the labels row
    was_found, search_result = find_tube_label_col(master_sheet, label_row)
    if not was_found:
        return False, search_result
    tube_label_col = search_result
    # Find the tissue column in the labels row
    was_found, search_result = find_tissue_col(master_sheet, label_row)
    if not was_found:
        return False, search_result
    tissue_col = search_result
    # Find a match for the tube and tissue based on the found columns
    tube_label_match_row = None
    tissue_match_row = None
    # Iterate over the rows in the master sheet starting from the first data row
    for row in range(first_data_row, master_sheet.max_row + 1):
        # Get the tube label and tissue title values from the respective columns
        tube_cell = master_sheet.cell(row=row, column=tube_label_col)
        tissue_cell = master_sheet.cell(row=row, column=tissue_col)
        tube_cell_value = str(tube_cell.value).strip().lower() if tube_cell.value else ''
        tissue_cell_value = str(tissue_cell.value).strip().lower() if tissue_cell.value else ''
        # Check if the tube label matches
        if tube_cell_value == tube_label.strip().lower():
            tube_label_match_row = row
            # Check if the tissue title is contained in the tissue cell value
            if tissue_title.strip().lower() in tissue_cell_value:
                tissue_match_row = row
                return True, row
    # Return appropriate results messages based on what was found
    if not tube_label_match_row:
        return False, f'Tube label "{tube_label}" was not found.'
    if not tissue_match_row:
        return False, f'Tissue title "{tissue_title}" was not found for tube label "{tube_label}".'
    return False, "Unknown error during tube match search."  # Fallback case


def find_tcid_cols(master_sheet, label_row):
    """Function to find all TCID columns in the master sheet based on identifiers.

    Parameters:
    label_row (int): The row containing the column labels.

    Returns:
    list: A list of column indices matching the TCID identifiers.
    """
    COL_IDENTIFIERS = ["tcid", "result"]
    tcid_cols = []
    # Search for TCID columns in the main label row
    for col in range(1, master_sheet.max_column + 1):
        cell = master_sheet.cell(row=label_row, column=col)
        cell_value = str(cell.value).strip().lower() if cell.value else ""
        # Should the cell's content satisfy all column identifiers, the column is picked
        if all(identifier in cell_value for identifier in COL_IDENTIFIERS):
            tcid_cols.append(col)
    return tcid_cols


def find_free_tcid_col(master_sheet, label_row, data_row):
    """Function to find the next empty TCID column for a given row in the master sheet.

    Returns:
    tuple: (was_found, search_result) where was_found is a boolean indicating success,
           and search_result is the column number if found, or an error message if not.
    """
    # Find all TCID columns
    tcid_cols = find_tcid_cols(master_sheet, label_row)
    if not tcid_cols:
        return False, "No TCID columns found."
    # Iterate through the TCID columns to find the next empty column
    for col in tcid_cols:
        data_cell = master_sheet.cell(row=data_row, column=col)
        if not data_cell.value:  # If the cell is empty, return the column
            return True, col
    return False, f"All found TCID columns ({len(tcid_cols)}) already contained values."




########## FILE CONVERTER ##########


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Function to convert deprecated .xls files into .xlsx files."""
    # Open the .xls file using xlrd
    workbook_xls = open_workbook(xls_path)
    workbook_xlsx = Workbook()
    for sheet_index in range(workbook_xls.nsheets):
        # One sheet at a time
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        if sheet_index == 0:
            # Activate the workbook to create its first sheet
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
        else:
            # Add them otherwise
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
        # One row at a time
        for row_index in range(sheet_xls.nrows):
            for col_index in range(sheet_xls.ncols):
                # Write each cell
                sheet_xlsx.cell(row=row_index + 1, column=col_index + 1).value = sheet_xls.cell_value(row_index, col_index)
    try:
        # Save the new .xlsx file using openpyxl
        workbook_xlsx.save(xlsx_path)
        # Delete the old .xls file using xlrds
        os.remove(xls_path)
    except PermissionError:
        # If couldn't remove the original file, remove the new one to avoid duplicates
        os.remove(xlsx_path)
        exit_by_error(f'File "{xls_path}" was already open in another program. Please close it and try again.')




########## CORE FUNCTIONS ##########


def directory_contains_file(directory, filename):
    """Check if a specific file exists in the given directory."""
    return filename in os.listdir(directory)


def get_excel_files(directory):
    """Function to find all excel files (.xls and .xlsx) in the given directory that do not start with '~$'."""
    return [file for file in os.listdir(directory) if (file.endswith('.xlsx') or file.endswith('.xls')) and not file.startswith('~$')]


def generate_mock_template():
    """Function to create mock tissues instance."""
    return ["Tissue 1",
            "Tissue 2",
            "Tissue 3"]


def create_config_template():
    """Function to create a template configuration file."""
    # Ensuring the config directory still exists.
    os.makedirs(CONFIG_FILES_DIRECTORY, exist_ok=True)
    print_task(f'Creating config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" in config directory.')
    # Determine the configuration file path
    config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    # Create the configuration file
    with open(config_file_path, "w", encoding="utf-8") as file:
        # Add template content
        mock_tissues = generate_mock_template()
        json.dump({CONFIG_PARAMETER: mock_tissues}, file, indent=4)
    print_success("Config file created, terminating script.")


def evaluate_script_environment():
    """Check if the required directories with acceptable content exist in the current working directory."""
    # Check for the existence of config files directory
    if not os.path.isdir(CONFIG_FILES_DIRECTORY):
        exit_by_error(f'Config files directory "{CONFIG_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Config files directory "{CONFIG_FILES_DIRECTORY}" located.')
    # Check if config files directory contains the scirpt's config file
    config_file = directory_contains_file(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
    if not config_file:
        # In case the config file cannot be located
        print_error(f'Config files directory didn\'t contain file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
        # Ask the user if the missing config file should be restored
        yes_no_value = get_user_input('The script will be terminated...\nRestore the missing config file template on exit? (yes/no)', validate_yes_no)
        if yes_no_value in ['y', 'yes']:
            # Restore the config file template
            create_config_template()
            # Stop the script
            exit_by_error()
        else:
            # Terminate the script
            exit_by_error("Terminating script.")
    print_progress(f'Config files directory contained file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    # Check for the existence of master files directory
    if not os.path.isdir(MASTER_FILES_DIRECTORY):
        exit_by_error(f'Master files directory "{MASTER_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Master files directory "{MASTER_FILES_DIRECTORY}" located.')
    # Check if master files directory contains at least one excel file that doesn't start with "~$"
    master_files = get_excel_files(MASTER_FILES_DIRECTORY)
    if not master_files:
        exit_by_error(f'Master files directory did not contain any excel files.')
    print_progress(f'Master files directory contained excel files.')
    # Check for the existence of data files directory
    if not os.path.isdir(DATA_FILES_DIRECTORY):
        exit_by_error(f'Data files directory "{DATA_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Data files directory "{DATA_FILES_DIRECTORY}" located.')
    # Check if data files directory contains at least one excel file that doesn't start with "~$"
    data_files = get_excel_files(DATA_FILES_DIRECTORY)
    if not data_files:
        exit_by_error(f'Data files directory did not contain any excel files.')
    print_progress(f'Data files directory contained excel files.')


def select_excel_file(directory):
    """Selects an Excel file from the specified directory."""
    # List all Excel files in the directory, excluding temporary files
    try:
        excel_files = get_excel_files(directory)
    except FileNotFoundError:
        exit_by_error(f'Directory "{directory}" could not be located.')
    # If no Excel files are found, exit the script with an error message
    if not excel_files:
        exit_by_error(f'Directory "{directory}" did not contain any excel files.')
    # Print choices for the user to select from
    for i, file in enumerate(excel_files):
        print_choice(i+1, file)
    # User chooses a file
    int_value = get_user_input(f'Please select the file', lambda input_str: validate_min_max(input_str, 1, len(excel_files)))
    return excel_files[int_value - 1]


def load_excel_file(directory, file_name):
    """Load an Excel file from the given path and return the workbook object."""
    try:
        # Convert .xls to .xlsx if necessary
        if file_name.endswith('.xls'):
            xls_path = directory + file_name
            file_name += 'x'
            xlsx_path = directory + file_name
            convert_xls_to_xlsx(xls_path, xlsx_path)
            print_progress("Converted the old .xls file into a new .xlsx file.")
        else:
            xlsx_path = directory + file_name
        # Load the Excel file
        file = load(xlsx_path)
        # Save the unchanged file to check if it is accessible
        file.save(xlsx_path)
        return file, file_name
    except FileNotFoundError:
        # Handle file not found error
        exit_by_error(f'File "{file_name}" could not be located.')
    except PermissionError:
        # Handle permission error if the file is already open
        exit_by_error(f'File "{file_name}" was already open in another program. Please close it and try again.')
    except Exception as e:
        # Handle any other exceptions that occur during file loading
        exit_by_error(f'Failed loading file "{file_name}".  > REASON: {e}.')


def load_configuration():
    """Attempt to Load configration from the configuration file."""
    try:
        # Determine the configuration file path
        config_file_path = os.path.join(CONFIG_FILES_DIRECTORY, f'{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}')
        with open(config_file_path, "r", encoding="utf-8") as file:
            # Load from the configuration file
            config = json.load(file)
            # Return the loaded configuration
            return config
    except FileNotFoundError:
        # Configuration file could not be found
        exit_by_error(f'Config file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be found.')
    except JSONDecodeError:
        # Configuration could not be loaded
        exit_by_error(f'Configuration from file "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}" could not be loaded.')
    

def apply_configuration(config):
    """Check that the configuration is valid and apply it."""
    # Fetch "analyses" from configuration
    tissues = config.get(CONFIG_PARAMETER, None)
    # Check that configuration is valid
    if tissues is None:
        # Configuration didn't contain "analyses"
        exit_by_error(f'Loaded configuration didn\'t contain "{CONFIG_PARAMETER}".')
    if not isinstance(tissues, list):
        # Configuration "analyses" wasn't a list
        exit_by_error(f'Configured "{CONFIG_PARAMETER}" wasn\'t a list.')
    if len(tissues) < 1:
        # Configuration "analyses" was empty
        exit_by_error(f'Configured "{CONFIG_PARAMETER}" didn\'t contain items')
    if not all(isinstance(item, str) for item in tissues):
        # Configuration "analysis" contained invalid item types like for example integers
        exit_by_error(f'Configured "{CONFIG_PARAMETER}" contained items of other type than string.')
    else:
        # Configuration was valid, add configured analyses to the list of choices
        ALL_TISSUES[:0] = tissues


def configure_script():
    """Initialize the script by loading and applying a configuration."""
    # Load the configuration
    config = load_configuration()
    print_progress("Configuration loaded.")
    # Apply the configuration
    apply_configuration(config)
    print_progress("Configuration applied.")


def select_datasheet(data_file):
    """Prompt the user to select a data sheet from the given data file."""
    # List all sheet names in the data file
    datasheets = [sheet for sheet in data_file.sheetnames]
    for i, sheet in enumerate(datasheets):
        print_choice(i+1, sheet)
    # User chooses a sheet
    int_value = get_user_input('Please select the sheet', lambda input_str: validate_min_max(input_str, 1, len(data_file.sheetnames)))
    return datasheets[int_value - 1]


def load_datasheet(data_file, data_sheet_name):
    """Load the specified data sheet from the given data file."""
    return data_file[data_sheet_name]


def define_tissues():
    """Define the tissues to be analyzed by collecting user inputs for each tissue's details."""
    tissues = []
    tissue_titles = sorted(ALL_TISSUES)
    tissue_counts = {tissue_title: 0 for tissue_title in tissue_titles}  # Initialize counter for each tissue type
    while True:
        # Display available tissue titles for selection
        for i, tissue_title in enumerate(tissue_titles):
            print_choice(i+1, tissue_title)
        # Get user input for the selected tissue type
        user_input = get_user_input("Please select the tissue type", lambda input_str: validate_min_max(input_str, 1, len(tissue_titles)))
        tissue_title = tissue_titles[int(user_input) - 1]
         # Increment the count for this tissue type
        tissue_counts[tissue_title] += 1
        ordinal = tissue_counts[tissue_title]
         # Create a new Tissue object and collect its details
        tissue = Tissue(tissue_title, ordinal)
        tissue.dil_rate = get_user_input(f"Please enter the dilution rate ({tissue_title}{ordinal})", validate_float, False)
        tissue.first_dil = get_user_input(f"Please enter the first dilution ({tissue_title}{ordinal})", validate_float, False)
        tissue.working_vol = get_user_input(f"Please enter the working volume ({tissue_title}{ordinal})", validate_float, False)
        sample_count = get_user_input(f"Please enter the sample count ({tissue_title}{ordinal})", validate_sample_count, False)
        parallel_count = get_user_input(f"Please enter the parallel count ({tissue_title}{ordinal})", validate_parallel_count, False)
        # Collect sample details for the tissue
        for i in range(sample_count):
            tube_label = get_user_input(f"Please enter tube label for sample {i+1} ({tissue_title}{ordinal})", validate_non_empty, False)
            tissue.add_sample(tube_label, parallel_count)
        # Add the defined tissue to the list of tissues
        tissues.append(tissue)
        print_progress(f'Tissue "{tissue_title}{ordinal}" defined.')
        # Ask if the user wants to define another tissue
        yes_no_value = get_user_input('Would you like to add another tissue? (yes/no)', validate_yes_no)
        if yes_no_value in ['n', 'no']:
            break
    return tissues


def extract_wellplate(data_sheet, start_row, start_col):
    """Get a segment of the worksheet based on the starting point, number of rows, and number of columns.

    Parameters:
    data_sheet (Worksheet): The sheet to get data from.
    start_row (int): The starting row of the segment.
    start_col (int): The starting column of the segment.

    Returns:
    list: A list of lists representing the segment.
    """
    # Determine the coordinate of the first cell in the segment
    first_cell = data_sheet.cell(row=start_row, column=start_col).coordinate
    # Determine the coordinate of the last cell in the segment
    last_cell = data_sheet.cell(row=start_row + WELLPLATE_ROWS, column=start_col + WELLPLATE_COLS).coordinate
    # Return the segment of the worksheet as a list of lists
    return data_sheet[first_cell:last_cell]


def get_wellplate_data(wellplate, tissues):
    """Creates a dictionary of absorbance values from the wellplate segment.

    Returns:
    dict: A dictionary with (row_char, col_num) as keys and absorbance values as values.
    """
    # Determine required data counts
    parallel_count = sum(len(sample.parallels) for tissue in tissues for sample in tissue.samples)
    required_total = parallel_count * WELLS_IN_A_PARALLEL
    realized_total = 0
    # Create the dictionary to hold the absorbance values
    wellplate_data = {}
    # Extract the data as new key (row character, column number) value (absorbance) pairs of the dictionary
    for i, row in enumerate(wellplate):
        if i == 0:
            continue  # Skip the first row which contains headers
        row_char = wellplate[i][0].value
        # Validate the row character
        is_valid, validation_result = validate_row_char(row_char)
        if not is_valid:
            exit_by_error(validation_result)
        row_char = validation_result
        for j, well in enumerate(row):
            if j == 0:
                continue  # Skip the first column which contains row headers
            col_num = wellplate[0][j].value
            # Validate the column number
            is_valid, validation_result = validate_col_num(col_num)
            if not is_valid:
                exit_by_error(validation_result)
            col_num = validation_result
            absorbance = well.value
            # Validate the absorbance value
            is_valid, validation_result = validate_absorbance(absorbance, row_char, int(col_num))
            if not is_valid:
                exit_by_error(validation_result)
            absorbance = validation_result
            # Add the valid absorbance value to the dictionary
            wellplate_data[(row_char, col_num)] = absorbance
            if absorbance:
                # If the well contains a value, increment the total found absorbance values counter
                realized_total += 1
    if realized_total < required_total:
        # If there weren't enough data in the wellplate to cover the requirements, terminate the script
        exit_by_error(f"Excess data provided: wellplate contains ({realized_total}), but you configured ({required_total}). Ensure that all data points are correctly configured.")
    elif required_total < realized_total:
        # If there was more data on the wellplate than required to cover the analyses, terminate the script
        exit_by_error(f"Insufficient data provided: wellplate contains ({realized_total}), but you configured ({required_total}). Ensure that all data points are correctly configured.")
    return wellplate_data


def extract_wellplate_data(data_sheet, tissues):
    """Extracts wellplate data from the specified data sheet."""
    # Find the starting row and column of the wellplate
    START_ROW, START_COL = find_wellplate_start(data_sheet)
    print_progress(f"Wellplate located in segment [{chr(64+START_COL)}{START_ROW}, {chr(64+START_COL+WELLPLATE_COLS)}{START_ROW+WELLPLATE_ROWS}].")
    # Extract the wellplate segment based on the starting point
    wellplate = extract_wellplate(data_sheet, START_ROW, START_COL)
    # Get the absorbance values from the wellplate segment
    wellplate_data = get_wellplate_data(wellplate, tissues)
    print_progress("Wellplate data was valid.")
    return wellplate_data


def get_first_key(wellplate_data):
    """Find the first key in the wellplate data that contains a numeric value.

    Parameters:
    wellplate_data (dict): Dictionary with (row_char, col_num) as keys and absorbance values as values.

    Returns:
    tuple: The first (row_char, col_num_padded) key that contains a numeric value, or None if no valid key is found.
    """
    # Initialize the starting key
    current_key = ('A', '01')
    for _ in range(WELLPLATE_ROWS * WELLPLATE_COLS):
        value = wellplate_data.get(current_key)
        if value is not None:
            return current_key
        current_key = get_next_key(current_key)
    return None  # Return None if no valid key is found


def get_next_key(key):
    """Update the row character and column index as per the required logic and return the full key.

    Parameters:
    current_key (tuple): The current (row_char, col_num) key.

    Returns:
    tuple: Updated (row_char, col_num_padded) key.
    """
    ASCII_A = ord('A')
    ASCII_B = ord('B')
    ASCII_JUMP = ASCII_A + WELLPLATE_ROWS
    ASCII_RESET = ASCII_B + WELLPLATE_ROWS
    # Retrieve the current key values
    row_char, col_num = key
    row_char_ascii = ord(row_char)
    col_num_int = int(col_num)
    # Determine the parallel segment columns
    segment_start = max([start_col for start_col in PARALLEL_FIRST_COLS if start_col <= col_num_int])
    segment_end = segment_start + WELLS_IN_A_PARALLEL
    # Determine the next column number
    next_col_num_int = col_num_int + 1
    # By default the next row character will remain the same unless deemed otherwise
    next_row_char = row_char
    # Check if the column number has reached the end of the current segment
    if segment_end <= next_col_num_int:
        # Reset the col number
        next_col_num_int = segment_start
        # Determine the next row character
        next_row_char_ascii = row_char_ascii + 2
        if ASCII_RESET <= next_row_char_ascii:
            # Jump back to the start after exhausting both odd and even rows
            next_row_char_ascii = ASCII_A
            next_col_num_int = segment_end
            if segment_start == PARALLEL_FIRST_COLS[-1]:
                return None # If the next key is outside the bounds of the wellplate, return None
        elif ASCII_JUMP <= next_row_char_ascii: 
            # Jump to the even row characters if the odd rows are exhausted
            next_row_char_ascii = ASCII_B
        next_row_char = chr(next_row_char_ascii)
    next_col_num = str(next_col_num_int).zfill(2)
    next_key = (next_row_char, next_col_num)
    return next_key


def integrate_wellplate_data(wellplate_data, tissues):
    """Integrates the collected wellplate data into the appropriate tissues."""
    current_key = get_first_key(wellplate_data)  # Find the first key with valid values
    if not current_key:
        # If the first key could not be determined, the wellplate didn't contain any valid absorbance data
        exit_by_error("Couldn't find any absorbance values in the wellplate.")
    for tissue in tissues:
        for sample in tissue.samples:
            for parallel in sample.parallels:
                for well_nbr in range(WELLS_IN_A_PARALLEL):
                    if not current_key:
                        # If the current key is None, the wellplate ran out of data and the user input or data was incorrect
                        exit_by_error(f'The wellplate did not contain enough absorbance data to integrate tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}".')
                    if current_key not in wellplate_data:
                        # If the wellplate didn't contain the key, the wellplate in the data file was faulty
                        exit_by_error(f"Couldn't find an absorbance value in well {current_key}.")
                    absorbance = wellplate_data[current_key]
                    # Add absorbance to controls or virals based on well number
                    if well_nbr < 2:
                        parallel.add_control(absorbance, current_key[0], current_key[1])
                    else:
                        parallel.add_viral(absorbance, current_key[0], current_key[1])
                    # Update to the next key
                    current_key = get_next_key(current_key)
            print_progress(f'Integrated data for tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}".')
        print_progress(f'Tissue "{tissue.title}{tissue.ordinal}" integrated.')
    return tissues  # Return the modified tissues list for further use


def create_or_clear_sheet(xlsx_file, sheet_name):
    """Create a new sheet or clear it if it already exists."""
    # Check if the sheet already exists
    if sheet_name in xlsx_file.sheetnames:
        # Delete the existing sheet
        print_progress(f'Sheet "{sheet_name}" data wiped.')
        del xlsx_file[sheet_name]
    else:
        print_progress(f'New sheet "{sheet_name}" created.')
    # Create and return a new sheet
    return xlsx_file.create_sheet(sheet_name)


def get_col_width(sheet, col_char, padding=2):
    """
    Calculate the width of a column based on the maximum length of the data in that column plus padding.

    Parameters:
    sheet (Worksheet): The worksheet containing the column.
    col_letter (str): The letter of the column to calculate the width for.
    padding (int): The amount of excess space to add to the maximum length. Default is 2.

    Returns:
    int: The calculated width of the column.
    """
    DEFAULT_WIDTH = 8.43
    # Get the lengths of all non-None values in the column
    lengths = [len(str(cell.value)) for cell in sheet[col_char] if cell.value is not None]
    # If lengths is empty, return the padding only
    if not lengths:
        return DEFAULT_WIDTH
    # Return the maximum length plus padding
    return max(lengths) + padding


def write_sample_analysis(analysis_sheet, sample_num, sample, analysis_results, write_row, write_col):
    """Function to write the analysis results of a sample to the analysis sheet.

    Parameters:
    analysis_sheet (Worksheet): The worksheet to write the analysis results into.
    sample (Sample): The sample object containing the tube label and parallels.
    analysis_results (dict): The dictionary containing analysis results such as averages and infection status.
    write_row (int): The starting row to write the results.
    write_col (int): The starting column to write the results.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Set the width of the column to be proportional to its content
    write_col_char = get_col_char(write_col)
    col_width = get_col_width(analysis_sheet, write_col_char)
    analysis_sheet.column_dimensions[write_col_char].width = col_width
    # Write the Sample number
    analysis_sheet.cell(row=write_row, column=write_col).value = f"Sample {sample_num}"
    analysis_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    # Write the sample tube label
    analysis_sheet.cell(row=write_row + 1, column=write_col).value = "Tube label"
    analysis_sheet.cell(row=write_row + 1, column=write_col + 1).value = sample.tube_label
    analysis_sheet.cell(row=write_row + 1, column=write_col + 1).font = Font(bold=True)
    analysis_sheet.cell(row=write_row + 1, column=write_col + 1).alignment = Alignment(horizontal='right')
    # Write control averages and standard deviations
    analysis_sheet.cell(row=write_row + 2, column=write_col).value = "Avg"
    analysis_sheet.cell(row=write_row + 2, column=write_col + 1).value = analysis_results['avg']
    analysis_sheet.cell(row=write_row + 3, column=write_col).value = "Avg-50%"
    analysis_sheet.cell(row=write_row + 3, column=write_col + 1).value = analysis_results['avg_50']
    analysis_sheet.cell(row=write_row + 4, column=write_col).value = "StDev"
    analysis_sheet.cell(row=write_row + 4, column=write_col + 1).value = analysis_results['dev']
    # Write absorbance values and infection status for each parallel
    for i, parallel in enumerate(sample.parallels):
        absorbance_row = write_row + 6 + i
        infection_row = write_row + 6 + len(sample.parallels) + i
        analysis_sheet.cell(row=absorbance_row, column=write_col).value = parallel.controls[0]['row_char']
        for j, well in enumerate(analysis_results["well_info"][i]):
            if i == 0:  # Only write column numbers during the first parallel iteration
                analysis_sheet.cell(row=write_row + 5, column=write_col + 1 + j).value = well['col_num']
            analysis_sheet.cell(row=absorbance_row, column=write_col + 1 + j).value = well['value']
            analysis_sheet.cell(row=infection_row, column=write_col + 1 + j).value = well['is_infected']
    # Write infection averages
    avg_row_number = write_row + 6 + len(sample.parallels) * 2
    for i, avg in enumerate(analysis_results['infection_avgs']):
        analysis_sheet.cell(row=avg_row_number, column=write_col + 1 + i).value = avg
    # Write the sum of averages at the end of the row
    analysis_sheet.cell(row=avg_row_number, column=write_col + 1 + WELLS_IN_A_PARALLEL).value = analysis_results['infection_avgs_sum']
    # Return the next column and row values
    return avg_row_number + 3, write_col


def run_analysis(analysis_sheet, tissue):
    """Function to run the analysis for each sample in a tissue and write the results to the analysis sheet."""
    next_row = 2
    next_col = 2
    for i, sample in enumerate(tissue.samples):
        # Get analysis results for the sample
        analysis_results = sample.get_analysis_results()
        # Write sample data to the analysis sheet
        next_row, next_col = write_sample_analysis(analysis_sheet, i+1, sample, analysis_results, next_row, next_col)


def analysis_to_data_file(data_file, tissues):
    """Function to log the analysis results to the data file."""
    for tissue in tissues:
        # Prepare new sheets for each tissue's analysis
        analysis_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_{tissue.title.strip().lower().replace(' ', '')}{tissue.ordinal}_ANALYSIS")
        # Run analysis for the current tissue
        run_analysis(analysis_sheet, tissue)
        print_progress(f'Analysis results of tissue "{tissue.title}{tissue.ordinal}" written to sheet "{analysis_sheet.title}".')


def write_calculation_labels(calc_sheet, samples, write_row, write_col):
    """Function to write the calculation labels to the calculation sheet.

    Parameters:
    calc_sheet (Worksheet): The worksheet where the labels will be written.
    samples (list): The list of samples to create labels for.
    write_row (int): The starting row for writing the labels.
    write_col (int): The starting column for writing the labels.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write sample number labels
    for i in range(len(samples)):
        calc_sheet.cell(row=write_row, column=write_col + i + 1).value = f"Sample {i + 1}"
        calc_sheet.cell(row=write_row, column=write_col + i + 1).alignment = Alignment(horizontal='right')
        sample_col_char = get_col_char(write_col + i + 1)
        col_width = get_col_width(calc_sheet, sample_col_char, 10)
        calc_sheet.column_dimensions[sample_col_char].width = col_width
    # Write headers for the calculation results
    calc_sheet.cell(row=write_row + 1, column=write_col).value = "Tube Label"
    calc_sheet.cell(row=write_row + 2, column=write_col).value = "SUM"
    calc_sheet.cell(row=write_row + 3, column=write_col).value = "SUM-0.5"
    calc_sheet.cell(row=write_row + 4, column=write_col).value = "Dil rate to the power of"
    calc_sheet.cell(row=write_row + 5, column=write_col).value = "1st dil (dil in tube * dil rate)"
    calc_sheet.cell(row=write_row + 6, column=write_col).value = "Dil rate x 1st dil (TCID/wv)"
    calc_sheet.cell(row=write_row + 7, column=write_col).value = "TCID/ml"
    calc_sheet.cell(row=write_row + 8, column=write_col).value = "PFU/ml"
    # Set the width of the column to be proportional to its content
    write_col_char = get_col_char(write_col)
    col_width = get_col_width(calc_sheet, write_col_char)
    calc_sheet.column_dimensions[write_col_char].width = col_width
    # Return the next row and column for subsequent writing
    return write_row + 1, write_col + 1


def write_sample_calculations(calc_sheet, sample, calc_results, write_row, write_col):
    """Function to write sample calculations to the calculation sheet.

    Parameters:
    calc_sheet (Worksheet): The worksheet where the calculations will be written.
    sample (Sample): The sample object containing tube label and calculation results.
    calc_results (dict): The dictionary containing calculated results for the sample.
    write_row (int): The starting row for writing the calculations.
    write_col (int): The starting column for writing the calculations.

    Returns:
    tuple: The next row and column values to continue writing.
    """
    # Write the tube label of the sample
    calc_sheet.cell(row=write_row + 0, column=write_col).value = sample.tube_label
    calc_sheet.cell(row=write_row + 0, column=write_col).alignment = Alignment(horizontal='right')
    # Write the calculated results to the sheet
    calc_sheet.cell(row=write_row + 1, column=write_col).value = calc_results["sum"]
    calc_sheet.cell(row=write_row + 2, column=write_col).value = calc_results["sum_minus_half"]
    calc_sheet.cell(row=write_row + 3, column=write_col).value = calc_results["dil_rate_to_the_power_of"]
    calc_sheet.cell(row=write_row + 4, column=write_col).value = calc_results["first_dil_multiplied"]
    calc_sheet.cell(row=write_row + 5, column=write_col).value = calc_results["dil_rate_x_first_dil_multiplied"]
    calc_sheet.cell(row=write_row + 6, column=write_col).value = calc_results["tcid"]
    calc_sheet.cell(row=write_row + 6, column=write_col).font = Font(italic=True, bold=True)
    calc_sheet.cell(row=write_row + 7, column=write_col).value = calc_results["pfu"]
    # Return the next row and column for subsequent writing
    return write_row, write_col + 1


def run_calculations(calc_sheet, tissue):
    """Function to perform calculations for a given tissue and write the results to a calculation sheet."""
    next_row = 1
    next_col = 1
    # Write sample numbers and labels
    next_row, next_col = write_calculation_labels(calc_sheet, tissue.samples, next_row, next_col)
    for sample in tissue.samples:
        # Get calculation results for the sample
        calc_results = sample.get_calc_results(tissue.dil_rate, tissue.first_dil, tissue.working_vol)
        # Write the sample calculations to the sheet and update the next row and column
        next_row, next_col = write_sample_calculations(calc_sheet, sample, calc_results, next_row, next_col)


def calculations_to_data_file(data_file, tissues):
    """Function to log the calculation results to the data file."""
    for tissue in tissues:
        # Prepare new sheets for each tissue's calculations
        calc_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_{tissue.title.strip().lower().replace(' ', '')}{tissue.ordinal}_CALC")
        # Run calculations for the current tissue
        run_calculations(calc_sheet, tissue)
        print_progress(f'Calculation results of tissue "{tissue.title}{tissue.ordinal}" written to sheet "{calc_sheet.title}".')


def results_to_master_file(master_sheet, tissues):
    """Function to log the EPD results to the master file."""
    # Find the main labels row in the master sheet
    label_row = find_main_labels_row(master_sheet)
    if not label_row:
        exit_by_error("No label row found.")
    # Write the EPD result of each sample of each tissue separately
    for tissue in tissues:
        for sample in tissue.samples:
            # Get the EPD result of the sample
            epd_value = sample.get_epd(tissue.dil_rate, tissue.first_dil, tissue.working_vol)
            # Find the row for the tube
            was_found, search_result = find_tube_match_row(master_sheet, sample.tube_label, tissue.title)
            if not was_found:
                print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" result ({epd_value}) FAILED to write.  > REASON: {search_result}')
                continue
            tube_match_row = search_result
            # Find the next empty TCID column
            was_found, search_result = find_free_tcid_col(master_sheet, label_row, tube_match_row)
            if not was_found:
                print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" result ({epd_value}) FAILED to write.  > REASON: {search_result}')
                continue
            free_tcid_col = search_result
            # Write the EPD result in the next empty TCID column for the matched row
            cell_to_write = master_sheet.cell(row=tube_match_row, column=free_tcid_col)
            cell_to_write.value = epd_value
            print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" result ({epd_value}) SUCCEEDED to write into cell "{cell_to_write.coordinate}".')


def get_tcid_values(master_sheet, data_row):
    """Retrieves TCID values for a specified data row in the master sheet."""
    # Find the main label row in the master sheet
    label_row = find_main_labels_row(master_sheet)
    if not label_row:
        exit_by_error("No label row found.")
    tcid_values = []
    # Find the TCID columns based on the label row
    tcid_cols = find_tcid_cols(master_sheet, label_row)
    for col in tcid_cols:
        cell_value = master_sheet.cell(row=data_row, column=col).value
        if not cell_value:
            continue
        try:
            # Convert cell value to float and add to the TCID values list
            tcid_value = float(cell_value)
            tcid_values.append(tcid_value)
        except ValueError:
            continue
    return tcid_values


def calculate_tcid_average(master_sheet, label_row, data_row):
    """Calculate the TCID average for a given row in the master sheet.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success, and the second value is 
           either the average value or an error message.
    """
    # Find all TCID values in the specified row
    tcid_cols = find_tcid_cols(master_sheet, label_row)
    if not tcid_cols:
        return False, "No TCID columns found."
    tcid_values = get_tcid_values(master_sheet, data_row)
    tcid_sum = sum(tcid_values)
    tcid_count = len(tcid_values)
    # Check if there are any valid TCID values
    if tcid_count < 1:
        return False, "No valid TCID values found."
    # Calculate the average of the TCID values
    tcid_average = tcid_sum / tcid_count
    return True, tcid_average


def averages_to_master_file(master_sheet, tissues):
    """Update the master file with the TCID averages for each sample."""
    # Find the main labels row in the master sheet
    label_row = find_main_labels_row(master_sheet)
    if not label_row:
        exit_by_error("No label row found.")
    # Find the tcid average column from master file
    was_found, search_result = find_tcid_avg_col(master_sheet, label_row)
    if not was_found:
        print_progress(f"Measurement average updates have all FAILED.  > REASON: {search_result}")
        return
    tcid_avg_col = search_result
    # Update the measurement average of each sample of each tissue separately
    for tissue in tissues:
        for sample in tissue.samples:
            # Find the row for the tube
            was_found, search_result = find_tube_match_row(master_sheet, sample.tube_label, tissue.title)
            if not was_found:
                print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" average FAILED to update.  > REASON: {search_result}')
                continue
            tube_match_row = search_result
            # Calculate the average TCID value
            was_found, search_result = calculate_tcid_average(master_sheet, label_row, tube_match_row)
            if not was_found:
                print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" average FAILED to update.  > REASON: {search_result}')
                continue
            average_value = round(search_result, 2)
            # Write the average TCID value in the appropriate column
            cell_to_write = master_sheet.cell(row=tube_match_row, column=tcid_avg_col)
            cell_to_write.value = average_value
            print_progress(f'Tissue "{tissue.title}{tissue.ordinal}", sample "{sample.tube_label}" average ({average_value}) SUCCEEDED to update in cell "{cell_to_write.coordinate}".')


def save_file(file, file_name, is_master=False):
    """Save the Excel file to the specified path.

    Parameters:
    file (Workbook): The Excel workbook to be saved.
    file_name (str): The name of the file to save.
    is_master (bool): Flag indicating if the file is a master file (default is False).
    """
    try:
        # Save master file to the 'master_files' directory
        if is_master:
            file.save(f"./master_files/{file_name}")
        # Save data file to the 'data_files' directory
        else:
            file.save(f"./data_files/{file_name}")
    except PermissionError:
        # Handle permission error if the master file is open in another program
        if is_master:
            exit_by_error(f'Unable to save master file "{file_name}". It might be opened in an Excel editor.')
        # Handle permission error if the data file is open in another program
        else:
            exit_by_error(f'Unable to save data file "{file_name}". It might be opened in an Excel editor.')




########## ACTUAL SCRIPT ##########


def run_script():
    print_action("Evaluating environment")
    print_task(f'Checking script environment structure.')
    evaluate_script_environment()
    print_success("Script environment structure accepted.")
    # NECESSARY DIRECTORIES HAVE BEEN LOCATED #

    print_action("Configuring script")
    print_task(f'Fetching script configuration from "{CONFIG_FILE_NAME}{CONFIG_FILE_TYPE}".')
    configure_script()
    print_success("Script configured.")
    # SCRIPT HAS BEEN CONFIGURED #

    print_action("Selecting master file")
    master_file_name = select_excel_file(MASTER_FILES_DIRECTORY)
    print_task(f'Loading master file "{master_file_name}".')
    # MASTER FILE HAS BEEN SELECTED #
    master_file, master_file_name = load_excel_file(MASTER_FILES_DIRECTORY, master_file_name)
    master_sheet = master_file.worksheets[0]
    # MASTER FILE HAS BEEN LOADED #
    print_success(f'Master file "{master_file_name}" loaded.')
    # MASTER SHEET HAS BEEN SELECTED #

    print_action("Selecting data file")
    data_file_name = select_excel_file(DATA_FILES_DIRECTORY)
    print_task(f'Loading data file "{data_file_name}".')
    # DATA FILE HAS BEEN SELECTED #
    data_file, data_file_name = load_excel_file(DATA_FILES_DIRECTORY, data_file_name)
    print_success(f'Data file "{data_file_name}" loaded.')
    # DATA FILE HAS BEEN LOADED #

    print_action("Selecting data sheet")
    data_sheet_name = select_datasheet(data_file)
    print_task(f'Loading data sheet "{data_sheet_name}".')
    # DATA SHEET HAS BEEN SELECTED #
    data_sheet = load_datasheet(data_file, data_sheet_name)
    print_success(f'Data sheet "{data_sheet_name}" loaded.')
    # DATA SHEET HAS BEEN LOADED #

    print_action("Defining tissues")
    tissues = define_tissues()
    print_success(f"Tissues ({len(tissues)}) have been defined.")
    # TISSUES HAVE BEEN DEFINED #

    print_action("Extracting wellplate data")
    print_task(f'Extracting wellplate data from data sheet "{data_sheet_name}".')
    wellplate_data = extract_wellplate_data(data_sheet, tissues)
    print_success("Wellplate data extracted.")
    # DATA HAS BEEN EXTRACTED #

    print_action("Integrating wellplate data")
    print_task("Integrating wellplate data into tissues.")
    tissues = integrate_wellplate_data(wellplate_data, tissues)
    print_success("Wellplate data integrated.")
    # DATA HAS BEEN INTEGRATED #

    print_action("Analyzing data")
    print_task(f'Analyzing data and creating analysis sheets in data file "{data_file_name}".')
    analysis_to_data_file(data_file, tissues)
    print_success("Data analysis completed.")
    # ANALYSIS HAS BEEN COMPLETED #

    print_action("Running calculations")
    print_task(f'Running calculations and creating calculation sheets in data file "{data_file_name}".')
    calculations_to_data_file(data_file, tissues)
    print_success("Calculations completed.")
    # CALCULATIONS HAVE BEEN MADE #

    print_action("Saving data file")
    print_task(f'Saving data file "{data_file_name}".')
    save_file(data_file, data_file_name)
    print_success(f'Data file "{data_file_name}" saved.')
    # DATA FILE HAS BEEN SAVED #

    print_action("Writing results")
    print_task(f'Writing EPD results in master file "{master_file_name}".')
    results_to_master_file(master_sheet, tissues)
    print_success("Writing results completed.")
    # RESULTS HAVE BEEN WRITTEN #

    print_action("Updating measurement averages")
    print_task(f'Calculating EPD averages in master file "{master_file_name}".')
    averages_to_master_file(master_sheet, tissues)
    print_success("Updating averages completed.")
    # UPDATING MEASUREMENT AVERAGES HAS COMPLETED #

    print_action("Saving master file")
    print_task(f'Saving master file "{data_file_name}".')
    save_file(master_file, master_file_name, True)
    print_success(f'Master file "{master_file_name}" saved.')
    # MASTER FILE HAS BEEN SAVED #




########## MAIN ENTRY POINT ##########


if __name__ == "__main__":
    handle_arguments(args)
    try:
        print_status("Started")
        run_script()
        print_status("Completed")
    except KeyboardInterrupt:
        exit_by_interruption()
