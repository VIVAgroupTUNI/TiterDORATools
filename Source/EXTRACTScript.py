"""
96-Wellplate Data Processing Script

Summary:
This script is designed to process and analyze data from excel files containing 96-wellplate experiments. The primary
functionalities include selecting and loading excel files, extracting and validating wellplate data, separating the
wellplates into excel files of their own, and zipping the created excel files into a zip container.
The script starts by ensuring the necessary libraries are installed and then initializes global attributes related to
the script setup. It includes user interaction functions for input validation and output formatting, ensuring
a smooth user experience.
Interruptions are handled gracefully, and the script provides clear progress and error messages throughout its execution.

Usage:
1. Ensure you have the required directories (`data_files` and `result_files`) with appropriate Excel files.
2. Run the script from the command line.
3. Follow the prompts to:
   - Select a data file.
   - Choose a data sheet from the data file.
4. The script will then:
   - Extract wellplates and their data.
   - Separate the wellplates into their own .xlsx files.
   - Zip all of the created files into a zip container.
"""




########## IMPORTS ##########


import os
import time
import textwrap
from sys import exit
from sys import argv as args
from datetime import datetime
from zipfile import ZipFile, ZIP_DEFLATED
try:
    from xlrd import open_workbook as open                          # type: ignore
except ImportError as e:
    exit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.\n')
try:
    from openpyxl import load_workbook as load                      # type: ignore
    from openpyxl.styles import Font, Alignment                     # type: ignore
    from openpyxl.workbook import Workbook                          # type: ignore
except ImportError as e:
    exit('\n> Error: You are missing the library to read .xlsx files called "openpyxl". Run the installer to update your libraries.\n')




########## ATTRIBUTES ##########

SCRIPT_VERSION = "0.1.0"                        # The public version number of the script

action_iterator = 1                             # A global iterator used for tracking actions
DELAY = 0.05                                    # The delay (in seconds) used for timing outputs and creating a user-friendly experience
DATA_FILES_DIRECTORY = "./data_files/"          # The directory within which the script looks for potential excel (.xls and .xlsx) data files.
RESULT_FILES_DIRECTORY = "./result_files/"      # The directory within which the script places its result files
WELLPLATE_ROWS = 8                              # The number of rows in the wellplate
WELLPLATE_COLS = 12                             # The number of columns in the wellplate
WELLS_IN_A_PARALLEL = 12                        # The number of wells in a parallel sample




########## CLASSES ##########


class Wellplate:
    """Class to encapsulate wellplate information."""

    def __init__(self, ordinal, data):
        """Initializes a Tissue object with a title."""
        self.ordinal = ordinal
        self.data = data

    def __str__(self):
        """String representation of the Tissue object."""
        return f"Plate_{self.ordinal}"




########## ARGUMENT HANDLING ##########


def info():
    """Function to print out instructions on how to use the script for the user."""
    try:
        # Print tutorial started status
        print_status("Tutorial Started", tailing_line_break=True)
        # Print each section of the tutorial with detailed instructions
        print_info("1) Evaluating Environment:", [
            "The script begins by checking the structure of the required directories.",
            "It checks for the existence of 'data_files' directory.",
            "It ensures that each directory contains at least one valid excel (.xls or .xlsx) file."
        ])
        print_info("2) Selecting Data File:", [
            "The user will then be prompted to select a data file from the 'data_files' directory.",
            "The available excel files will be listed, and the user will again choose one by its corresponding number."
        ])
        print_info("3) Selecting Data Sheet:", [
            "The user will select a sheet from the chosen data file.",
            "The available sheets will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("4) Extracting Wellplate Data:", [
            "The script will then read the selected data sheet and extract its wellplates.",
            f"It will look for the {WELLPLATE_ROWS}x{WELLPLATE_COLS} wellplate segments, starting from the appropriate row and column, and collect the absorbance values."
        ])
        print_info("5) Creating Excel Files:", [
            "Finally the scrip will create a new .xlsx file for each of the extracted wellplates and zip them into a zip container.",
            "The .xlsx files will be named as \"Plate_[plate_num]_[timestamp]_([origin_file]).xlsx\" to make them distinct and recognizable."
            "The zip container on the other hand will named as \"Extracted_plates_[timestamp]_([origin_file]).zip\"."
        ])
        # Print additional notes
        print_info("Notes:", [
            "The script may be interrupted at any point of its execution safely by hitting (CTRL+C).",
            "Should the user pick a .xls file as data file, the script will automatically conver it into a .xlsx file and replace the old .xls file altogether.",
            "The script looks for data files within the './data_files/' directory.",
            "The script places generated files into the './result_files/' directory.",
            "The timestamp in the name of the created files will be of format \"YYYYmmdd_HHMMSS\" so that the files can be sort ordered based on name."
        ])
        # Print explanation for each type of user prompt that the script uses
        print_info("User Prompts:", [
            "Input Prompt: When the script requires user input, it will display a prompt with a '>' symbol (e.g., \"Please select the data file > \"). The user will enter their response after the '>' symbol.",
            "Choices: Whenever the script presents multiple options to choose from, they will be listed with a corresponding command in square brackets (e.g., \"[1] data_file_1.xlsx\"). The user will select an option by typing its command.",
            "Task: The script will print task status messages to indicate what it is currently doing. These messages will start with \"> Task: \" followed by the task description.",
            "Progress: During longer operations, the script will print progress messages to keep the user informed. These messages will start with \"> Progress: \" followed by the progress description.",
            "Success: Upon successful completion of an operation, the script will print success messages starting with \"> Success: \" followed by a brief description of the success.",
            "Error: If an error occurs during an operation, the script will print error messages starting with \"> Error: \" followed by the error description."
        ])
        print_status("Tutorial completed", leading_line_break=False)
    except KeyboardInterrupt:
        exit_by_interruption(True)


def handle_arguments(args):
    """Handles the command line arguments provided to the script."""
    # Check if any arguments were given at launch
    if 1 < len(args):
        param = args[1].lower()
        if param in ["i", "info", "h", "help"]:
            # If the first argument was either "info" or "help", starts tutorial
            info()
            exit()




########## EXIT ROUTES ##########


def exit_by_error(message):
    """Exits the script with a given error message."""
    try:
        time.sleep(DELAY)
        print_error(message)
        time.sleep(DELAY)
        print_status("Failed")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")


def exit_by_interruption(during_info=False):
    """Exits the script due to an interruption (e.g., KeyboardInterrupt, CTRL+C)."""
    try:
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        if during_info:
            print_status("Tutorial stopped")    
        else:
            print_status("Stopped")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")




########## USER INTERACTION ##########


def prompt_input(prompt_str):
    """Gets user input after displaying a prompt."""
    try:
        time.sleep(DELAY)
        user_input = input(f"{prompt_str} > ").strip()
    except KeyboardInterrupt:
        exit_by_interruption()
    return user_input


def get_user_input(prompt, validation_func=None, type_select=True):
    """Gets and validates user input.

    Parameters:
    prompt (str): The prompt message to display.
    validation_func (function): The function to validate user input. Default is None.
    type_select (bool): Determines if input is for selection (True) or entry (False). Default is True.

    Returns:
    str: The validated user input.
    """
    # Print a blank line if type_select is True
    if type_select:
        print()
    while True:
        # Get user input
        user_input = prompt_input(prompt)
        # If no validation function is provided, return the input
        if not validation_func:
            if type_select:
                print()
            return user_input
        # Validate the user input
        validation_result, validated_input = validation_func(user_input)
        # If validation is successful, return the validated input
        if validation_result is True:
            if type_select:
                print()
            return validated_input
        # Print error message if validation fails
        print_error(validated_input)




########## PRINT FUNCTIONS ##########


def print_status(status_string, leading_line_break=True, tailing_line_break=False):
    """Prints a status message with version in the top border."""
    try:
        time.sleep(DELAY)
        text = f"EXTRACT SCRIPT {status_string.upper()}"
        # Middle line padding
        asterisks = "*" * 3
        spaces = " " * 5
        middle_line = f"{asterisks}{spaces}{text}{spaces}{asterisks}"
        # Top border with version
        top_border_text = f" v{SCRIPT_VERSION} "
        total_width = len(middle_line)
        top_line = "*" * ((total_width - len(top_border_text)) // 2) + top_border_text
        if len(top_line) < total_width:
            top_line += "*" * (total_width - len(top_line))
        bottom_line = "*" * len(middle_line)

        if leading_line_break:
            time.sleep(DELAY)
            print()
        time.sleep(DELAY)
        print(top_line)
        time.sleep(DELAY)
        print(middle_line)
        time.sleep(DELAY)
        print(bottom_line)
        if tailing_line_break or status_string.strip().lower() == "completed":
            time.sleep(DELAY)
            print()
    except KeyboardInterrupt:
        exit_by_interruption()


def print_action(action_str):
    """Prints a separator line for actions."""
    try:
        global action_iterator
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        print(f"========== Action {action_iterator}: {action_str.title()} ==========")
        action_iterator += 1
    except KeyboardInterrupt:
        exit_by_interruption()


def print_task(task_str):
    """Prints a task message."""
    try:
        time.sleep(DELAY)
        print(f"> Task: {task_str}..")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_choice(choice_cmd, choice_str):
    """Prints a choice option."""
    try:
        time.sleep(DELAY)
        print(f"[{choice_cmd}] {choice_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_progress(progress_str):
    """Prints a progress message."""
    try:
        time.sleep(DELAY)
        print(f"> Progress: {progress_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_success(success_str):
    """Prints a success message."""
    try:
        time.sleep(DELAY)
        print(f"> Success: {success_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_error(error_str):
    """Prints an error message."""
    try:
        time.sleep(DELAY)
        print(f"> Error: {error_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_info(header, messages, max_length=100):
    """Prints information with a delay and text wrapping."""
    try:
        INDENT = 3
        time.sleep(DELAY)
        print(header.title())
        for message in messages:
            wrapped_message = textwrap.fill(f" - {message}", max_length, subsequent_indent=' ' * INDENT)
            time.sleep(DELAY)
            print(wrapped_message)
        time.sleep(DELAY)
        print()
    except KeyboardInterrupt:
        exit_by_interruption(True)




########## VALIDATION FUNCTIONS ##########


def validate_digit(input_str):
    """Validate if the input string is a digit.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the integer value if valid, or an error message if not.
    """
    if input_str.isdigit():
        return True, int(input_str)
    return False, f'Your input of "{input_str}" was not a whole number.'


def validate_float(input_str):
    """Validate if the input string is a float, allowing for both Finnish (,) and English (.) decimal separators.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the float value if valid, or an error message if not.
    """
    try:
        input_str = input_str.replace(',', '.')
        float_value = float(input_str)
        return True, float_value
    except ValueError:
        return False, f'Your input "{input_str}" was not a number.'


def validate_non_empty(input_str):
    """Validate if the input string is not empty.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the input string if valid, or an error message if not.
    """
    if input_str.strip():
        return True, input_str
    return False, f"Your input was empty."


def validate_min_max(input_str, min_choice, max_choice, escape_str=None):
    """Validate if the input string is a digit within a specified range, with an optional escape string.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer or 0 if escaped, or an error message if not.
    """
    if escape_str is not None and input_str.strip() == escape_str:
        return True, 0
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if min_choice <= validation_result <= max_choice:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [{min_choice}, {max_choice}].'


def validate_yes_no(input_str):
    """Validate if the input string is a 'yes' or 'no' response.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated input string if valid, or an error message if not.
    """
    input_str = input_str.lower().strip()
    if input_str in ['y', 'yes', 'n', 'no']:
        return True, input_str
    return False, f'Your input "{input_str}" was invalid. Choose either yes or no.'


def validate_absorbance(absorbance):
    """Check if the absorbance value is valid."""
    if isinstance(absorbance, (int, float)):
        return True
    return False


def validate_parallel(sheet, row, first_col):
    """Check if the row contains valid float values in all wells of the parallel."""
    for col in range(first_col, WELLS_IN_A_PARALLEL + first_col):
        cell_value = sheet.cell(row=row, column=col).value
        if not validate_absorbance(cell_value):
            return False
    return True




########## SEARCH FUNCTIONS ##########


def find_wellplate(sheet, start_row):
    """Finds the start, end, and next row of a wellplate by scanning the entire sheet.
    
    Returns:
    dict: {'start': (row, col), 'end': (row, col), 'next_row': row} or None if not found.
    """
    START_COL = 1
    END_COL = 5
    VALID_ROW_THRESHOLD = 4  # Minimum valid rows required to consider a wellplate found
    # Iterate through the entire sheet, checking each row in the first column for valid absorbance values
    for row in range(start_row, sheet.max_row + 1):
        for col in range(START_COL, END_COL + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if not validate_absorbance(cell_value):
                continue
            # Count how many valid rows of absorbance values there are
            valid_row_count = 0
            for check_row in range(row, WELLPLATE_ROWS + row):
                if not validate_parallel(sheet, check_row, col):
                    continue
                valid_row_count += 1
            # If not enough valid rows were discovered we didn't find a wellpalte
            if valid_row_count < VALID_ROW_THRESHOLD:
                continue
            # Wellplate was found returning its information
            return {
                "start": (row, col),  # Starting cell of the wellplate
                "end": (row + valid_row_count - 1, WELLPLATE_COLS + col - 1),  # Ending cell of the wellplate
                "next_row": row + valid_row_count  # Row after the wellplate for further search
            }
    return None # No wellplates were found




########## FILE CONVERTER ##########


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Function to convert deprecated .xls files into .xlsx files."""
    # Open the .xls file using xlrd
    workbook_xls = open(xls_path)
    workbook_xlsx = Workbook()
    for sheet_index in range(workbook_xls.nsheets):
        # One sheet at a time
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        if sheet_index == 0:
            # Activate the workbook to create its first sheet
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
        else:
            # Add them otherwise
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
        # One row at a time
        for row_index in range(sheet_xls.nrows):
            for col_index in range(sheet_xls.ncols):
                # Write each cell
                sheet_xlsx.cell(row=row_index + 1, column=col_index + 1).value = sheet_xls.cell_value(row_index, col_index)
    try:
        # Save the new .xlsx file using openpyxl
        workbook_xlsx.save(xlsx_path)
        # Delete the old .xls file using xlrds
        os.remove(xls_path)
    except PermissionError:
        # If couldn't remove the original file, remove the new one to avoid duplicates
        os.remove(xlsx_path)
        exit_by_error(f'File "{xls_path}" was already open in another program. Please close it and try again.')




########## CORE FUNCTIONS ##########


def get_excel_files(directory):
    """Function to find all excel (.xls and .xlsx) files in the given directory that do not start with '~$'."""
    return [file for file in os.listdir(directory) if (file.endswith('.xlsx') or file.endswith('.xls')) and not file.startswith('~$')]


def evaluate_script_environment():
    """Check if the required directories with acceptable content exist in the current working directory."""
    # Check for the existence of data files directory
    if not os.path.isdir(DATA_FILES_DIRECTORY):
        exit_by_error(f'Data files directory "{DATA_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Data files directory "{DATA_FILES_DIRECTORY}" located.')
    # Check if data files directory contains at least one excel file that doesn't start with "~$"
    data_files = get_excel_files(DATA_FILES_DIRECTORY)
    if not data_files:
        exit_by_error(f'Data files directory did not contain any excel files.')
    print_progress(f'Data files directory contained excel files.')
    # Check for the existence of result files directory
    if not os.path.isdir(RESULT_FILES_DIRECTORY):
        exit_by_error(f'Result files directory "{RESULT_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Result files directory "{RESULT_FILES_DIRECTORY}" located.')


def select_excel_file(directory):
    """Selects an Excel file from the specified directory."""
    # List all Excel files in the directory, excluding temporary files
    try:
        excel_files = get_excel_files(directory)
    except FileNotFoundError:
        exit_by_error(f'Directory "{directory}" could not be located.')
    # If no Excel files are found, exit the script with an error message
    if not excel_files:
        exit_by_error(f'Directory "{directory}" did not contain any excel files.')
    # Print choices for the user to select from
    for i, file in enumerate(excel_files):
        print_choice(i+1, file)
    # User chooses a file
    int_value = get_user_input(f'Please select the file', lambda input_str: validate_min_max(input_str, 1, len(excel_files)))
    return excel_files[int_value - 1]


def load_excel_file(directory, file_name):
    """Load an Excel file from the given path and return the workbook object."""
    try:
        # Convert .xls to .xlsx if necessary
        if file_name.endswith('.xls'):
            xls_path = directory + file_name
            file_name += 'x'
            xlsx_path = directory + file_name
            convert_xls_to_xlsx(xls_path, xlsx_path)
            print_progress("Converted the old .xls file into a new .xlsx file.")
        else:
            xlsx_path = directory + file_name
        # Load the Excel file
        file = load(xlsx_path)
        # Save the unchanged file to check if it is accessible
        file.save(xlsx_path)
        return file, file_name
    except FileNotFoundError:
        # Handle file not found error
        exit_by_error(f'File "{file_name}" could not be located.')
    except PermissionError:
        # Handle permission error if the file is already open
        exit_by_error(f'File "{file_name}" was already open in another program. Please close it and try again.')
    except Exception as e:
        # Handle any other exceptions that occur during file loading
        exit_by_error(f'Failed loading file "{file_name}".  > REASON: {e}.')


def select_datasheet(data_file):
    """Prompt the user to select a data sheet from the given data file."""
    # List all sheet names in the data file
    datasheets = [sheet for sheet in data_file.sheetnames]
    for i, sheet in enumerate(datasheets):
        print_choice(i+1, sheet)
    # User chooses a sheet
    int_value = get_user_input('Please select the sheet', lambda input_str: validate_min_max(input_str, 1, len(data_file.sheetnames)))
    return datasheets[int_value - 1]


def load_datasheet(data_file, data_sheet_name):
    """Load the specified data sheet from the given data file."""
    return data_file[data_sheet_name]


def get_wellplate_data(wellplate):
    """Creates a dictionary of absorbance values from the wellplate segment with manually generated row chars and col nums.

    Returns:
    dict: A dictionary with (row_char, col_num) as keys and absorbance values as values.
    """
    # Create the dictionary to hold the absorbance values
    wellplate_data = {}
    # Extract the data as new key (row character, column number) value (absorbance) pairs of the dictionary
    for i, row in enumerate(wellplate):
        # Manually create the row character based on the row
        row_char = chr(64+i+1)
        for j, well in enumerate(row):
            # Manually create the column number based on the well
            col_num = j+1
            absorbance = well.value
            # Add the absorbance value to the dictionary
            wellplate_data[(row_char, col_num)] = absorbance
    return wellplate_data


def extract_segment(data_sheet, start_row, start_col, end_row, end_col):
    """Extracts a specified segment of data from a datasheet."""
    # Determine the coordinates of the first and last cell of the segment
    first_cell = data_sheet.cell(row=start_row, column=start_col).coordinate
    last_cell = data_sheet.cell(row=end_row, column=end_col).coordinate
    # Extract the segment
    segment = data_sheet[first_cell:last_cell]
    return segment


def extract_wellplates(data_sheet):
    """Extracts wellplates from the specified data sheet."""
    plate_num = 1
    next_row = 1
    wellplates = []
    while True:
        # Find a wellplate in the sheet
        wellplate_info = find_wellplate(data_sheet, next_row)
        # If no wellplate was found then the sheet contains no more plates
        if not wellplate_info:
            break
        start_row, start_col = wellplate_info["start"]
        end_row, end_col = wellplate_info["end"]
        # Extract the wellplate segment from the sheet
        wellplate_segment = extract_segment(data_sheet, start_row, start_col, end_row, end_col)
        # Get the absorbance values from the wellplate segment
        wellplate_data = get_wellplate_data(wellplate_segment)
        # Create a wellplate object
        wellplate = Wellplate(plate_num, wellplate_data)
        # Add the object as part of a list
        wellplates.append(wellplate)
        print_progress(f"Wellplate {plate_num} extracted from [{chr(64+start_col)}{start_row}, {chr(64+end_col)}{end_row}].")
        # Update the row to read onwards from
        next_row = wellplate_info["next_row"]
        plate_num += 1
    # Exit with an error in case no wellplates were found on the sheet
    if len(wellplates) < 1:
        exit_by_error(f"Could not locate any wellplates in data sheet '{data_sheet.title}'.")
    # Return the wellplates
    return wellplates


def write_wellplate_data(sheet, wellplate):
    """Function to write wellplate data into a worksheet."""
    data = wellplate.data
    # Extract unique row characters and column numbers, and sort them
    row_chars = sorted(set(row for row, _ in data.keys()))
    col_nums = sorted(set(col for _, col in data.keys()))
    # Write row characters as row headers
    for i, row_char in enumerate(row_chars):
        sheet.cell(row=i+2, column=1).value = row_char
        sheet.cell(row=i+2, column=1).font = Font(bold=True)
        sheet.cell(row=i+2, column=1).alignment = Alignment(horizontal='center')
    # Write column col numbers as column headers
    for i, col_num in enumerate(col_nums):
        sheet.cell(row=1, column=i+2).value = col_num
        sheet.cell(row=1, column=i+2).font = Font(bold=True)
        sheet.cell(row=1, column=i+2).alignment = Alignment(horizontal='center')
    # Write absorbance data in the table
    for (row_char, col_num), absorbance in data.items():
        write_row = row_chars.index(row_char) + 2
        write_col = col_nums.index(col_num) + 2
        sheet.cell(row=write_row, column=write_col).value = absorbance
        sheet.cell(row=write_row, column=write_col).alignment = Alignment(horizontal='center')


def create_wellplate_file(wellplate, file_path):
    """Function to create an excel file for a wellplate."""
    # Create a new workbook and add a new worksheet for the wellplate data
    workbook = Workbook()
    platesheet = workbook.active
    platesheet.title = f"Plate_{wellplate.ordinal}"
    # Write the wellplate data into the sheet
    write_wellplate_data(platesheet, wellplate)
    # Save the workbook to the specified file path
    workbook.save(file_path)


def wellplates_to_files(wellplates, data_file_name):
    """Function to log wellplates into files of their own and zip them."""
    # Get current DateTime in 'YYYYMMDD_HHmmSS' format
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    # Zip file name e.g.: "Extracted_Plates_20250101_235959_(data_file.xls).zip"
    zip_file_name = f"Extracted_Plates_{timestamp}_({data_file_name}).zip"
    zip_file_path = os.path.join(RESULT_FILES_DIRECTORY, zip_file_name)
    # Create the zip file
    with ZipFile(zip_file_path, 'w', ZIP_DEFLATED) as zipf:
        # Loop through wellplates and create individual files
        for plate in wellplates:
            # Wellplate excel file name e.g.: "Plate_3_20250101_235959_(data_file.xls).xlsx"
            plate_file_name = f"Plate_{plate.ordinal}_{timestamp}_({data_file_name}).xlsx"
            plate_file_path = os.path.join(RESULT_FILES_DIRECTORY, plate_file_name)
            # Create the Excel file for each wellplate
            create_wellplate_file(plate, plate_file_path)
            # Add the wellplate file to the zip
            zipf.write(plate_file_path, plate_file_name)
            # Remove the individual wellplate file after adding to zip
            os.remove(plate_file_path)
            print_progress(f"Plate {plate.ordinal} done.")
    return zip_file_name




########## ACTUAL SCRIPT ##########


def run_script():
    print_action("Evaluating environment")
    print_task(f'Checking script environment structure.')
    evaluate_script_environment()
    print_success("Script environment structure accepted.")
    # NECESSARY DIRECTORIES HAVE BEEN LOCATED #

    print_action("Selecting data file")
    data_file_name = select_excel_file(DATA_FILES_DIRECTORY)
    print_task(f'Loading data file "{data_file_name}".')
    # DATA FILE HAS BEEN SELECTED #
    data_file, data_file_name = load_excel_file(DATA_FILES_DIRECTORY, data_file_name)
    print_success(f'Data file "{data_file_name}" loaded.')
    # DATA FILE HAS BEEN LOADED #

    print_action("Selecting data sheet")
    data_sheet_name = select_datasheet(data_file)
    print_task(f'Loading data sheet "{data_sheet_name}".')
    # DATA SHEET HAS BEEN SELECTED #
    data_sheet = load_datasheet(data_file, data_sheet_name)
    print_success(f'Data sheet "{data_sheet_name}" loaded.')
    # DATA SHEET HAS BEEN LOADED #

    print_action("Extracting wellplates")
    print_task(f'Extracting wellplates from data sheet "{data_sheet_name}".')
    wellplates = extract_wellplates(data_sheet)
    print_success(f"Total of {len(wellplates)} wellplates extracted.")
    # DATA HAS BEEN EXTRACTED #

    print_action("Creating excel files")
    print_task(f'Creating files for each extracted wellplate and storing them in "{RESULT_FILES_DIRECTORY}".')
    zip_file_name = wellplates_to_files(wellplates, data_file_name)
    print_success(f'Excel files created and stored into "{zip_file_name}".')
    # WELLPLATE FILES HAVE BEEN CREATED #




########## MAIN ENTRY POINT ##########


if __name__ == "__main__":
    handle_arguments(args)
    try:
        print_status("Started")
        run_script()
        print_status("Completed")
    except KeyboardInterrupt:
        exit_by_interruption()
