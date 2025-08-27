"""
4PL-Result Statistical Testing Script

Summary:
This script is designed to run statistical tests on the data produced by the 4PLScript. The primary functionalities include
selecting and loading data files, selecting the statistical test method, extracting 4PL data, performing the selected statistical test,
and writing the results back to Excel files.
The script starts by ensuring the necessary libraries are installed and then initializes global attributes related to
the experiment setup. It includes user interaction functions for input validation and output formatting, ensuring
a smooth user experience.
Interruptions are handled gracefully, and the script provides clear progress and error messages throughout its execution.

Usage:
1. Ensure you have the required directory (`data_files`) with appropriate files.
2. Run the script from the command line.
3. Follow the prompts to:
   - Select a data file.
   - Choose a data sheet from the data file.
   - Select the statistical testing method.
4. The script will then:
   - Extract 4PL-analysis data.
   - Perform statistical testing.
   - Write the statistical results back into the selected file.
5. The updated data file will be saved with the new statistics and graphs.
"""




########## IMPORTS ##########


import os
import time
import textwrap
from sys import exit
from sys import argv as args
from itertools import combinations
try:
    from xlrd import open_workbook                                  # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xls files called "xlrd". Run the installer to update your libraries.\n')
try:
    from openpyxl import load_workbook as load                      # type: ignore
    from openpyxl.styles import Font, Alignment                     # type: ignore
    from openpyxl.workbook import Workbook                          # type: ignore
    from openpyxl.utils import get_column_letter                    # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library to read .xlsx files called "openpyxl". Run the installer to update your libraries.\n')
try:
    from scipy.stats import kruskal, mannwhitneyu                   # type: ignore
except ImportError:
    exit('\n> Error: You are missing the library for curve fitting called "scipy". Run the installer to update your libraries.\n')
try:
    import pandas as panda                                          # type: ignore
except ImportError:
    exit('\n> Error: You are missing the data handling library called "pandas". Run the installer to update your libraries.\n')
try:
    import scikit_posthocs as posthoc                               # type: ignore
except ImportError:
    exit('\n> Error: You are missing the statistical test library called "scikit-posthocs". Run the installer to update your libraries.\n')




########## ATTRIBUTES ##########


action_iterator = 1                                                 # A global iterator used for tracking actions
DELAY = 0.05                                                        # The delay (in seconds) used for timing outputs and creating a user-friendly experience
ENVIRONMENT = "SCR"                                                 # The environment mode for the script, typically set to 'SCR' (short for scripted)
KRUSKAL_WALLIS = "Kruskal-Wallis"                                   # Name of the test for Kruskal-Wallis
MANN_WHITNEY = "Mann-Whitney"                                       # Name of the test for Mann-Whitney
STATISTICAL_TESTS = [ KRUSKAL_WALLIS, MANN_WHITNEY ]                # Statistical tests that are supported
CONFIG_PARAMETER = "analysis_types"                                 # Parameter that is configured in the config file
DATA_FILES_DIRECTORY = "./data_files/"                              # The directory within which the script looks for potential excel (.xls and.xlsx) data files
ALPHA = 0.05                                                        # The maximum allowed probability of a false positive
SIGNIFICANCE_LEVELS = [                                             # Stores the p-value to symbol pairs
    (0.0001, "****"),
    (0.001, "***"),
    (0.01, "**"),
    (0.05, "*"),
    (float('inf'), "ns"),
]




########## ARGUMENT HANDLING ##########


def info():
    """Function to print out instructions on how to use the script for the user."""
    try:
        # Print tutorial started status
        print_status("Tutorial Started", tailing_line_break=True)
        # Print each section of the tutorial with detailed instructions
        print_info("1) Evaluating Environment:", [
            "The script begins by checking the structure of the required directories.",
            "It checks for the existence of 'data_files' directory.",
            "It ensures that it contains valid files (.xls or .xlsx)."
        ])
        print_info("2) Selecting Data File:", [
            "The user will then be prompted to select a data file from the 'data_files' directory.",
            "The available excel files will be listed, and the user will again choose one by its corresponding number."
        ])
        print_info("3) Selecting Data Sheet:", [
            "The user will select a sheet from the chosen data file.",
            "The available sheets will be listed, and the user will choose one by its corresponding number."
        ])
        print_info("4) Selecting Statistical Test:", [
            "The user will select the statistical test method."
        ])
        print_info("5) Extracting Test Data:", [
            "The script will then read the selected data sheet and extract the test data.",
            "The statistical tests require at least two (2) datapoints to work."
        ])
        print_info("6) Integrating Wellplate Data:", [
            "The script will integrate the extracted wellplate data into the configured groups.",
            "Each sample and parallel within the groups will be populated with the corresponding absorbance values from the wellplate data."
        ])
        print_info("7) Running Statistical Test:", [
            "The script will run the selected statistical test and calculate the result values for the tested data.",
            "The script will then write the test results into the data file.",
        ])
        print_info("8) Saving Data File:", [
            "Finally the script will save the updated data file with the new test sheets to ensure that the intermediate results are preserved."
        ])
        
        # Print additional notes
        print_info("Notes:", [
            "The script may be interrupted at any point of its execution safely by hitting (CTRL+C).",
            "Should the user pick a .xls file as data file, the script will automatically conver it into a .xlsx file and replace the old .xls file altogether.",
            "The script looks for data files within the './data_files/' directory.",
        ])
        # Print explanation for each type of user prompt that the script uses
        print_info("User Prompts:", [
            "Input Prompt: When the script requires user input, it will display a prompt with a '>' symbol (e.g., \"Please select the data file > \"). The user will enter their response after the '>' symbol.",
            "Choices: Whenever the script presents multiple options to choose from, they will be listed with a corresponding command in square brackets (e.g., \"[1] data_file_1.xlsx\"). The user will select an option by typing its command.",
            "Task: The script will print task status messages to indicate what it is currently doing. These messages will start with \"> Task: \" followed by the task description.",
            "Progress: During longer operations, the script will print progress messages to keep the user informed. These messages will start with \"> Progress: \" followed by the progress description.",
            "Success: Upon successful completion of an operation, the script will print success messages starting with \"> Success: \" followed by a brief description of the success.",
            "Error: If an error occurs during an operation, the script will print error messages starting with \"> Error: \" followed by the error description."
        ])
        print_status("Tutorial completed", leading_line_break=False)
    except KeyboardInterrupt:
        exit_by_interruption(True)


def handle_arguments(args):
    """Handles the command line arguments provided to the script."""
    # Check if any arguments were given at launch
    if 1 < len(args):
        param = args[1].lower()
        if param in ["i", "info", "h", "help"]:
            # If the first argument was either "info" or "help", start tutorial
            info()
            exit()




########## EXIT ROUTES ##########


def exit_by_error(message=None):
    """Exits the script with a given error message."""
    try:
        if message:
            time.sleep(DELAY)
            print_error(message)
        time.sleep(DELAY)
        print_status("Failed")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")


def exit_by_interruption(during_info=False):
    """Exits the script due to an interruption (e.g., KeyboardInterrupt, CTRL+C)."""
    try:
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        if during_info:
            print_status("Tutorial stopped")
        else:
            print_status("Stopped")
    except KeyboardInterrupt:
        exit_by_interruption()
    exit("")




########## USER INTERACTION ##########


def prompt_input(prompt_str):
    """Gets user input after displaying a prompt."""
    try:
        time.sleep(DELAY)
        user_input = input(f"{prompt_str} > ").strip()
    except KeyboardInterrupt:
        exit_by_interruption()
    return user_input


def get_user_input(prompt, validation_func=None, type_select=True):
    """Gets and validates user input.

    Parameters:
    prompt (str): The prompt message to display.
    validation_func (function): The function to validate user input. Default is None.
    type_select (bool): Determines if input is for selection (True) or entry (False). Default is True.

    Returns:
    str: The validated user input.
    """
    # Print a blank line if type_select is True
    if type_select:
        print()
    while True:
        # Get user input
        user_input = prompt_input(prompt)
        # If no validation function is provided, return the input
        if not validation_func:
            if type_select:
                print()
            return user_input
        # Validate the user input
        validation_result, validated_input = validation_func(user_input)
        # If validation is successful, return the validated input
        if validation_result is True:
            if type_select:
                print()
            return validated_input
        # Print error message if validation fails
        print_error(validated_input)




########## PRINT FUNCTIONS ##########


def print_status(status_string, leading_line_break=True, tailing_line_break=False):
    """Prints a status message."""
    try:
        time.sleep(DELAY)
        asterisks = "*" * 3
        spaces = " " * 5
        text = f"4PL SCRIPT {status_string.upper()}"
        middle_line = f"{asterisks}{spaces}{text}{spaces}{asterisks}"
        top_bottom_line = '*' * len(middle_line)
        if leading_line_break:
            time.sleep(DELAY)
            print()
        time.sleep(DELAY)
        print(top_bottom_line)
        time.sleep(DELAY)
        print(middle_line)
        time.sleep(DELAY)
        print(top_bottom_line)
        if tailing_line_break or status_string.strip().lower() == "completed":
            time.sleep(DELAY)
            print()
    except KeyboardInterrupt:
        exit_by_interruption()


def print_action(action_str):
    """Prints a separator line for actions."""
    try:
        global action_iterator
        time.sleep(DELAY)
        print()
        time.sleep(DELAY)
        print(f"========== Action {action_iterator}: {action_str.title()} ==========")
        action_iterator += 1
    except KeyboardInterrupt:
        exit_by_interruption()


def print_task(task_str):
    """Prints a task message."""
    try:
        time.sleep(DELAY)
        print(f"> Task: {task_str}..")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_choice(choice_cmd, choice_str):
    """Prints a choice option."""
    try:
        time.sleep(DELAY)
        print(f"[{choice_cmd}] {choice_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_progress(progress_str):
    """Prints a progress message."""
    try:
        time.sleep(DELAY)
        print(f"> Progress: {progress_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_success(success_str):
    """Prints a success message."""
    try:
        time.sleep(DELAY)
        print(f"> Success: {success_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_error(error_str):
    """Prints an error message."""
    try:
        time.sleep(DELAY)
        print(f"> Error: {error_str}")
    except KeyboardInterrupt:
        exit_by_interruption()


def print_info(header, messages, max_length=100):
    """Prints information with a delay and text wrapping."""
    try:
        INDENT = 3
        time.sleep(DELAY)
        print(header.title())
        for message in messages:
            wrapped_message = textwrap.fill(f" - {message}", max_length, subsequent_indent=' ' * INDENT)
            time.sleep(DELAY)
            print(wrapped_message)
        time.sleep(DELAY)
        print()
    except KeyboardInterrupt:
        exit_by_interruption(True)




########## VALIDATION FUNCTIONS ##########


def validate_digit(input_str):
    """Validate if the input string is a digit.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the digit value if valid, or an error message if not.
    """
    if input_str.isdigit():
        return True, int(input_str)
    return False, f'Your input of "{input_str}" was not a positive whole number.'


def validate_float(input_str):
    """Validate if the input string is a float, allowing for both Finnish (,) and English (.) decimal separators.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the float value if valid, or an error message if not.
    """
    try:
        input_str = str(input_str).replace(',', '.')
        float_value = float(input_str)
        return True, float_value
    except ValueError:
        return False, f'Your input "{input_str}" was not a number.'


def validate_min_max(input_str, min_choice, max_choice, escape_str=None):
    """Validate if the input string is a digit within a specified range, with an optional escape string.

    Returns:
    tuple: A tuple where the first value is a boolean indicating success,
           and the second value is the validated integer or 0 if escaped, or an error message if not.
    """
    if escape_str is not None and input_str.strip() == escape_str:
        return True, 0
    is_valid, validation_result = validate_digit(input_str)
    if not is_valid:
        return False, validation_result
    if min_choice <= validation_result <= max_choice:
        return True, validation_result
    return False, f'Your input "{input_str}" was out of the allowed range [{min_choice}, {max_choice}].'


def validate_group_title(value):
    """Validate if the value contains an appropriate group title like 'Group...'."""
    if isinstance(value, str) and value.casefold().startswith("group"):
        return True
    return False


def validate_sample_string(value):
    """Validate if the input matches the string of 'sample' case insensitively."""
    if isinstance(value, str) and "sample" in value.casefold():
        return True
    return False


def validate_sample_label(value):
    """Validate if the value is an acceptable sample label."""
    if isinstance(value, str) and value.strip():
        return True
    is_valid, validation_result = validate_float(value)
    if is_valid:
        return True
    return False


def validate_result_string(value):
    """Validate if the input matches an acceptable result string."""
    if isinstance(value, str) and value.strip():
        return True
    is_valid, validation_result = validate_float(value)
    if is_valid:
        return True
    return False


def validate_sample_result(value):
    """Validate if the value is a valid absorbance value."""
    is_valid, validation_result = validate_float(value)
    if is_valid:
        return True
    return False




########## FILE CONVERTER ##########


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """Function to convert deprecated .xls files into .xlsx files."""
    # Open the .xls file using xlrd
    workbook_xls = open_workbook(xls_path)
    workbook_xlsx = Workbook()
    for sheet_index in range(workbook_xls.nsheets):
        # One sheet at a time
        sheet_xls = workbook_xls.sheet_by_index(sheet_index)
        if sheet_index == 0:
            # Activate the workbook to create its first sheet
            sheet_xlsx = workbook_xlsx.active
            sheet_xlsx.title = sheet_xls.name
        else:
            # Add them otherwise
            sheet_xlsx = workbook_xlsx.create_sheet(title=sheet_xls.name)
        # One row at a time
        for row_index in range(sheet_xls.nrows):
            for col_index in range(sheet_xls.ncols):
                # Write each cell
                sheet_xlsx.cell(row=row_index + 1, column=col_index + 1).value = sheet_xls.cell_value(row_index, col_index)
    try:
        # Save the new .xlsx file using openpyxl
        workbook_xlsx.save(xlsx_path)
        # Delete the old .xls file using xlrds
        os.remove(xls_path)
    except PermissionError:
        # If couldn't remove the original file, remove the new one to avoid duplicates
        os.remove(xlsx_path)
        exit_by_error(f'File "{xls_path}" was already open in another program. Please close it and try again.')




########## CORE FUNCTIONS ##########


def directory_contains_file(directory, filename):
    """Check if a specific file exists in the given directory."""
    return filename in os.listdir(directory)


def get_excel_files(directory):
    """Function to find all excel files (.xls and .xlsx) in the given directory that do not start with '~$'."""
    return [file for file in os.listdir(directory) if (file.endswith('.xlsx') or file.endswith('.xls')) and not file.startswith('~$')]


def evaluate_script_environment():
    """Check if the required directories with acceptable content exist in the current working directory."""
    # Check for the existence of data files directory
    if not os.path.isdir(DATA_FILES_DIRECTORY):
        exit_by_error(f'Data files directory "{DATA_FILES_DIRECTORY}" could not be located.')
    print_progress(f'Data files directory "{DATA_FILES_DIRECTORY}" located.')
    # Check if data files directory contains at least one excel file that doesn't start with "~$"
    data_files = get_excel_files(DATA_FILES_DIRECTORY)
    if not data_files:
        exit_by_error(f'Data files directory did not contain any excel files.')
    print_progress(f'Data files directory contained excel files.')


def select_excel_file(directory):
    """Selects an Excel file from the specified directory."""
    # List all Excel files in the directory, excluding temporary files
    try:
        excel_files = get_excel_files(directory)
    except FileNotFoundError:
        exit_by_error(f'Directory "{directory}" could not be located.')
    # If no Excel files are found, exit the script with an error message
    if not excel_files:
        exit_by_error(f'Directory "{directory}" did not contain any excel files.')
    # Print choices for the user to select from
    for i, file in enumerate(excel_files):
        print_choice(i+1, file)
    # User chooses a file
    int_value = get_user_input(f'Please select the file', lambda input_str: validate_min_max(input_str, 1, len(excel_files)))
    return excel_files[int_value - 1]


def load_excel_file(directory, file_name):
    """Load an Excel file from the given path and return the workbook object."""
    try:
        # Convert .xls to .xlsx if necessary
        if file_name.endswith('.xls'):
            xls_path = directory + file_name
            file_name += 'x'
            xlsx_path = directory + file_name
            convert_xls_to_xlsx(xls_path, xlsx_path)
            print_progress("Converted the old .xls file into a new .xlsx file.")
        else:
            xlsx_path = directory + file_name
        # Load the Excel file
        file = load(xlsx_path)
        # Save the unchanged file to check if it is accessible
        file.save(xlsx_path)
        return file, file_name
    except FileNotFoundError:
        # Handle file not found error
        exit_by_error(f'File "{file_name}" could not be located.')
    except PermissionError:
        # Handle permission error if the file is already open
        exit_by_error(f'File "{file_name}" was already open in another program. Please close it and try again.')
    except Exception as e:
        # Handle any other exceptions that occur during file loading
        exit_by_error(f'Failed loading file "{file_name}".  > REASON: {e}.')


def select_datasheet(data_file):
    """Prompt the user to select a data sheet from the given data file."""
    # List all sheet names in the data file
    datasheets = [sheet for sheet in data_file.sheetnames]
    for i, sheet in enumerate(datasheets):
        print_choice(i+1, sheet)
    # User chooses a sheet
    int_value = get_user_input('Please select the sheet', lambda input_str: validate_min_max(input_str, 1, len(data_file.sheetnames)))
    return datasheets[int_value - 1]


def load_datasheet(data_file, data_sheet_name):
    """Load the specified data sheet from the given data file."""
    return data_file[data_sheet_name]


def select_statistical_test():
    """Select the statistical test method by asking the user for a test number."""
    for i, test in enumerate(STATISTICAL_TESTS):
        print_choice(i+1, test)
    # User chooses a test
    int_value = get_user_input('Please select the statistical test', lambda input_str: validate_min_max(input_str, 1, len(STATISTICAL_TESTS)))
    return STATISTICAL_TESTS[int_value - 1]


def extract_segment(data_sheet, start_row, start_col, end_row, end_col):
    """Extracts a specified segment of data from a datasheet."""
    # Determine the coordinates of the first and last cell of the segment
    first_cell = data_sheet.cell(row=start_row, column=start_col).coordinate
    last_cell = data_sheet.cell(row=end_row, column=end_col).coordinate
    # Extract the segment
    segment = data_sheet[first_cell:last_cell]
    return segment


def extract_group_data(sheet, start_row):
    """Finds the information of a group by scanning the entire sheet.
    
    Returns:
    dict: {'title', 'data', 'next_row'} or None if not found.
    """
    group_data = []
    START_COL = 1
    END_COL = 5
    # Iterate through the entire sheet, checking each row's first 5 cols for a valid group.
    for row in range(start_row, sheet.max_row + 1):
        for col in range(START_COL, END_COL + 1):
            group_title = sheet.cell(row=row, column=col).value
            if not validate_group_title(group_title):
                continue
            sample_string = sheet.cell(row=row+1, column=col).value
            if not validate_sample_string(sample_string):
                continue
            result_string = sheet.cell(row=row+2, column=col).value
            if not validate_result_string(result_string):
                continue
            i = 0
            while (True):
                sample_label = sheet.cell(row=row+1, column=col+1+i).value
                if not validate_sample_label(sample_label):
                    break
                sample_result = sheet.cell(row=row+2, column=col+1+i).value
                if not validate_sample_result(sample_result):
                    break
                group_data.append(sample_result)
                i += 1
            if not len(group_data):
                continue
            # Group was found, returning its details
            return {
                "title": group_title,
                "data": group_data,
                "next_row": row + 3
            }
    return None # No groups were found


def extract_test_data(data_sheet):
    """Extracts test data from the specified data sheet."""
    # Create the data dict
    test_data = {}
    search_row = 1
    while True:
        # Find a group data segment in the sheet
        group_info = extract_group_data(data_sheet, search_row)
        # If no group was found then the sheet contains no more valid data
        if not group_info:
            break
        # Extract the group data segment
        title = group_info['title']
        data = group_info['data']
        test_data[title] = data
        print_progress(f'{title} with {len(data)} data points extracted.')
        search_row = group_info['next_row']
    if len(test_data) < 2:
        exit_by_error(f'Could not find enough groups on sheet "{data_sheet.title}" to perform statistical analysis on, at least 2 groups are required.')
    return test_data


def create_or_clear_sheet(xlsx_file, sheet_name, position=None):
    """Create a new sheet or clear it if it already exists."""
    # Check if the sheet already exists
    if sheet_name in xlsx_file.sheetnames:
        # Delete the existing sheet
        print_progress(f'Sheet "{sheet_name}" data wiped.')
        del xlsx_file[sheet_name]
    else:
        print_progress(f'New sheet "{sheet_name}" created.')
    # Create and return a new sheet
    sheet = xlsx_file.create_sheet(sheet_name)
    if position:
        sheets = xlsx_file._sheets
        sheets.insert(position, sheets.pop(sheets.index(sheet)))
    return sheet


def get_significance_symbol(p_value):
    """Returns a significance symbol based on the p-value according to predefined thresholds."""
    for threshold, symbol in SIGNIFICANCE_LEVELS:
        if p_value <= threshold:
            return symbol


def get_col_width(sheet, col_char, padding=2):
    """Calculate the width of a column based on the maximum length of the data in that column plus padding."""
    DEFAULT_WIDTH = 8.43
    lengths = [len(str(cell.value)) for cell in sheet[col_char] if cell.value is not None]
    if not lengths:
        return DEFAULT_WIDTH
    return max(lengths) + padding


def autosize_cols(sheet, padding=2):
    """Auto-size all columns in the given Excel worksheet using calculated content width plus padding."""
    for col_cells in sheet.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = get_col_width(sheet, col_letter, padding=padding)
        sheet.column_dimensions[col_letter].width = width


def kw_to_data_file(test_sheet, statistics):
    """Writes Kruskal-Wallis and Dunn's test results into an Excel sheet with inline pairwise comparisons."""
    write_row, write_col = 1, 1
    # Kruskal-Wallis results
    test_sheet.cell(row=write_row, column=write_col).value = "KRUSKAL-WALLIS RESULTS"
    test_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    test_sheet.cell(row=write_row + 1, column=write_col).value = "P-value"
    test_sheet.cell(row=write_row + 1, column=write_col).font = Font(bold=True)
    kw_p_val = statistics["p_value"]
    test_sheet.cell(row=write_row + 1, column=write_col + 1).value = kw_p_val
    test_sheet.cell(row=write_row + 2, column=write_col).value = "Significance"
    test_sheet.cell(row=write_row + 2, column=write_col).font = Font(bold=True)
    symbol = get_significance_symbol(kw_p_val)
    test_sheet.cell(row=write_row + 2, column=write_col + 1).value = symbol
    test_sheet.cell(row=write_row + 2, column=write_col + 1).alignment = Alignment(horizontal='right')
    # Dunn's results header
    test_sheet.cell(row=write_row + 4, column=write_col).value = "DUNN'S RESULTS"
    test_sheet.cell(row=write_row + 4, column=write_col).font = Font(bold=True)
    # Write column headers for Dunn's inline comparisons
    test_sheet.cell(row=write_row + 5, column=write_col).value = "Comparison"
    test_sheet.cell(row=write_row + 5, column=write_col).font = Font(bold=True)
    test_sheet.cell(row=write_row + 5, column=write_col).alignment = Alignment(horizontal='left')
    test_sheet.cell(row=write_row + 5, column=write_col + 1).value = "P-value"
    test_sheet.cell(row=write_row + 5, column=write_col + 1).font = Font(bold=True)
    test_sheet.cell(row=write_row + 5, column=write_col + 1).alignment = Alignment(horizontal='right')
    test_sheet.cell(row=write_row + 5, column=write_col + 2).value = "Significance"
    test_sheet.cell(row=write_row + 5, column=write_col + 2).font = Font(bold=True)
    test_sheet.cell(row=write_row + 5, column=write_col + 2).alignment = Alignment(horizontal='right')
    # Write Dunn's pairwise comparisons inline
    list_row = write_row + 6
    dunn_df = statistics["dunn_matrix"]
    for i, group1 in enumerate(dunn_df.index):
        for j, group2 in enumerate(dunn_df.columns):
            if j <= i:
                continue  # Skip lower triangle and diagonal
            dunn_p_val = dunn_df.loc[group1, group2]
            comparison = f"{group1} vs. {group2}"
            test_sheet.cell(row=list_row, column=write_col).value = comparison
            test_sheet.cell(row=list_row, column=write_col).alignment = Alignment(horizontal='left')
            test_sheet.cell(row=list_row, column=write_col + 1).value = dunn_p_val
            test_sheet.cell(row=list_row, column=write_col + 1).alignment = Alignment(horizontal='right')
            test_sheet.cell(row=list_row, column=write_col + 1).number_format = "0.0##"
            symbol = get_significance_symbol(dunn_p_val)
            test_sheet.cell(row=list_row, column=write_col + 2).value = symbol
            test_sheet.cell(row=list_row, column=write_col + 2).alignment = Alignment(horizontal='right')
            list_row += 1
    autosize_cols(test_sheet)


def run_kruskal_wallis(data_file, statistical_test, test_data):
    """Runs Kruskal-Wallis test coupled with Dunn's test with bonferroni correction and writes the resutls into stats sheet."""
    # Perform Kruskal-Wallis test
    h_stat, kw_p_value = kruskal(*test_data.values())
    significant = kw_p_value < ALPHA
    # Prepare data for Dunn's test
    long_data = []
    for group, values in test_data.items():
        for val in values:
            long_data.append((group, val))
    # Perform Dunn's test if Kruskal-Wallis is significant
    dunn_matrix = None
    data_frame = panda.DataFrame(long_data, columns=["Group", "P-value"])
    dunn_matrix = posthoc.posthoc_dunn(data_frame, val_col='P-value', group_col='Group', p_adjust='bonferroni')
    statistics = {
        "dunn_matrix": dunn_matrix,
        "p_value": kw_p_value,
        "significant": significant
    }
    print_progress(f"{statistical_test} test completed.")
    # Create a new results sheet
    test_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_kw_TEST")
    # Write the Kruskal-Wallis results onto the stats sheet
    kw_to_data_file(test_sheet, statistics)
    print_progress(f'Statistical data written into "{test_sheet.title}".')


def mw_to_data_file(test_sheet, statistics):
    """Function to handle mann-whitney writing into stats sheet."""
    write_row, write_col = 1, 1
    # Write column headers
    test_sheet.cell(row=write_row, column=write_col).value = "MANN-WHITNEY RESULTS"
    test_sheet.cell(row=write_row, column=write_col).font = Font(bold=True)
    test_sheet.cell(row=write_row+1, column=write_col).value = "Comparison"
    test_sheet.cell(row=write_row+1, column=write_col).font = Font(bold=True)
    test_sheet.cell(row=write_row+1, column=write_col).alignment = Alignment(horizontal='left')
    test_sheet.cell(row=write_row+1, column=write_col + 1).value = "P-value"
    test_sheet.cell(row=write_row+1, column=write_col + 1).font = Font(bold=True)
    test_sheet.cell(row=write_row+1, column=write_col + 1).alignment = Alignment(horizontal='right')
    test_sheet.cell(row=write_row+1, column=write_col + 2).value = "Significance"
    test_sheet.cell(row=write_row+1, column=write_col + 2).font = Font(bold=True)
    test_sheet.cell(row=write_row+1, column=write_col + 2).alignment = Alignment(horizontal='right')
    write_row += 2
    # Write the results of each group into the results sheet
    for statistic in statistics:
        test_sheet.cell(row=write_row, column=write_col).value = f'{statistic["g1"]} vs {statistic["g2"]}'
        test_sheet.cell(row=write_row, column=write_col).alignment = Alignment(horizontal='left')
        p_val = statistic["p_value"]
        test_sheet.cell(row=write_row, column=write_col + 1).value = p_val
        test_sheet.cell(row=write_row, column=write_col + 1).alignment = Alignment(horizontal='right')
        test_sheet.cell(row=write_row, column=write_col + 1).number_format = "0.0##"
        symbol = get_significance_symbol(p_val)
        test_sheet.cell(row=write_row, column=write_col + 2).value = symbol
        test_sheet.cell(row=write_row, column=write_col + 2).alignment = Alignment(horizontal='right')
        write_row += 1
    autosize_cols(test_sheet)


def run_mann_whitney(data_file, statistical_test, test_data):
    """Runs Mann-Whitney test with bonferroni correction and writes the resutls into stats sheet."""
    # Initialize the statistics list
    statistics = []
    # Specify group pairs to compare
    group_pairs = list(combinations(test_data.keys(), 2))
    m = len(group_pairs)
    # Calculate the bonferroni alpha
    bonferroni_alpha = ALPHA / m
    # Perform Mann-Whitney U test for specified group comparisons with Bonferroni correction
    for g1, g2 in group_pairs:
        u_stat, p_raw = mannwhitneyu(test_data[g1], test_data[g2])
        p_corrected = min(p_raw * m, 1.0)
        significant = p_raw < bonferroni_alpha
        # Create and add the group vs group statistic
        statistic = {
            "g1": g1,
            "g2": g2,
            "p_value": p_corrected,
            "significant": significant
        }
        statistics.append(statistic)
    print_progress(f"{statistical_test} test completed.")
    # Create a new results sheet
    test_sheet = create_or_clear_sheet(data_file, f"{ENVIRONMENT}_mw_TEST")
    # Fill the sheet with test statistics
    mw_to_data_file(test_sheet, statistics)
    print_progress(f'Statistical data written into "{test_sheet.title}".')


def run_statistical_test(data_file, statistical_test, test_data):
    """Function to perform the selected statistical test and write its results a the test sheet of the data file."""
    # Run statistical test
    if statistical_test == KRUSKAL_WALLIS:
        run_kruskal_wallis(data_file, statistical_test, test_data)
    elif statistical_test == MANN_WHITNEY:
        run_mann_whitney(data_file, statistical_test, test_data)
    else:
        exit_by_error("Unknown statistical test selected.")


def save_file(file, file_name):
    """Save the Excel file to the specified path.

    Parameters:
    file (Workbook): The Excel workbook to be saved.
    file_name (str): The name of the file to save.
    """
    try:
        # Save data file to the 'data_files' directory
        file.save(f"./data_files/{file_name}")
    except PermissionError:
        # Handle permission error if the data file is open in another program
        exit_by_error(f'Unable to save data file "{file_name}". It might be opened in an Excel editor.')




########## ACTUAL SCRIPT ##########


def run_script():
    print_action("Evaluating environment")
    print_task(f'Checking script environment structure.')
    evaluate_script_environment()
    print_success("Script environment structure accepted.")
    # NECESSARY DIRECTORIES HAVE BEEN LOCATED #

    print_action("Selecting data file")
    data_file_name = select_excel_file(DATA_FILES_DIRECTORY)
    print_task(f'Loading data file "{data_file_name}".')
    # DATA FILE HAS BEEN SELECTED #
    data_file, data_file_name = load_excel_file(DATA_FILES_DIRECTORY, data_file_name)
    print_success(f'Data file "{data_file_name}" loaded.')
    # DATA FILE HAS BEEN LOADED #

    print_action("Selecting data sheet")
    data_sheet_name = select_datasheet(data_file)
    print_task(f'Loading data sheet "{data_sheet_name}".')
    # DATA SHEET HAS BEEN SELECTED #
    data_sheet = load_datasheet(data_file, data_sheet_name)
    print_success(f'Data sheet "{data_sheet_name}" loaded.')
    # DATA SHEET HAS BEEN LOADED #

    print_action("Selecting statistical test")
    statistical_test = select_statistical_test()
    print_success(f"Statistical test selected.")
    # ANALYSIS HAS BEEN DEFINED #

    print_action("Extracting test data")
    print_task(f'Extracting the data for {statistical_test} test from data sheet "{data_sheet.title}".')
    test_data = extract_test_data(data_sheet)
    print_success(f"Test data extracted.")
    # ANALYSIS RUN COMPLETED #

    print_action("Running statistical test")
    print_task(f'Running {statistical_test} and creating a test sheet in data file "{data_file_name}".')
    run_statistical_test(data_file, statistical_test, test_data)
    print_success("Statistical test completed.")
    # STATISTICAL TEST COMPLETED #

    print_action("Saving data file")
    print_task(f'Saving data file "{data_file_name}".')
    save_file(data_file, data_file_name)
    print_success(f'Data file "{data_file_name}" saved.')
    # DATA FILE HAS BEEN SAVED #




########## MAIN ENTRY POINT ##########


if __name__ == "__main__":
    handle_arguments(args)
    try:
        print_status("Started")
        run_script()
        print_status("Completed")
    except KeyboardInterrupt:
        exit_by_interruption()
